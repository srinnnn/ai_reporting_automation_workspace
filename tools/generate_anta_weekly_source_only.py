from __future__ import annotations

import logging
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SOURCE_ROOT = (
    Path.cwd()
    / "ai_report_config_materials"
    / "01_data_processing"
    / "01-3_weekly_report"
    / "anta_weekly_report"
    / "01_raw_data"
    / "2026_07_week2"
)
OUTPUT_ROOT = (
    Path.cwd()
    / "ai_report_config_materials"
    / "01_data_processing"
    / "01-3_weekly_report"
    / "anta_weekly_report"
    / "02_manual_deliverables"
    / "anta_weekly_materials"
    / "2026_07_weekly_reports"
    / "week2"
)
MAIN_OUTPUT_NAME = "2026-07-06_2026-07-12_anta_kids_weekly_analysis_ai_source.xlsx"
SELECTION_OUTPUT_NAME = "2026-07-06_2026-07-12_anta_kids_weekly_selection_ai_source.xlsx"


@dataclass(frozen=True)
class ProductMetric:
    platform: str
    product_name: str
    sku_code: str
    style_code: str
    quantity: Decimal
    amount: Decimal
    major_category: str
    category: str

    def __post_init__(self) -> None:
        for field_name, value in (
            ("platform", self.platform),
            ("product_name", self.product_name),
            ("style_code", self.style_code),
        ):
            if not isinstance(value, str) or not value.strip():
                raise ValueError(f"{field_name} must be non-empty str")
        if not isinstance(self.quantity, Decimal) or not isinstance(self.amount, Decimal):
            raise TypeError("quantity and amount must be Decimal")


@dataclass(frozen=True)
class WeeklySourceData:
    meituan_products: tuple[ProductMetric, ...]
    jd_products: tuple[ProductMetric, ...]
    official_sales_by_style: dict[str, Decimal]
    short_title_by_style: dict[str, str]
    source_files: tuple[str, ...]

    def __post_init__(self) -> None:
        if not self.meituan_products:
            raise ValueError("meituan_products must not be empty")
        if not self.jd_products:
            raise ValueError("jd_products must not be empty")
        if not self.source_files:
            raise ValueError("source_files must not be empty")


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    source_data = load_weekly_source_data(SOURCE_ROOT)
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    main_path = OUTPUT_ROOT / MAIN_OUTPUT_NAME
    selection_path = OUTPUT_ROOT / SELECTION_OUTPUT_NAME
    build_main_workbook(source_data, main_path)
    build_selection_workbook(source_data, selection_path)
    assert main_path.exists()
    assert selection_path.exists()
    logging.info("generated source-only weekly reports: %s | %s", main_path, selection_path)


def load_weekly_source_data(source_root: Path) -> WeeklySourceData:
    if not isinstance(source_root, Path):
        raise TypeError("source_root must be pathlib.Path")
    if not source_root.exists() or not source_root.is_dir():
        raise ValueError(f"source_root does not exist: {source_root}")
    meituan_path = _single_matching_file(source_root, "meituan_product_data")
    jd_path = _single_matching_file(source_root, "jd_product_analysis")
    catalog_path = _single_matching_file(source_root, "product_export")
    meituan_products = tuple(_read_platform_products(meituan_path, "美团", "美团"))
    jd_products = tuple(_read_platform_products(jd_path, "京东", "京东"))
    style_codes = {item.style_code for item in meituan_products + jd_products}
    official_sales = _read_official_sales(catalog_path)
    short_titles = _read_short_titles(catalog_path, style_codes)
    result = WeeklySourceData(
        meituan_products=meituan_products,
        jd_products=jd_products,
        official_sales_by_style=official_sales,
        short_title_by_style=short_titles,
        source_files=(meituan_path.name, jd_path.name, catalog_path.name),
    )
    assert len(result.source_files) == 3
    return result


def build_main_workbook(source_data: WeeklySourceData, output_path: Path) -> None:
    _require_source_data(source_data)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_source_sheet(workbook.create_sheet("数据来源"), source_data)
    _write_metric_sheet(workbook.create_sheet("美团"), "美团", source_data.meituan_products)
    _write_metric_sheet(workbook.create_sheet("京东"), "京东", source_data.jd_products)
    _write_product_dimension_sheet(workbook.create_sheet("商品维度数据"), source_data)
    _write_selection_sheet(workbook.create_sheet("本周pyq选品情况"), source_data)
    _write_missing_sheet(workbook.create_sheet("上周pyq销售情况"), "缺少上周已选品清单或上周销量源数据，无法由7月WEEK2源数据产出。")
    _write_missing_sheet(workbook.create_sheet("7月营销节奏表"), "缺少7月营销日历、活动节奏或目标拆解源数据，无法由7月WEEK2源数据产出。")
    _write_missing_sheet(workbook.create_sheet("上周专题海报推品销售情况"), "缺少上周专题海报推品清单和对应销售源数据，无法由7月WEEK2源数据产出。")
    _write_category_sheet(workbook.create_sheet("美团周商品销售数据"), source_data.meituan_products)
    _write_category_sheet(workbook.create_sheet("京东周商品销售数据"), source_data.jd_products)
    _save_workbook(workbook, output_path)


def build_selection_workbook(source_data: WeeklySourceData, output_path: Path) -> None:
    _require_source_data(source_data)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_selection_sheet(workbook.create_sheet("本周pyq选品"), source_data)
    _write_missing_sheet(workbook.create_sheet("上周pyq销售情况"), "缺少上周已选品清单或上周销量源数据，无法由7月WEEK2源数据产出。")
    _write_source_sheet(workbook.create_sheet("数据来源"), source_data)
    _save_workbook(workbook, output_path)


def _write_source_sheet(sheet: Worksheet, source_data: WeeklySourceData) -> None:
    rows = [
        ["说明", "本文件只使用7月WEEK2源数据自动生成；参考成品仅用于理解表名和结构，未复制参考文件中的图片、文案、素材或人工结论。"],
        ["周区间", "2026-07-06 至 2026-07-12"],
        ["源文件", "；".join(source_data.source_files)],
        ["不可产出项", "流量数据、上周环比、营销节奏、专题海报、图片素材：本次源数据未提供，已在对应工作表留空。"],
    ]
    _write_rows(sheet, rows)
    _style_basic_sheet(sheet, title_rows=1)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 120


def _write_metric_sheet(sheet: Worksheet, platform: str, products: tuple[ProductMetric, ...]) -> None:
    if not isinstance(platform, str) or not platform.strip():
        raise ValueError("platform must not be empty")
    if not products:
        raise ValueError("products must not be empty")
    amount = sum((item.amount for item in products), Decimal("0"))
    quantity = sum((item.quantity for item in products), Decimal("0"))
    style_count = len({item.style_code for item in products})
    average_amount = amount / Decimal(style_count) if style_count else Decimal("0")
    rows = [
        ["一、周报数据核心概览", "", "", "", ""],
        ["平台", platform, "本周7.6-7.12", "上周6.29-7.5", "周环比"],
        ["商品数据", "商品实付销售额", amount, "", ""],
        ["", "商品销售数量", quantity, "", ""],
        ["", "有销售款号数", style_count, "", ""],
        ["", "单款平均销售额", average_amount, "", ""],
        ["流量数据", "曝光人数", "", "", ""],
        ["", "入店人数", "", "", ""],
        ["", "下单人数", "", "", ""],
        ["", "入店率", "", "", ""],
        ["", "下单率", "", "", ""],
        ["", "新客下单率/占比", "", "", ""],
        ["说明", "流量数据、上周数据和周环比缺少源文件，按要求留空。", "", "", ""],
    ]
    _write_rows(sheet, rows)
    _style_basic_sheet(sheet, title_rows=1)
    _apply_numeric_format(sheet, "C3:E12")
    sheet.freeze_panes = "A3"


def _write_product_dimension_sheet(sheet: Worksheet, source_data: WeeklySourceData) -> None:
    rows = [[
        "平台",
        "款号",
        "商品名称",
        "短标题",
        "大类",
        "品类",
        "本周销量",
        "本周销售额",
        "官网近30天销量",
    ]]
    for item in sorted(source_data.meituan_products + source_data.jd_products, key=lambda value: (value.platform, -value.amount)):
        rows.append([
            item.platform,
            item.style_code,
            item.product_name,
            source_data.short_title_by_style.get(item.style_code, ""),
            item.major_category,
            item.category,
            item.quantity,
            item.amount,
            source_data.official_sales_by_style.get(item.style_code, Decimal("0")),
        ])
    _write_rows(sheet, rows)
    _style_table_sheet(sheet)
    _apply_numeric_format(sheet, f"G2:I{len(rows)}")


def _write_selection_sheet(sheet: Worksheet, source_data: WeeklySourceData) -> None:
    rows = [[
        "平台",
        "款号",
        "短标题",
        "大类",
        "品类",
        "官网近30天销量",
        "美团近7天销量",
        "京东近7天销量",
        "美团近7天销售额",
        "京东近7天销售额",
        "推荐理由",
        "图片",
    ]]
    for platform, item in _selection_candidates(source_data):
        official_sales = source_data.official_sales_by_style.get(item.style_code, Decimal("0"))
        meituan_match = _find_product(source_data.meituan_products, item.style_code)
        jd_match = _find_product(source_data.jd_products, item.style_code)
        rows.append([
            platform,
            item.style_code,
            source_data.short_title_by_style.get(item.style_code, item.product_name[:30]),
            item.major_category,
            item.category,
            official_sales,
            meituan_match.quantity if meituan_match else "",
            jd_match.quantity if jd_match else "",
            meituan_match.amount if meituan_match else "",
            jd_match.amount if jd_match else "",
            _selection_reason(item, official_sales),
            "",
        ])
    _write_rows(sheet, rows)
    _style_table_sheet(sheet)
    _apply_numeric_format(sheet, f"F2:J{len(rows)}")
    sheet.column_dimensions["K"].width = 42
    sheet.column_dimensions["L"].width = 12


def _write_category_sheet(sheet: Worksheet, products: tuple[ProductMetric, ...]) -> None:
    if not products:
        raise ValueError("products must not be empty")
    totals = _category_totals(products)
    total_amount = sum((values["amount"] for values in totals.values()), Decimal("0"))
    rows = [["大类", "品类", "销售额", "销量", "销售额占比"]]
    for (major, category), values in sorted(totals.items(), key=lambda item: item[1]["amount"], reverse=True):
        share = values["amount"] / total_amount if total_amount else Decimal("0")
        rows.append([major, category, values["amount"], values["quantity"], share])
    _write_rows(sheet, rows)
    _style_table_sheet(sheet)
    _apply_numeric_format(sheet, f"C2:D{len(rows)}")
    for row_index in range(2, len(rows) + 1):
        sheet.cell(row_index, 5).number_format = "0.0%"


def _write_missing_sheet(sheet: Worksheet, reason: str) -> None:
    if not isinstance(reason, str) or not reason.strip():
        raise ValueError("reason must not be empty")
    rows = [
        ["状态", "留空"],
        ["原因", reason],
        ["需要补充的源数据", "对应后台导出的原始表或业务维护清单"],
    ]
    _write_rows(sheet, rows)
    _style_basic_sheet(sheet, title_rows=0)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 90


def _selection_candidates(source_data: WeeklySourceData) -> list[tuple[str, ProductMetric]]:
    _require_source_data(source_data)
    candidates: list[tuple[str, ProductMetric]] = []
    candidates.extend(("美团", item) for item in sorted(source_data.meituan_products, key=lambda product: product.amount, reverse=True)[:10])
    candidates.extend(("京东", item) for item in sorted(source_data.jd_products, key=lambda product: product.amount, reverse=True)[:10])
    result: list[tuple[str, ProductMetric]] = []
    seen: set[tuple[str, str]] = set()
    for platform, item in candidates:
        key = (platform, item.style_code)
        if key in seen:
            continue
        seen.add(key)
        result.append((platform, item))
    assert result
    return result


def _selection_reason(item: ProductMetric, official_sales: Decimal) -> str:
    if not isinstance(item, ProductMetric):
        raise TypeError("item must be ProductMetric")
    if not isinstance(official_sales, Decimal):
        raise TypeError("official_sales must be Decimal")
    parts = [f"{item.platform}本周销售额{_format_decimal(item.amount)}，销量{_format_decimal(item.quantity)}"]
    if official_sales > 0:
        parts.append(f"官网近30天销量{_format_decimal(official_sales)}")
    return "；".join(parts)


def _find_product(products: tuple[ProductMetric, ...], style_code: str) -> ProductMetric | None:
    if not isinstance(style_code, str) or not style_code.strip():
        raise ValueError("style_code must not be empty")
    return next((item for item in products if item.style_code == style_code), None)


def _category_totals(products: tuple[ProductMetric, ...]) -> dict[tuple[str, str], dict[str, Decimal]]:
    if not products:
        raise ValueError("products must not be empty")
    totals: dict[tuple[str, str], dict[str, Decimal]] = defaultdict(lambda: {"amount": Decimal("0"), "quantity": Decimal("0")})
    for item in products:
        key = (item.major_category or "未识别", item.category or "未识别")
        totals[key]["amount"] += item.amount
        totals[key]["quantity"] += item.quantity
    result = dict(totals)
    assert result
    return result


def _read_platform_products(path: Path, sheet_name: str, platform: str) -> list[ProductMetric]:
    if not isinstance(path, Path):
        raise TypeError("path must be pathlib.Path")
    if not path.exists() or not path.is_file():
        raise ValueError(f"path does not exist: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = _clean_row(next(rows))
        indexes = _indexes(headers)
        required = ("商品名称", "款号", "透视销量", "透视金额", "大类", "品类")
        _require_headers(indexes, required, path.name, sheet_name)
        products: list[ProductMetric] = []
        for row in rows:
            values = _clean_row(row)
            style_code = _cell(values, indexes["款号"])
            if not style_code:
                continue
            products.append(
                ProductMetric(
                    platform=platform,
                    product_name=_cell(values, indexes["商品名称"]) or "未命名商品",
                    sku_code=_cell(values, indexes.get("商品SKU码", indexes.get("SKUID", -1))),
                    style_code=style_code,
                    quantity=_decimal(_cell(values, indexes["透视销量"]), "透视销量"),
                    amount=_decimal(_cell(values, indexes["透视金额"]), "透视金额"),
                    major_category=_cell(values, indexes["大类"]) or "未识别",
                    category=_cell(values, indexes["品类"]) or "未识别",
                )
            )
    finally:
        workbook.close()
    if not products:
        raise ValueError(f"no products read from {path}")
    logging.info("read %s %s products from %s", len(products), platform, path.name)
    return products


def _read_official_sales(path: Path) -> dict[str, Decimal]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)
        next(rows, None)
        headers = _clean_row(next(rows))
        indexes = _indexes(headers)
        _require_headers(indexes, ("款号", "求和项:近30天销量（实时）（款）"), path.name, "Sheet1")
        sales: dict[str, Decimal] = {}
        for row in rows:
            values = _clean_row(row)
            style_code = _cell(values, indexes["款号"])
            if not style_code:
                continue
            sales[style_code] = _decimal(_cell(values, indexes["求和项:近30天销量（实时）（款）"]), "官网近30天销量")
    finally:
        workbook.close()
    assert isinstance(sales, dict)
    logging.info("read official sales for %s styles", len(sales))
    return sales


def _read_short_titles(path: Path, style_codes: set[str]) -> dict[str, str]:
    if not isinstance(style_codes, set):
        raise TypeError("style_codes must be set[str]")
    if not style_codes:
        raise ValueError("style_codes must not be empty")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = next(ws for ws in workbook.worksheets if ws.title.startswith("商品下载-"))
        rows = sheet.iter_rows(values_only=True)
        headers = _clean_row(next(rows))
        indexes = _indexes(headers)
        _require_headers(indexes, ("款号", "短标题", "标题"), path.name, sheet.title)
        titles: dict[str, str] = {}
        for row in rows:
            values = _clean_row(row)
            style_code = _cell(values, indexes["款号"])
            if style_code not in style_codes or style_code in titles:
                continue
            titles[style_code] = _cell(values, indexes["短标题"]) or _cell(values, indexes["标题"])
            if len(titles) == len(style_codes):
                break
    finally:
        workbook.close()
    assert isinstance(titles, dict)
    logging.info("read short titles for %s styles", len(titles))
    return titles


def _single_matching_file(root: Path, keyword: str) -> Path:
    if not isinstance(root, Path):
        raise TypeError("root must be pathlib.Path")
    if not isinstance(keyword, str) or not keyword.strip():
        raise ValueError("keyword must not be empty")
    matches = [path for path in root.glob("*.xlsx") if keyword in path.name and not path.name.startswith("~$")]
    if len(matches) != 1:
        raise ValueError(f"expected one {keyword} file under {root}, got {len(matches)}")
    result = matches[0]
    assert result.exists()
    return result


def _clean_row(row: Iterable[object]) -> list[str]:
    result = ["" if value is None else str(value).replace("\t", "").strip() for value in row]
    assert isinstance(result, list)
    return result


def _indexes(headers: list[str]) -> dict[str, int]:
    if not isinstance(headers, list) or not headers:
        raise ValueError("headers must be non-empty list")
    result = {header: index for index, header in enumerate(headers) if header}
    assert isinstance(result, dict)
    return result


def _cell(values: list[str], index: int) -> str:
    if index < 0 or index >= len(values):
        return ""
    return values[index]


def _require_headers(indexes: dict[str, int], required: tuple[str, ...], file_name: str, sheet_name: str) -> None:
    missing = [header for header in required if header not in indexes]
    if missing:
        raise ValueError(f"{file_name}/{sheet_name} 缺少字段：{'、'.join(missing)}")


def _decimal(value: object, field_name: str) -> Decimal:
    if not isinstance(field_name, str) or not field_name.strip():
        raise ValueError("field_name must not be empty")
    text = "" if value is None else str(value).strip().replace(",", "")
    if text in ("", "-", "--"):
        return Decimal("0")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be numeric, got {value}") from exc


def _format_decimal(value: Decimal) -> str:
    if not isinstance(value, Decimal):
        raise TypeError("value must be Decimal")
    if value == value.to_integral_value():
        return str(value.to_integral_value())
    return str(value.quantize(Decimal("0.01")))


def _write_rows(sheet: Worksheet, rows: list[list[object]]) -> None:
    if not isinstance(sheet, Worksheet):
        raise TypeError("sheet must be Worksheet")
    if not isinstance(rows, list) or not rows:
        raise ValueError("rows must be non-empty list")
    for row in rows:
        sheet.append([_excel_value(value) for value in row])
    assert sheet.max_row == len(rows)


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    return value


def _style_basic_sheet(sheet: Worksheet, title_rows: int) -> None:
    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="EAF1FF")
    thin = Side(style="thin", color="D8DEE8")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    if title_rows >= 1:
        for cell in sheet[1]:
            cell.fill = title_fill
            cell.font = Font(bold=True, color="FFFFFF")
    if sheet.max_row >= 2:
        for cell in sheet[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="1F2933")
    _autofit(sheet)


def _style_table_sheet(sheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="0F766E")
    thin = Side(style="thin", color="D8DEE8")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autofit(sheet)


def _apply_numeric_format(sheet: Worksheet, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"


def _autofit(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells), default=0)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 36)
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 22


def _save_workbook(workbook: Workbook, output_path: Path) -> None:
    if not isinstance(workbook, Workbook):
        raise TypeError("workbook must be Workbook")
    if not isinstance(output_path, Path):
        raise TypeError("output_path must be pathlib.Path")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    if not output_path.exists() or output_path.stat().st_size <= 0:
        raise AssertionError(f"failed to save workbook: {output_path}")


def _require_source_data(source_data: WeeklySourceData) -> None:
    if not isinstance(source_data, WeeklySourceData):
        raise TypeError("source_data must be WeeklySourceData")


if __name__ == "__main__":
    main()
