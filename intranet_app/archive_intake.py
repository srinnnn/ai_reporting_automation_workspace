from __future__ import annotations

import csv
import logging
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


INDEX_HEADERS = (
    "归档时间",
    "优先级",
    "项目",
    "品牌",
    "平台",
    "业务方",
    "文件名称",
    "资料类型",
    "日期范围",
    "用途说明",
    "是否脱敏",
    "口径负责人",
    "原始提交路径",
    "归档路径",
    "判定置信度",
    "判定理由",
    "文件大小KB",
    "文件类型",
)

DICTIONARY_HEADERS = (
    "更新时间",
    "优先级",
    "项目",
    "资料类型",
    "字段名",
    "出现次数",
    "最近来源文件",
    "最近归档路径",
    "示例值",
    "备注",
)

SUPPORTED_FILE_SUFFIXES = (".csv", ".xlsx", ".xls", ".html", ".htm", ".docx", ".pptx", ".pdf", ".png", ".jpg", ".jpeg")
CATALOG_EXCLUDED_DIR_NAMES = ("00_intake", "00_index_dictionary")
CATALOG_EXCLUDED_FILE_SUFFIXES = ("_AI生成.xlsx",)


@dataclass(frozen=True)
class ArchiveIntakeConfig:
    package_root: Path

    def __post_init__(self) -> None:
        if not isinstance(self.package_root, Path):
            raise TypeError("package_root must be pathlib.Path")
        if not self.package_root.exists():
            raise ValueError(f"资料包目录不存在：{self.package_root}")

    @property
    def intake_pending_dir(self) -> Path:
        return self.package_root / "00_intake" / "01_pending"

    @property
    def intake_done_dir(self) -> Path:
        return self.package_root / "00_intake" / "02_processed"

    @property
    def intake_unresolved_dir(self) -> Path:
        return self.package_root / "00_intake" / "03_unresolved"

    @property
    def index_path(self) -> Path:
        return self.package_root / "00_index_dictionary" / "archive_index.csv"

    @property
    def dictionary_path(self) -> Path:
        return self.package_root / "00_index_dictionary" / "data_dictionary.csv"


@dataclass(frozen=True)
class SourceProfile:
    file_path: Path
    file_name: str
    suffix: str
    headers: tuple[str, ...]
    sample_values: dict[str, str]
    text: str

    def __post_init__(self) -> None:
        if not isinstance(self.file_path, Path):
            raise TypeError("file_path must be pathlib.Path")
        if not self.file_name.strip():
            raise ValueError("file_name must not be empty")
        if not isinstance(self.headers, tuple):
            raise TypeError("headers must be tuple[str, ...]")
        if not isinstance(self.sample_values, dict):
            raise TypeError("sample_values must be dict[str, str]")
        if not self.text.strip():
            raise ValueError("text must not be empty")


@dataclass(frozen=True)
class FileClassification:
    priority: str
    project: str
    brand: str
    platform: str
    business_party: str
    material_type: str
    destination_dir: Path
    confidence: int
    reason: str

    def __post_init__(self) -> None:
        for field_name, value in (
            ("priority", self.priority),
            ("project", self.project),
            ("brand", self.brand),
            ("platform", self.platform),
            ("business_party", self.business_party),
            ("material_type", self.material_type),
            ("reason", self.reason),
        ):
            if not isinstance(value, str) or not value.strip():
                raise ValueError(f"{field_name} must be non-empty str")
        if not isinstance(self.destination_dir, Path):
            raise TypeError("destination_dir must be pathlib.Path")
        if self.confidence < 0 or self.confidence > 100:
            raise ValueError("confidence must be between 0 and 100")


@dataclass(frozen=True)
class ArchiveFileOutcome:
    file_name: str
    status: str
    priority: str
    project: str
    material_type: str
    archive_path: str
    reason: str

    def __post_init__(self) -> None:
        for field_name, value in (
            ("file_name", self.file_name),
            ("status", self.status),
            ("priority", self.priority),
            ("project", self.project),
            ("material_type", self.material_type),
            ("archive_path", self.archive_path),
            ("reason", self.reason),
        ):
            if not isinstance(value, str):
                raise TypeError(f"{field_name} must be str")


@dataclass(frozen=True)
class ArchiveIntakeResult:
    processed_count: int
    unresolved_count: int
    skipped_count: int
    index_path: Path
    dictionary_path: Path
    outcomes: tuple[ArchiveFileOutcome, ...]

    def __post_init__(self) -> None:
        for value, name in (
            (self.processed_count, "processed_count"),
            (self.unresolved_count, "unresolved_count"),
            (self.skipped_count, "skipped_count"),
        ):
            if not isinstance(value, int) or value < 0:
                raise ValueError(f"{name} must be non-negative int")
        if not isinstance(self.index_path, Path) or not isinstance(self.dictionary_path, Path):
            raise TypeError("index_path and dictionary_path must be pathlib.Path")
        if not isinstance(self.outcomes, tuple):
            raise TypeError("outcomes must be tuple[ArchiveFileOutcome, ...]")


def ensure_intake_workspace(config: ArchiveIntakeConfig) -> None:
    if not isinstance(config, ArchiveIntakeConfig):
        raise TypeError("config must be ArchiveIntakeConfig")
    for path in (config.intake_pending_dir, config.intake_done_dir, config.intake_unresolved_dir, config.index_path.parent):
        path.mkdir(parents=True, exist_ok=True)
        if not path.exists():
            raise AssertionError(f"failed to create {path}")
    _ensure_csv(config.index_path, INDEX_HEADERS)
    _ensure_csv(config.dictionary_path, DICTIONARY_HEADERS)
    logging.info("archive intake workspace ensured")


def run_archive_intake(config: ArchiveIntakeConfig) -> ArchiveIntakeResult:
    if not isinstance(config, ArchiveIntakeConfig):
        raise TypeError("config must be ArchiveIntakeConfig")
    ensure_intake_workspace(config)
    files = _pending_files(config.intake_pending_dir)
    outcomes: list[ArchiveFileOutcome] = []
    index_rows: list[dict[str, str]] = []
    dictionary_updates: list[tuple[FileClassification, Path, SourceProfile]] = []
    today_dir = datetime.now().strftime("%Y%m%d")
    done_dir = config.intake_done_dir / today_dir
    unresolved_dir = config.intake_unresolved_dir / today_dir
    processed_count = 0
    unresolved_count = 0
    skipped_count = 0

    for file_path in files:
        if file_path.suffix.lower() not in SUPPORTED_FILE_SUFFIXES:
            skipped_count += 1
            outcomes.append(_outcome(file_path.name, "跳过", "", "", "", "", "暂不支持该文件类型"))
            continue
        profile = build_source_profile(file_path)
        classification = classify_source(config.package_root, profile)
        if classification is None:
            unresolved_path = _move_with_unique_name(file_path, unresolved_dir)
            unresolved_count += 1
            outcomes.append(_outcome(file_path.name, "无法判定", "", "", "", str(unresolved_path), "文件名和字段未命中已知规则"))
            continue
        classification.destination_dir.mkdir(parents=True, exist_ok=True)
        archive_path = _copy_with_unique_name(file_path, classification.destination_dir)
        index_rows.append(_index_row(profile, classification, archive_path))
        dictionary_updates.append((classification, archive_path, profile))
        _move_with_unique_name(file_path, done_dir)
        processed_count += 1
        outcomes.append(
            _outcome(
                file_path.name,
                "已归档",
                classification.priority,
                classification.project,
                classification.material_type,
                str(archive_path),
                classification.reason,
            )
        )

    if index_rows:
        _append_csv_rows(config.index_path, INDEX_HEADERS, index_rows)
    if dictionary_updates:
        _update_dictionary(config.dictionary_path, dictionary_updates)
    result = ArchiveIntakeResult(
        processed_count=processed_count,
        unresolved_count=unresolved_count,
        skipped_count=skipped_count,
        index_path=config.index_path,
        dictionary_path=config.dictionary_path,
        outcomes=tuple(outcomes),
    )
    logging.info("archive intake finished: processed=%s unresolved=%s skipped=%s", processed_count, unresolved_count, skipped_count)
    return result


def rebuild_archive_catalog(config: ArchiveIntakeConfig) -> ArchiveIntakeResult:
    if not isinstance(config, ArchiveIntakeConfig):
        raise TypeError("config must be ArchiveIntakeConfig")
    ensure_intake_workspace(config)
    files = _catalog_files(config.package_root)
    outcomes: list[ArchiveFileOutcome] = []
    index_rows: list[dict[str, str]] = []
    dictionary_updates: list[tuple[FileClassification, Path, SourceProfile]] = []
    processed_count = 0
    unresolved_count = 0
    skipped_count = 0

    for file_path in files:
        if not file_path.exists() or not file_path.is_file():
            skipped_count += 1
            logging.info("catalog file disappeared before profiling: %s", file_path)
            continue
        if file_path.suffix.lower() not in SUPPORTED_FILE_SUFFIXES:
            skipped_count += 1
            continue
        profile = build_source_profile(file_path)
        classification = classify_source(config.package_root, profile)
        if classification is None:
            unresolved_count += 1
            outcomes.append(_outcome(file_path.name, "无法判定", "", "", "", str(file_path), "本地资料路径和字段未命中已知规则"))
            continue
        index_rows.append(_index_row(profile, classification, file_path))
        dictionary_updates.append((classification, file_path, profile))
        processed_count += 1
        outcomes.append(
            _outcome(
                file_path.name,
                "已登记",
                classification.priority,
                classification.project,
                classification.material_type,
                str(file_path),
                classification.reason,
            )
        )

    _write_csv_rows(config.index_path, INDEX_HEADERS, index_rows)
    _write_csv_rows(config.dictionary_path, DICTIONARY_HEADERS, [])
    if dictionary_updates:
        _update_dictionary(config.dictionary_path, dictionary_updates)
    result = ArchiveIntakeResult(
        processed_count=processed_count,
        unresolved_count=unresolved_count,
        skipped_count=skipped_count,
        index_path=config.index_path,
        dictionary_path=config.dictionary_path,
        outcomes=tuple(outcomes),
    )
    logging.info("archive catalog rebuilt: processed=%s unresolved=%s skipped=%s", processed_count, unresolved_count, skipped_count)
    return result


def build_source_profile(file_path: Path) -> SourceProfile:
    if not isinstance(file_path, Path):
        raise TypeError("file_path must be pathlib.Path")
    if not file_path.exists() or not file_path.is_file():
        raise ValueError(f"file_path must be an existing file: {file_path}")
    suffix = file_path.suffix.lower()
    headers, sample_values = _extract_table_profile(file_path, suffix)
    path_text = str(file_path.parent).replace("\\", " ")
    joined = " ".join((path_text, file_path.stem, suffix, " ".join(headers)))
    profile = SourceProfile(file_path=file_path, file_name=file_path.name, suffix=suffix, headers=headers, sample_values=sample_values, text=joined.lower())
    assert profile.file_name == file_path.name
    return profile


def classify_source(package_root: Path, profile: SourceProfile) -> FileClassification | None:
    if not isinstance(package_root, Path):
        raise TypeError("package_root must be pathlib.Path")
    if not isinstance(profile, SourceProfile):
        raise TypeError("profile must be SourceProfile")
    brand = _detect_brand(profile.text)
    platform = _detect_platform(profile.text)
    material_type = _detect_material_type(profile.text)
    date_range = _detect_date_range(profile.file_name)
    _ = date_range

    if _is_short_message(profile):
        brand_name = brand if brand != "未识别" else "博西"
        return FileClassification(
            "P1",
            "短彩信数据处理",
            brand_name,
            platform if platform != "未识别" else "短彩信",
            "待补充",
            "原始导出数据",
            package_root / "01_data_processing" / "01-1_sms_mms_processing" / _brand_dir_name(brand_name) / "01_raw_exports",
            92,
            "命中文件名或字段中的短彩信数据处理规则",
        )
    if "日报" in profile.text:
        brand_name = brand if brand != "未识别" else "标准模板包"
        return FileClassification("P1", "日报", brand_name, platform, "待补充", material_type, package_root / "01_data_processing" / "01-2_daily_report" / _brand_dir_name(brand_name) / _data_subdir(material_type), 86, "命中文件名中的日报规则")
    if "周报" in profile.text or "weekly" in profile.text:
        project = "anta_weekly_report" if brand == "安踏" else "standard_weekly_template_pack"
        return FileClassification("P1", "周报", brand, platform, "待补充", material_type, package_root / "01_data_processing" / "01-3_weekly_report" / project / _data_subdir(material_type), 86, "命中文件名中的周报规则")
    if "月报" in profile.text or "monthly" in profile.text:
        project = "anta_monthly_report" if brand == "安踏" else "standard_material_template"
        return FileClassification("P1", "月报", brand, platform, "待补充", material_type, package_root / "01_data_processing" / "01-4_monthly_report" / project / _data_subdir(material_type), 86, "命中文件名中的月报规则")

    content_destination = _content_destination(package_root, profile.text)
    if content_destination is not None:
        content_type, destination = content_destination
        return FileClassification("P2", "品牌内容资料", brand, platform, "待补充", content_type, destination, 84, "命中品牌内容或 AI 选品资料规则")

    retail_destination = _anta_retail_destination(package_root, profile.text)
    if retail_destination is not None:
        retail_type, destination, confidence, reason = retail_destination
        return FileClassification("P3", "安踏即时零售", "安踏", platform, "待补充", retail_type, destination, confidence, reason)

    page_destination = _page_review_destination(package_root, profile.text)
    if page_destination is not None:
        page_type, destination = page_destination
        return FileClassification("P4", "页面巡检复盘", brand, platform, "待补充", page_type, destination, 82, "命中页面巡检复盘资料规则")
    return None


def _pending_files(pending_dir: Path) -> tuple[Path, ...]:
    if not isinstance(pending_dir, Path):
        raise TypeError("pending_dir must be pathlib.Path")
    files = tuple(path for path in sorted(pending_dir.rglob("*")) if path.is_file() and not path.name.startswith("~$"))
    assert isinstance(files, tuple)
    return files


def _catalog_files(package_root: Path) -> tuple[Path, ...]:
    if not isinstance(package_root, Path):
        raise TypeError("package_root must be pathlib.Path")
    excluded_roots = {package_root / dir_name for dir_name in CATALOG_EXCLUDED_DIR_NAMES}
    files: list[Path] = []
    for path in sorted(package_root.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.name.endswith(CATALOG_EXCLUDED_FILE_SUFFIXES):
            continue
        if any(_is_relative_to(path, excluded_root) for excluded_root in excluded_roots):
            continue
        files.append(path)
    result = tuple(files)
    assert isinstance(result, tuple)
    return result


def _is_relative_to(path: Path, root: Path) -> bool:
    if not isinstance(path, Path) or not isinstance(root, Path):
        raise TypeError("path and root must be pathlib.Path")
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _extract_table_profile(file_path: Path, suffix: str) -> tuple[tuple[str, ...], dict[str, str]]:
    if suffix == ".csv":
        return _extract_csv_profile(file_path)
    if suffix == ".xlsx":
        return _extract_xlsx_profile(file_path)
    if suffix in (".html", ".htm"):
        return _extract_html_profile(file_path)
    return (), {}


def _extract_csv_profile(file_path: Path) -> tuple[tuple[str, ...], dict[str, str]]:
    try:
        headers, first_row = _read_csv_header_and_sample(file_path, "utf-8-sig")
    except UnicodeDecodeError:
        logging.info("csv profile falls back to gb18030: %s", file_path)
        headers, first_row = _read_csv_header_and_sample(file_path, "gb18030")
    sample = {header: str(first_row.get(header, "")).strip() for header in headers} if first_row is not None else {}
    return headers, sample


def _read_csv_header_and_sample(file_path: Path, encoding: str) -> tuple[tuple[str, ...], dict[str, str] | None]:
    if not isinstance(file_path, Path):
        raise TypeError("file_path must be pathlib.Path")
    if not isinstance(encoding, str) or not encoding.strip():
        raise ValueError("encoding must not be empty")
    with file_path.open("r", encoding=encoding, newline="") as handle:
        reader = csv.DictReader(handle)
        headers = tuple(str(name).strip() for name in (reader.fieldnames or ()) if str(name).strip())
        first_row = next(reader, None)
    assert isinstance(headers, tuple)
    return headers, first_row


def _extract_xlsx_profile(file_path: Path) -> tuple[tuple[str, ...], dict[str, str]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(min_row=1, max_row=2, values_only=True)
        header_row = next(rows, ())
        sample_row = next(rows, ())
        headers = tuple("" if value is None else str(value).strip() for value in header_row)
        cleaned_headers = tuple(header for header in headers if header)
        sample = {
            header: "" if index >= len(sample_row) or sample_row[index] is None else str(sample_row[index]).strip()
            for index, header in enumerate(headers)
            if header
        }
        return cleaned_headers, sample
    finally:
        workbook.close()


def _extract_html_profile(file_path: Path) -> tuple[tuple[str, ...], dict[str, str]]:
    text = file_path.read_text(encoding="utf-8", errors="ignore")
    headers = tuple(_clean_html_text(value) for value in re.findall(r"<th[^>]*>(.*?)</th>", text, flags=re.IGNORECASE | re.DOTALL))
    cleaned = tuple(header for header in headers if header)
    return cleaned, {}


def _is_short_message(profile: SourceProfile) -> bool:
    required_headers = {"发送量", "到达量", "点击量", "订单量", "成交金额"}
    return "短彩信" in profile.text or "短信" in profile.text or required_headers.issubset(set(profile.headers))


def _content_destination(package_root: Path, text: str) -> tuple[str, Path] | None:
    base = package_root / "02_brand_content_materials" / "required_materials"
    if "品牌规范" in text or "品牌调性" in text:
        return "品牌规范", base / "01_brand_guidelines"
    if "商品卖点" in text or "卖点" in text:
        return "商品卖点", base / "02_product_selling_points"
    if "历史文案" in text or "优秀文案" in text:
        return "历史优秀文案", base / "03_historical_copy_examples"
    if "禁用词" in text or "敏感词" in text:
        return "禁用词敏感词", base / "04_forbidden_sensitive_words"
    if "活动日历" in text or "营销日历" in text:
        return "活动日历", base / "05_campaign_calendar"
    if "ai选品" in text or "选品" in text:
        return "AI选品资料", package_root / "02_brand_content_materials" / "anta_project_enrichment"
    return None


def _anta_retail_destination(package_root: Path, text: str) -> tuple[str, Path, int, str] | None:
    base = package_root / "03_config_automation_materials" / "03-1_anta_instant_retail"
    if "黑名单" in text:
        return "黑名单筛选", base / "03-1-3_blacklist_filter", 92, "命中黑名单关键词"
    if "素材" in text or "主图" in text or "待核验" in text:
        return "素材筛选", base / "03-1-2_material_filter", 88, "命中素材筛选关键词"
    if "上下架" in text or "上架" in text or "下架" in text or "货盘" in text:
        return "上下架筛选", base / "03-1-1_listing_filter", 88, "命中上下架筛选关键词"
    if "app_spu_code" in text or "商家商品编码" in text or "spu编码" in text or "商品下载" in text:
        return "上下架筛选", base / "03-1-1_listing_filter", 86, "命中美团或京东商品原表字段"
    if "安踏" in text and ("美团" in text or "京东" in text):
        return "上下架筛选", base / "03-1-1_listing_filter", 74, "命中安踏即时零售平台关键词"
    return None


def _page_review_destination(package_root: Path, text: str) -> tuple[str, Path] | None:
    base = package_root / "04_page_audit_review" / "required_materials"
    if "页面链接" in text or "url" in text or "链接清单" in text:
        return "页面链接清单", base / "01_page_url_list"
    if "巡检标准" in text or "巡检" in text:
        return "巡检标准", base / "02_audit_standards"
    if "历史问题" in text or "问题清单" in text:
        return "历史问题清单", base / "03_historical_issue_list"
    if "复盘" in text:
        return "活动复盘样例", base / "04_campaign_review_examples"
    if "竞品" in text or "对账" in text:
        return "竞品与对账资料", base / "05_competitor_reconciliation"
    return None


def _data_subdir(material_type: str) -> str:
    if material_type == "人工成品报表":
        return "02_manual_deliverables"
    if material_type == "指标口径说明":
        return "03_metric_definitions"
    if material_type == "异常案例":
        return "04_exception_cases"
    if material_type == "测试数据_脱敏":
        return "05_desensitized_test_data"
    return "01_raw_data"


def _detect_material_type(text: str) -> str:
    if "人工成品" in text or "成品报表" in text or "结果样例" in text:
        return "人工成品报表"
    if "指标口径" in text or "字段规则" in text or "规则说明" in text:
        return "指标口径说明"
    if "异常" in text or "失败案例" in text:
        return "异常案例"
    if "脱敏" in text or "测试数据" in text:
        return "测试数据_脱敏"
    return "原始数据"


def _detect_brand(text: str) -> str:
    brand_rules = (
        ("博西", ("博西", "bosch")),
        ("安踏", ("安踏", "anta")),
        ("NIKE", ("nike", "耐克")),
        ("TOMMY", ("tommy",)),
        ("NES", ("nes",)),
        ("armani", ("armani", "阿玛尼")),
    )
    for brand, keywords in brand_rules:
        if any(keyword in text for keyword in keywords):
            return brand
    return "未识别"


def _brand_dir_name(brand: str) -> str:
    if not isinstance(brand, str) or not brand.strip():
        raise ValueError("brand must be non-empty str")
    brand_dirs = {
        "博西": "bosch",
        "安踏": "anta",
        "安踏儿童": "anta_kids",
        "标准模板包": "standard_template_pack",
        "标准资料模板": "standard_material_template",
        "未识别": "unknown",
    }
    result = brand_dirs.get(brand, re.sub(r"[^A-Za-z0-9_-]+", "_", brand).strip("_").lower() or "unknown")
    assert result.strip()
    return result


def _detect_platform(text: str) -> str:
    platform_rules = (
        ("美团", ("美团", "meituan")),
        ("京东", ("京东", "jd", "jingdong")),
        ("天猫", ("天猫", "tmall")),
        ("淘宝", ("淘宝", "taobao")),
        ("抖音", ("抖音", "douyin")),
        ("微信", ("微信", "wechat")),
        ("短彩信", ("短彩信", "短信", "彩信", "sms")),
    )
    for platform, keywords in platform_rules:
        if any(keyword in text for keyword in keywords):
            return platform
    return "未识别"


def _detect_date_range(file_name: str) -> str:
    patterns = (
        r"20\d{2}[-_.]?\d{2}[-_.]?\d{2}\s*[-~至_]\s*20\d{2}[-_.]?\d{2}[-_.]?\d{2}",
        r"\d{4}[-_.]?\d{2}[-_.]?\d{2}\s*[-~至_]\s*\d{4}[-_.]?\d{2}[-_.]?\d{2}",
        r"\d{4}[-_.]?\d{2}[-_.]?\d{2}",
        r"\d{4}[-_.]?\d{2}",
        r"\d{4}",
    )
    for pattern in patterns:
        match = re.search(pattern, file_name)
        if match:
            return match.group(0)
    return "待补充"


def _index_row(profile: SourceProfile, classification: FileClassification, archive_path: Path) -> dict[str, str]:
    if not isinstance(profile, SourceProfile):
        raise TypeError("profile must be SourceProfile")
    if not isinstance(classification, FileClassification):
        raise TypeError("classification must be FileClassification")
    return {
        "归档时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "优先级": classification.priority,
        "项目": classification.project,
        "品牌": classification.brand,
        "平台": classification.platform,
        "业务方": classification.business_party,
        "文件名称": profile.file_name,
        "资料类型": classification.material_type,
        "日期范围": _detect_date_range(profile.file_name),
        "用途说明": classification.reason,
        "是否脱敏": "待确认",
        "口径负责人": "待补充",
        "原始提交路径": str(profile.file_path),
        "归档路径": str(archive_path),
        "判定置信度": str(classification.confidence),
        "判定理由": classification.reason,
        "文件大小KB": str(max(1, round(profile.file_path.stat().st_size / 1024))),
        "文件类型": profile.suffix,
    }


def _update_dictionary(dictionary_path: Path, updates: list[tuple[FileClassification, Path, SourceProfile]]) -> None:
    if not isinstance(dictionary_path, Path):
        raise TypeError("dictionary_path must be pathlib.Path")
    existing_rows = _read_csv_rows(dictionary_path)
    row_by_key = {
        (row.get("优先级", ""), row.get("项目", ""), row.get("资料类型", ""), row.get("字段名", "")): dict(row)
        for row in existing_rows
        if row.get("字段名", "").strip()
    }
    for classification, archive_path, profile in updates:
        for header in profile.headers:
            cleaned_header = header.strip()
            if not cleaned_header:
                continue
            key = (classification.priority, classification.project, classification.material_type, cleaned_header)
            row = row_by_key.get(key)
            if row is None:
                row = {header_name: "" for header_name in DICTIONARY_HEADERS}
                row["出现次数"] = "0"
            count = int(row.get("出现次数", "0") or "0") + 1
            row.update(
                {
                    "更新时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "优先级": classification.priority,
                    "项目": classification.project,
                    "资料类型": classification.material_type,
                    "字段名": cleaned_header,
                    "出现次数": str(count),
                    "最近来源文件": profile.file_name,
                    "最近归档路径": str(archive_path),
                    "示例值": profile.sample_values.get(cleaned_header, ""),
                    "备注": row.get("备注", ""),
                }
            )
            row_by_key[key] = row
    _write_csv_rows(dictionary_path, DICTIONARY_HEADERS, list(row_by_key.values()))


def _ensure_csv(path: Path, headers: tuple[str, ...]) -> None:
    if not path.exists():
        _write_csv_rows(path, headers, [])
        return
    if path.stat().st_size == 0:
        _write_csv_rows(path, headers, [])


def _append_csv_rows(path: Path, headers: tuple[str, ...], rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = path.exists() and path.stat().st_size > 0
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        if not file_exists:
            writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _write_csv_rows(path: Path, headers: tuple[str, ...], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})
    if not path.exists():
        raise AssertionError(f"failed to write {path}")


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _copy_with_unique_name(source_path: Path, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination_path = _unique_path(destination_dir / source_path.name)
    shutil.copy2(source_path, destination_path)
    if not destination_path.exists():
        raise AssertionError(f"failed to copy {source_path}")
    return destination_path


def _move_with_unique_name(source_path: Path, destination_dir: Path) -> Path:
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination_path = _unique_path(destination_dir / source_path.name)
    shutil.move(str(source_path), str(destination_path))
    if not destination_path.exists():
        raise AssertionError(f"failed to move {source_path}")
    return destination_path


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 10000):
        candidate = path.with_name(f"{path.stem}_{index:03d}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise ValueError(f"cannot create unique path for {path}")


def _clean_html_text(value: str) -> str:
    without_tags = re.sub(r"<[^>]+>", "", value)
    return re.sub(r"\s+", " ", without_tags).strip()


def _outcome(file_name: str, status: str, priority: str, project: str, material_type: str, archive_path: str, reason: str) -> ArchiveFileOutcome:
    return ArchiveFileOutcome(
        file_name=file_name,
        status=status,
        priority=priority,
        project=project,
        material_type=material_type,
        archive_path=archive_path,
        reason=reason,
    )


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    package_root = Path(__file__).resolve().parent.parent / "ai_report_config_materials"
    result = run_archive_intake(ArchiveIntakeConfig(package_root))
    print(f"已归档：{result.processed_count}")
    print(f"无法判定：{result.unresolved_count}")
    print(f"已跳过：{result.skipped_count}")
    print(f"资料索引：{result.index_path}")
    print(f"数据字典：{result.dictionary_path}")
    for outcome in result.outcomes:
        print(f"{outcome.status} | {outcome.file_name} | {outcome.priority} | {outcome.project} | {outcome.material_type}")


if __name__ == "__main__":
    main()
