from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

from .domain import ValidationError


def read_table(file_name: str, stream: BinaryIO) -> list[dict[str, str]]:
    if not isinstance(file_name, str) or not file_name.strip():
        raise ValueError("file_name must not be empty")
    if not hasattr(stream, "read"):
        raise TypeError("stream must be file-like")
    suffix = Path(file_name).suffix.lower()
    content = stream.read()
    if not isinstance(content, bytes):
        raise TypeError("stream.read() must return bytes")
    if not content:
        raise ValidationError("上传文件不能为空")
    if suffix == ".csv":
        rows = _read_csv(content)
    elif suffix == ".xlsx":
        rows = _read_xlsx(content)
    else:
        raise ValidationError("当前只支持 .csv 或 .xlsx 文件")
    if not rows:
        raise ValidationError("上传文件没有可处理的数据行")
    assert rows
    return rows


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    if not isinstance(path, Path):
        raise TypeError("path must be pathlib.Path")
    if not isinstance(rows, list):
        raise TypeError("rows must be list")
    if not rows:
        raise ValidationError("结果数据不能为空")
    fieldnames = list(rows[0].keys())
    if not fieldnames:
        raise ValidationError("结果数据缺少字段")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    if not path.exists() or path.stat().st_size == 0:
        raise AssertionError("failed to write csv result")


def _read_csv(content: bytes) -> list[dict[str, str]]:
    text = _decode_csv_text(content)
    reader = csv.DictReader(io.StringIO(text))
    rows = [{str(k).strip(): ("" if v is None else str(v).strip()) for k, v in row.items()} for row in reader]
    return [row for row in rows if any(value for value in row.values())]


def _decode_csv_text(content: bytes) -> str:
    if not isinstance(content, bytes) or not content:
        raise ValueError("content must be non-empty bytes")
    for encoding in ("utf-8-sig", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValidationError("CSV 文件编码暂不支持，请另存为 UTF-8 或 GBK")


def _read_xlsx(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                continue
            headers = ["" if value is None else str(value).strip() for value in header_row]
            if not any(headers):
                continue
            rows: list[dict[str, str]] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = {
                    headers[index]: "" if value is None else str(value).strip()
                    for index, value in enumerate(values)
                    if index < len(headers) and headers[index]
                }
                if any(row.values()):
                    rows.append(row)
            if rows:
                return rows
    finally:
        workbook.close()
    raise ValidationError("Excel 缺少可读取的数据表")
