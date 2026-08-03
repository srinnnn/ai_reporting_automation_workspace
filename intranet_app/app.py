from __future__ import annotations

import csv
import hashlib
import html
import json
import logging
import mimetypes
import re
import shutil
import sys
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from email import policy
from email.parser import BytesParser
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Callable
from urllib.parse import parse_qs, quote, unquote, urlparse

from backend.adapters.report_task_adapter import build_daily_report_task_payload
from backend.core.config import load_core_config
from backend.core.container import ApplicationContainer, build_application_container
from backend.repositories.sqlite.foundation_repository import SQLiteFoundationRepository
from backend.repositories.sqlite.report_repository import SQLiteReportRepository
from backend.repositories.sqlite.task_repository import SQLiteTaskRepository
from backend.services.ai_content_service import AIContentService
from backend.services.assets.asset_service import ResultAssetService
from backend.services.assets.providers.local_provider import LocalStorageProvider
from backend.services.ai_service import AIService
from backend.services.dashboard_service import DashboardService
from backend.services.data_foundation_service import DataFoundationService
from backend.services.permission_service import PermissionService
from backend.services.report_service import ReportService
from backend.services.system_status_service import SystemStatusService
from backend.services.task_console_service import TaskConsoleFilters, TaskConsoleService
from backend.services.task_query_service import TaskQueryService
from backend.services.task_result_service import TaskResultService
from backend.services.task_service import TaskService
from backend.workers.contracts import TaskResult, TaskType, WorkerTaskStatus
from backend.workers.executors.ai_content_executor import AIContentExecutor
from backend.workers.executors.data_import_executor import DataImportExecutor
from backend.workers.executors.report_executor import ReportExecutor
from backend.workers.task_runner import TaskRunner
from backend.workers.task_submitter import TaskSubmitter

from .ai_gateway import AiGatewayError, BailianClient, BailianSettings, save_bailian_api_key
from .archive_intake import ArchiveIntakeConfig, ArchiveIntakeResult, ensure_intake_workspace, rebuild_archive_catalog, run_archive_intake
from .auth import new_session_token, verify_password
from .config import AppConfig, DEFAULT_CONFIG
from .content_pipeline import ANTA_DEFAULT_BRAND_PROFILE, DEFAULT_FORBIDDEN_WORDS, MODULE_KEY as P2_CONTENT_MODULE_KEY, P2ContentRequest, build_p2_content_pack
from .data_foundation import IngestionPlan, UploadMetadata, build_ingestion_plan
from .domain import ProcessingResult, ValidationError, ensure_runtime_dirs
from .io_utils import read_table, write_csv
from .processors import ai_selection, anta_blacklist, anta_listing, anta_meituan_reporting, anta_reporting, bosch_sms, copy_content
from .roadmap import CAPABILITY_STATUSES, MATERIAL_REQUIREMENTS, ROADMAP_WEEKS, daily_task_date, roadmap_day_count
from .scenarios import ANTA_RETAIL_KEY, build_scenarios
from .storage import AppStorage, AutomationTaskRecord, EfficiencyMappingRecord, JobRecord, ProjectFeedbackRecord, UserRecord, _normalize_duration_hours_text
from tools.meituan_download_assistant_sync import MEITUAN_REPORT_KEYWORDS, SyncConfig, SyncedFile, copy_new_files, default_config as default_meituan_sync_config


Processor = Callable[[list[dict[str, str]]], ProcessingResult]
ANTA_RETAIL_DEFAULT_URL = "http://127.0.0.1:8766"
PRIORITY_SECTIONS = (
    ("P1", "数据提效", "统一沉淀基础数据层，优先自动化日报、周报、月报和指标分析。"),
    ("P2", "内容提效", "围绕选品、策略、卖点、文案和视觉Brief做内容生产提效。"),
    ("P3", "配置提效", "用规则引擎和网页工具承接重复配置，并保留人工确认节点。"),
    ("P4", "巡查", "沉淀页面巡查、复盘和问题判断口径，成熟后逐步自动化。"),
)
PLATFORM_SECTIONS = (
    ("CRM", "覆盖短信、彩信、会员触达和 CRM 数据处理项目。"),
    ("美团", "覆盖美团渠道的数据报表、即时零售和商品配置项目。"),
    ("京东", "覆盖京东渠道的数据报表、即时零售和商品配置项目。"),
    ("天猫", "覆盖天猫渠道的数据报表、页面巡检和内容配置项目。"),
    ("小程序", "覆盖品牌小程序渠道的数据报表、页面巡检和内容配置项目。"),
    ("官网", "覆盖品牌官网渠道的数据报表、页面巡检和内容配置项目。"),
    ("飞猪", "覆盖飞猪渠道的品牌资料、会员触达和运营数据项目。"),
    ("经销", "覆盖经销渠道的品牌销售、数据处理和内容支持项目。"),
    ("企微/社群", "覆盖企微、社群触达、社群维护和私域内容项目。"),
    ("多渠道通用", "覆盖不绑定单一渠道、可跨品牌跨平台复用的项目。"),
)
PROJECT_STAGE_ROWS = (
    ("P1", "开发完成", "博西短彩信数据处理", "博西", "数据处理"),
    ("P1", "已经开发", "安踏周报/月报", "安踏儿童", "数据处理"),
    ("P2", "正在开发", "AI选品辅助", "多品牌", "AI选品"),
    ("P2", "正在开发", "文案内容辅助", "多品牌", "文案内容"),
    ("P3", "开发完成", "安踏即时零售", "安踏即时零售", "配置自动化"),
    ("P4", "正在开发", "页面巡检复盘", "多品牌", "页面巡检复盘"),
)
PROJECT_STAGE_HEADERS = ("项目", "优先级", "原人工耗时", "现在处理耗时", "业务反馈", "后续迭代需求")
CHANNEL_BRAND_DISTRIBUTION_PATH = Path.home() / "Desktop" / "工作文件" / "西门子" / "短信数据" / "5.30数据done" / "平台-品牌-渠道分布.xlsx"
DEVELOPMENT_PLAN_PATH = Path(__file__).resolve().parent.parent / "AI自动化替代开发规划与排期_含品牌业务方_不用API流程_含资料收集规范.xlsx"

PROCESSORS: dict[str, Processor] = {
    bosch_sms.MODULE_KEY: bosch_sms.process,
    anta_listing.MODULE_KEY: anta_listing.process,
    anta_blacklist.MODULE_KEY: anta_blacklist.process,
    ai_selection.MODULE_KEY: ai_selection.process,
    copy_content.MODULE_KEY: copy_content.process,
}


def _report_task_mode() -> str:
    try:
        return load_core_config().report_task_mode
    except (TypeError, ValueError) as exc:
        logging.error("report task mode config failed closed to legacy: %s", exc)
        return "legacy"


@dataclass(frozen=True)
class UploadedFile:
    file_name: str
    content: bytes

    def __post_init__(self) -> None:
        if not self.file_name.strip():
            raise ValueError("file_name must not be empty")
        if not self.content:
            raise ValidationError("上传文件不能为空")


@dataclass(frozen=True)
class RequestContext:
    user: UserRecord | None
    token: str | None


@dataclass(frozen=True)
class GroupProjectTreeItem:
    priority: str
    branch: str
    project: str
    brand: str
    business_type: str
    original_hours: Decimal
    source_detail: str

    def __post_init__(self) -> None:
        for field_name, field_value in (
            ("priority", self.priority),
            ("branch", self.branch),
            ("project", self.project),
            ("brand", self.brand),
            ("business_type", self.business_type),
            ("source_detail", self.source_detail),
        ):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if self.priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        if not isinstance(self.original_hours, Decimal):
            raise TypeError("original_hours must be Decimal")
        if self.original_hours < Decimal("0"):
            raise ValueError("original_hours must not be negative")


@dataclass(frozen=True)
class ChannelBrandNode:
    brand: str
    platform: str
    is_developed: bool
    source_project_count: int

    def __post_init__(self) -> None:
        for field_name, field_value in (("brand", self.brand), ("platform", self.platform)):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if not isinstance(self.is_developed, bool):
            raise TypeError("is_developed must be bool")
        if not isinstance(self.source_project_count, int):
            raise TypeError("source_project_count must be int")
        if self.source_project_count <= 0:
            raise ValueError("source_project_count must be positive")


@dataclass(frozen=True)
class WorkItemCoverage:
    channels: tuple[str, ...]
    brands: tuple[str, ...]

    def __post_init__(self) -> None:
        for field_name, values in (("channels", self.channels), ("brands", self.brands)):
            if not isinstance(values, tuple):
                raise TypeError(f"{field_name} must be tuple")
            for value in values:
                if not isinstance(value, str) or not value.strip():
                    raise ValueError(f"{field_name} item must not be empty")


@dataclass(frozen=True)
class CompletedFeedbackItem:
    priority: str
    project: str
    brand: str
    feedback_key: str
    legacy_project: str
    manual_time_project: str

    def __post_init__(self) -> None:
        for field_name, field_value in (
            ("priority", self.priority),
            ("project", self.project),
            ("brand", self.brand),
            ("feedback_key", self.feedback_key),
            ("legacy_project", self.legacy_project),
            ("manual_time_project", self.manual_time_project),
        ):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if self.priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")


@dataclass(frozen=True)
class PriorityDevelopmentStats:
    priority: str
    total_count: int
    developed_count: int
    pending_count: int

    def __post_init__(self) -> None:
        if self.priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        for field_name, field_value in (
            ("total_count", self.total_count),
            ("developed_count", self.developed_count),
            ("pending_count", self.pending_count),
        ):
            if not isinstance(field_value, int) or field_value < 0:
                raise ValueError(f"{field_name} must be a non-negative int")
        if self.developed_count > self.total_count:
            raise ValueError("developed_count must not exceed total_count")
        if self.pending_count != self.total_count - self.developed_count:
            raise ValueError("pending_count must equal total_count - developed_count")


@dataclass(frozen=True)
class ProjectPlatformGroup:
    platform: str
    description: str
    brands: tuple[ChannelBrandNode, ...]
    source_project_count: int

    def __post_init__(self) -> None:
        if not isinstance(self.platform, str) or not self.platform.strip():
            raise ValueError("platform must not be empty")
        if not isinstance(self.description, str) or not self.description.strip():
            raise ValueError("description must not be empty")
        if not isinstance(self.brands, tuple):
            raise TypeError("brands must be tuple")
        for brand in self.brands:
            if not isinstance(brand, ChannelBrandNode):
                raise TypeError("brand must be ChannelBrandNode")
        if not isinstance(self.source_project_count, int):
            raise TypeError("source_project_count must be int")
        if self.source_project_count <= 0:
            raise ValueError("source_project_count must be positive")


@dataclass(frozen=True)
class EfficiencyMappingBrand:
    task_name: str
    brand_name: str
    business_owners: tuple[str, ...]
    is_improved: bool
    note: EfficiencyMappingRecord | None

    def __post_init__(self) -> None:
        for field_name, field_value in (("task_name", self.task_name), ("brand_name", self.brand_name)):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if not isinstance(self.business_owners, tuple):
            raise TypeError("business_owners must be tuple")
        for owner in self.business_owners:
            if not isinstance(owner, str) or not owner.strip():
                raise ValueError("business_owners must contain non-empty strings")
        if not isinstance(self.is_improved, bool):
            raise TypeError("is_improved must be bool")
        if self.note is not None and not isinstance(self.note, EfficiencyMappingRecord):
            raise TypeError("note must be EfficiencyMappingRecord or None")


@dataclass(frozen=True)
class EfficiencyMappingTask:
    priority: str
    replacement_type: str
    task_name: str
    monthly_hours: Decimal
    replacement_reason: str
    brands: tuple[EfficiencyMappingBrand, ...]

    def __post_init__(self) -> None:
        for field_name, field_value in (
            ("priority", self.priority),
            ("replacement_type", self.replacement_type),
            ("task_name", self.task_name),
            ("replacement_reason", self.replacement_reason),
        ):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if self.priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        if not isinstance(self.monthly_hours, Decimal):
            raise TypeError("monthly_hours must be Decimal")
        if self.monthly_hours < Decimal("0"):
            raise ValueError("monthly_hours must not be negative")
        if not isinstance(self.brands, tuple) or not self.brands:
            raise ValueError("brands must be a non-empty tuple")
        for brand in self.brands:
            if not isinstance(brand, EfficiencyMappingBrand):
                raise TypeError("brand must be EfficiencyMappingBrand")


GROUP_PROJECT_TREE_ITEMS: tuple[GroupProjectTreeItem, ...] = (
    GroupProjectTreeItem("P1", "数据报表", "日报（日常销售报数）", "全组多品牌", "数据处理", Decimal("658"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "周报（周数据整合）", "全组多品牌", "数据处理", Decimal("640"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "月报（月报表整合/规划项）", "全组多品牌", "数据处理", Decimal("472"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "活动复盘（含临时需求及活动数据整合）", "全组多品牌", "数据处理", Decimal("335"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "往期活动数据对比（季度及周环比）", "全组多品牌", "数据处理", Decimal("282"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "货盘整合（货盘数据及日常数据处理）", "全组多品牌", "数据处理", Decimal("253"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "数据报表", "项目对账（平台及品牌结算对账）", "全组多品牌", "数据处理", Decimal("255"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P1", "短彩信数据", "短彩信数据追踪", "全组多品牌", "数据处理", Decimal("148"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "AI选品", "日常选品（页面+社群推送）", "全组多品牌", "AI选品", Decimal("574"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "内容生产", "日常内容优化（视觉项展示/文案项调整）", "全组多品牌", "文案内容", Decimal("542"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "内容生产", "社群文案撰写", "全组多品牌", "文案内容", Decimal("356"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "内容生产", "线框搭建（页面+社群推送）", "全组多品牌", "视觉Brief", Decimal("328"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "内容生产", "活动内容整合", "全组多品牌", "内容资料", Decimal("214"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "内容生产", "短彩信触达文案撰写", "全组多品牌", "文案内容", Decimal("66"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "策略辅助", "竞品分析", "全组多品牌", "策略分析", Decimal("255"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "策略辅助", "社群触达排期规划", "全组多品牌", "策略分析", Decimal("115"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P2", "策略辅助", "短彩信触达规划", "全组多品牌", "策略分析", Decimal("22"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "活动机制配置", "全组多品牌", "配置自动化", Decimal("430"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "活动商品配置", "全组多品牌", "配置自动化", Decimal("433"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "上下架处理", "全组多品牌", "配置自动化", Decimal("372"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "短彩信发信配置", "全组多品牌", "配置自动化", Decimal("178"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "短彩信模板配置", "全组多品牌", "配置自动化", Decimal("147"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "社群推送设置", "全组多品牌", "配置自动化", Decimal("192"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "社群活动配置", "全组多品牌", "配置自动化", Decimal("183"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "触达人群圈选", "全组多品牌", "配置自动化", Decimal("166"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "会员活动配置", "全组多品牌", "配置自动化", Decimal("164"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "库存调整/发货处理", "全组多品牌", "配置自动化", Decimal("170"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "优惠券配置", "全组多品牌", "配置自动化", Decimal("155"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "商品价格调整", "全组多品牌", "配置自动化", Decimal("148"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "OMS配置操作", "全组多品牌", "配置自动化", Decimal("147"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "门店/导购信息配置", "全组多品牌", "配置自动化", Decimal("109"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "导购推送设置", "全组多品牌", "配置自动化", Decimal("77"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "直播间配置", "全组多品牌", "配置自动化", Decimal("199"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "SI金额及数量预提", "全组多品牌", "配置自动化", Decimal("90"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "采购PR申请及下单流程", "全组多品牌", "配置自动化", Decimal("80"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P3", "配置自动化", "社群1v1推送", "全组多品牌", "配置自动化", Decimal("28"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "页面巡检复盘", "页面装修", "全组多品牌", "页面巡检复盘", Decimal("443"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "页面巡检复盘", "巡店（会员页面/大促前后/日常巡店）", "全组多品牌", "页面巡检复盘", Decimal("438"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "页面巡检复盘", "大促/活动档期前后整体页面优化", "全组多品牌", "页面巡检复盘", Decimal("277"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "页面巡检复盘", "货品进销存管理", "全组多品牌", "页面巡检复盘", Decimal("135"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "人工协同", "项目协同对接（多维度协同沟通）", "全组多品牌", "人工协同", Decimal("853"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "人工协同", "活动规划拆解（店铺活动及触达预算拆解）", "全组多品牌", "人工协同", Decimal("88"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "人工协同", "社群维护管理", "全组多品牌", "人工协同", Decimal("89"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
    GroupProjectTreeItem("P4", "人工协同", "临时需求", "全组多品牌", "人工协同", Decimal("35"), "内容任务耗时统计.xlsx / 运营 / 工作内容"),
)

NON_AI_EFFICIENCY_WORK_ITEMS = frozenset(
    (
        "项目协同对接（多维度协同沟通）",
        "触达人群圈选",
        "社群触达排期规划",
        "活动规划拆解（店铺活动及触达预算拆解）",
        "临时需求",
        "短彩信触达规划",
    )
)


@dataclass(frozen=True)
class LocalReportSource:
    kind: str
    path: Path
    rows: list[dict[str, str]]
    start_date: str
    end_date: str

    def __post_init__(self) -> None:
        if not isinstance(self.path, Path):
            raise TypeError("path must be pathlib.Path")
        if not isinstance(self.rows, list) or not self.rows:
            raise ValidationError("report source rows must not be empty")
        for field_name, field_value in (
            ("kind", self.kind),
            ("start_date", self.start_date),
            ("end_date", self.end_date),
        ):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if self.start_date > self.end_date:
            raise ValidationError("source start_date must not be later than end_date")


class IntranetApp:
    def __init__(self, config: AppConfig, container: ApplicationContainer | None = None) -> None:
        if not isinstance(config, AppConfig):
            raise TypeError("config must be AppConfig")
        if container is not None and not isinstance(container, ApplicationContainer):
            raise TypeError("container must be ApplicationContainer")
        self.container = container
        self.config = config
        self.storage = AppStorage(config.database_path)
        self.scenarios = build_scenarios(config.template_root)

    def initialize(self) -> None:
        ensure_runtime_dirs((self.config.upload_dir, self.config.result_dir))
        self.storage.initialize(self.config.default_admin_password)
        logging.info("intranet app initialized")

    def close(self) -> None:
        container = getattr(self, "container", None)
        if isinstance(container, ApplicationContainer):
            container.close()

    def make_handler(self) -> type[BaseHTTPRequestHandler]:
        app = self

        class Handler(BaseHTTPRequestHandler):
            def do_GET(self) -> None:
                app.handle_get(self)

            def do_POST(self) -> None:
                app.handle_post(self)

            def log_message(self, fmt: str, *args: object) -> None:
                logging.info("%s - %s", self.address_string(), fmt % args)

        return Handler

    def handle_get(self, handler: BaseHTTPRequestHandler) -> None:
        path = urlparse(handler.path).path
        context = self._context(handler)
        if path == "/login":
            self._send_html(handler, self._login_page(""))
            return
        if path == "/static/style.css":
            self._send_file(handler, Path(__file__).resolve().parent / "static" / "style.css")
            return
        if path == "/logout":
            if context.token:
                self.storage.delete_session(context.token)
            self._redirect(handler, "/login", clear_cookie=True)
            return
        if context.user is None:
            if path.startswith("/api/"):
                self._send_json(handler, {"error": "unauthorized"}, status=401)
                return
            self._redirect(handler, "/login")
            return
        if path == "/":
            self._send_html(handler, self._dashboard(context.user))
            return
        if path == "/console":
            self._send_html(handler, self._console_dashboard_page(context.user))
            return
        if path == "/console/tasks":
            self._send_html(handler, self._console_tasks_page(context.user))
            return
        if path == "/console/environment":
            self._send_html(handler, self._console_environment_page(context.user))
            return
        if path == "/archive-intake":
            self._send_html(handler, self._archive_intake_page(context.user, ""))
            return
        if path == "/data-foundation":
            self._send_html(handler, self._data_foundation_page(context.user, "", ""))
            return
        if path == "/automation-runs":
            self._send_html(handler, self._automation_runs_page(context.user, "", ""))
            return
        if path == "/p2-content-center":
            self._send_html(handler, self._p2_content_center_page(context.user, ""))
            return
        if path == "/archive-index":
            self._send_html(handler, self._archive_csv_page(context.user, "资料索引", self._archive_index_path(), "所有已归档资料的自动登记清单。", "/archive-index/download"))
            return
        if path == "/data-dictionary":
            self._send_html(handler, self._archive_csv_page(context.user, "数据字典", self._data_dictionary_path(), "系统从已归档表格中识别出的字段清单和来源。", "/data-dictionary/download"))
            return
        if path == "/project-stages":
            self._send_html(handler, self._project_stages_page(context.user, "", ""))
            return
        if path == "/work-item-planning":
            self._send_html(handler, self._work_item_planning_page(context.user))
            return
        if path == "/efficiency-mapping":
            self._send_html(handler, self._efficiency_mapping_page(context.user, ""))
            return
        if path == "/development-roadmap":
            self._send_html(handler, self._development_roadmap_page(context.user))
            return
        if path == "/development-roadmap/download":
            roadmap_path = self._development_roadmap_workbook_path()
            if roadmap_path.exists():
                self._send_file(handler, roadmap_path, download_name=roadmap_path.name)
                return
            self._send_html(handler, self._page("排期文件不存在", "<p>详细排期文件尚未生成。</p>"), status=404)
            return
        if path == "/admin/ai-settings":
            self._send_html(handler, self._ai_settings_page(context.user, "", "", self._is_loopback_request(handler)))
            return
        if path == "/archive-index/download":
            self._send_file(handler, self._archive_index_path(), download_name="archive_index.csv")
            return
        if path == "/data-dictionary/download":
            self._send_file(handler, self._data_dictionary_path(), download_name="data_dictionary.csv")
            return
        if path == "/api/console/dashboard":
            self._handle_console_dashboard_api(handler, context.user)
            return
        if path == "/api/system/health":
            self._handle_system_health_api(handler, context.user)
            return
        if path == "/api/system/config/status":
            self._handle_system_config_status_api(handler, context.user)
            return
        if path == "/tasks":
            self._send_html(handler, self._tasks_page(context.user))
            return
        if path.startswith("/tasks/"):
            self._send_html(handler, self._task_detail_page(context.user, path))
            return
        if path.startswith("/api/tasks/") and path.endswith("/download"):
            self._handle_task_api_download(handler, path, context.user)
            return
        if path.startswith("/api/tasks/"):
            self._handle_task_api_get(handler, path, context.user)
            return
        if path == "/api/tasks":
            self._handle_task_api_list(handler, context.user)
            return
        if path == "/anta-retail":
            self._send_html(handler, self._anta_retail_page(context.user, "", self._anta_retail_url(handler)))
            return
        if path == "/anta-reporting":
            self._send_html(handler, self._anta_reporting_page(context.user, ""))
            return
        if path.startswith("/scenario/") and path.endswith("/template"):
            scenario_key = unquote(path.removeprefix("/scenario/").removesuffix("/template")).strip("/")
            scenario = self.scenarios.get(scenario_key)
            if scenario is not None and scenario.template_path.exists():
                self._send_file(handler, scenario.template_path, download_name=scenario.template_path.name)
                return
            self._send_html(handler, self._page("模板不存在", "<p>当前场景尚未配置可下载模板。</p>"), status=404)
            return
        if path.startswith("/priority/"):
            priority = unquote(path.removeprefix("/priority/")).strip("/").upper()
            if priority in {item[0] for item in PRIORITY_SECTIONS}:
                self._send_html(handler, self._priority_page(context.user, priority))
                return
        if path.startswith("/scenario/"):
            scenario_key = unquote(path.removeprefix("/scenario/")).strip("/")
            if scenario_key in self.scenarios:
                if scenario_key == ANTA_RETAIL_KEY:
                    self._redirect(handler, "/anta-retail")
                    return
                self._send_html(handler, self._scenario_page(context.user, scenario_key, ""))
                return
        if path.startswith("/jobs/") and path.endswith("/download"):
            self._download_job_result(handler, path)
            return
        self._send_html(handler, self._page("未找到页面", "<p>页面不存在。</p>"), status=404)

    def handle_post(self, handler: BaseHTTPRequestHandler) -> None:
        path = urlparse(handler.path).path
        if path == "/login":
            self._handle_login(handler)
            return
        context = self._context(handler)
        if context.user is None:
            if path.startswith("/api/"):
                self._send_json(handler, {"error": "unauthorized"}, status=401)
                return
            self._redirect(handler, "/login")
            return
        if path == "/api/tasks":
            self._handle_task_api_submit(handler, context.user)
            return
        if path.startswith("/scenario/") and path.endswith("/run"):
            scenario_key = unquote(path.removeprefix("/scenario/").removesuffix("/run")).strip("/")
            self._handle_run(handler, context.user, scenario_key)
            return
        if path == "/anta-retail/blacklist/run":
            self._handle_anta_blacklist_run(handler, context.user)
            return
        if path == "/archive-intake/run":
            self._handle_archive_intake_run(handler, context.user)
            return
        if path == "/archive-catalog/rebuild":
            self._handle_archive_catalog_rebuild(handler, context.user)
            return
        if path == "/archive-intake/upload":
            self._handle_archive_intake_upload(handler, context.user)
            return
        if path == "/data-foundation/check":
            self._handle_data_foundation_check(handler, context.user)
            return
        if path == "/automation-runs/create":
            self._handle_automation_task_create(handler, context.user)
            return
        if path == "/automation-runs/toggle":
            self._handle_automation_task_toggle(handler, context.user)
            return
        if path == "/automation-runs/record":
            self._handle_automation_run_record(handler, context.user)
            return
        if path == "/automation-runs/sync":
            self._handle_automation_download_sync(handler, context.user)
            return
        if path == "/automation-runs/execute":
            self._handle_automation_execute(handler, context.user)
            return
        if path == "/p2-content-center/run":
            self._handle_p2_content_center_run(handler, context.user)
            return
        if path == "/anta-reporting/weekly/run":
            self._handle_anta_reporting_run(handler, context.user, "weekly")
            return
        if path == "/anta-reporting/monthly/run":
            self._handle_anta_reporting_run(handler, context.user, "monthly")
            return
        if path == "/anta-reporting/meituan-daily/run":
            self._handle_anta_meituan_reporting_run(handler, context.user, "daily")
            return
        if path == "/anta-reporting/meituan-weekly/run":
            self._handle_anta_meituan_reporting_run(handler, context.user, "weekly")
            return
        if path == "/admin/ai-settings/test":
            self._handle_ai_connection_test(handler, context.user)
            return
        if path == "/admin/ai-settings/save":
            self._handle_ai_key_save(handler, context.user)
            return
        if path == "/project-stages/feedback":
            self._handle_project_feedback(handler, context.user)
            return
        if path == "/efficiency-mapping/save":
            self._handle_efficiency_mapping_save(handler, context.user)
            return
        if path == "/efficiency-mapping/add-brand":
            self._handle_efficiency_mapping_brand_add(handler, context.user)
            return
        self._send_html(handler, self._page("未找到页面", "<p>页面不存在。</p>"), status=404)

    def _handle_efficiency_mapping_save(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        task_name = fields.get("task_name", [""])[0].strip()
        brand_name = fields.get("brand_name", [""])[0].strip()
        valid_pairs = self._efficiency_mapping_pairs()
        if (task_name, brand_name) not in valid_pairs:
            self._send_html(handler, self._dashboard(user), status=400)
            return
        try:
            self.storage.save_efficiency_mapping_note(
                task_name=task_name,
                brand_name=brand_name,
                not_improved_reason=fields.get("not_improved_reason", [""])[0],
                schedule_plan=fields.get("schedule_plan", [""])[0],
                updated_by=user.username,
                is_improved=fields.get("is_improved", [""])[0] == "on",
                is_manual_brand=fields.get("is_manual_brand", [""])[0] == "1",
            )
            self._send_html(handler, self._efficiency_mapping_page(user, "映射信息已保存。"))
        except (TypeError, ValueError) as exc:
            logging.error("failed to save efficiency mapping note: %s", exc)
            self._send_html(handler, self._efficiency_mapping_page(user, str(exc)), status=400)

    def _handle_efficiency_mapping_brand_add(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        task_name = fields.get("task_name", [""])[0].strip()
        brand_name = self._normalize_brand(fields.get("brand_name", [""])[0].strip()) if fields.get("brand_name", [""])[0].strip() else ""
        valid_tasks = {item.task_name for item in self._high_efficiency_mapping_items()}
        if task_name not in valid_tasks or not brand_name:
            self._send_html(handler, self._efficiency_mapping_page(user, "请选择任务并填写品牌名称。"), status=400)
            return
        try:
            self.storage.save_efficiency_mapping_note(
                task_name=task_name,
                brand_name=brand_name,
                not_improved_reason=fields.get("not_improved_reason", [""])[0],
                schedule_plan=fields.get("schedule_plan", [""])[0] or self._default_efficiency_schedule(task_name),
                updated_by=user.username,
                is_improved=fields.get("is_improved", [""])[0] == "on",
                is_manual_brand=True,
            )
            self._send_html(handler, self._efficiency_mapping_page(user, f"“{brand_name}”已加入“{task_name}”。"))
        except (TypeError, ValueError) as exc:
            logging.error("failed to add efficiency mapping brand: %s", exc)
            self._send_html(handler, self._efficiency_mapping_page(user, str(exc)), status=400)

    def _handle_project_feedback(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        fields = self._read_urlencoded(handler)
        project = fields.get("project", [""])[0].strip()
        return_to = fields.get("return_to", [""])[0].strip()
        valid_projects = self._project_feedback_names()
        if project not in valid_projects:
            page = self._dashboard(user) if return_to == "/" else self._project_stages_page(user, "", "项目不存在，无法保存反馈。")
            self._send_html(handler, page, status=400)
            return
        try:
            self.storage.save_project_feedback(
                project=project,
                original_manual_time=fields.get("original_manual_time", [""])[0],
                current_processing_time=fields.get("current_processing_time", [""])[0],
                business_feedback=fields.get("business_feedback", [""])[0],
                iteration_need=fields.get("iteration_need", [""])[0],
                updated_by=user.username,
            )
            if return_to == "/":
                self._send_html(handler, self._dashboard(user))
            else:
                self._send_html(handler, self._project_stages_page(user, f"“{project}”的业务反馈已保存。", ""))
        except (TypeError, ValueError) as exc:
            page = self._dashboard(user) if return_to == "/" else self._project_stages_page(user, "", str(exc))
            self._send_html(handler, page, status=400)

    def _handle_automation_task_create(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        try:
            self.storage.save_automation_task(
                task_name=fields.get("task_name", [""])[0],
                business_unit=fields.get("business_unit", [""])[0],
                brand_id=fields.get("brand_id", [""])[0],
                brand_name=fields.get("brand_name", [""])[0],
                platform=fields.get("platform", [""])[0],
                channel=fields.get("channel", [""])[0],
                file_type=fields.get("file_type", [""])[0],
                frequency=fields.get("frequency", [""])[0],
                scheduled_time=fields.get("scheduled_time", [""])[0],
                date_window=fields.get("date_window", [""])[0],
                enabled=fields.get("enabled", [""])[0] == "on",
                output_folder=fields.get("output_folder", [""])[0],
                owner=fields.get("owner", [""])[0],
                notes=fields.get("notes", [""])[0],
            )
            self._send_html(handler, self._automation_runs_page(user, "任务已新增。", ""))
        except (TypeError, ValueError) as exc:
            self._send_html(handler, self._automation_runs_page(user, "", str(exc)), status=400)

    def _handle_automation_task_toggle(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        try:
            task_id = int(fields.get("task_id", ["0"])[0])
            enabled = fields.get("enabled", [""])[0] == "1"
            self.storage.set_automation_task_enabled(task_id, enabled)
            self._send_html(handler, self._automation_runs_page(user, "任务状态已更新。", ""))
        except (TypeError, ValueError) as exc:
            self._send_html(handler, self._automation_runs_page(user, "", str(exc)), status=400)

    def _handle_automation_run_record(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        try:
            self.storage.save_automation_run(
                task_id=int(fields.get("task_id", ["0"])[0]),
                run_date=fields.get("run_date", [""])[0],
                status=fields.get("status", [""])[0],
                downloaded_file_count=int(fields.get("downloaded_file_count", ["0"])[0] or "0"),
                synced_file_count=int(fields.get("synced_file_count", ["0"])[0] or "0"),
                message=fields.get("message", [""])[0],
                executed_by=user.username,
            )
            self._send_html(handler, self._automation_runs_page(user, "执行记录已保存。", ""))
        except (TypeError, ValueError) as exc:
            self._send_html(handler, self._automation_runs_page(user, "", str(exc)), status=400)

    def _handle_automation_download_sync(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            synced = self._sync_meituan_download_sources()
            enabled_tasks = [task for task in self.storage.list_automation_tasks() if task.enabled]
            task_id = enabled_tasks[0].id if enabled_tasks else self.storage.list_automation_tasks()[0].id
            self.storage.save_automation_run(
                task_id=task_id,
                run_date=date.today().isoformat(),
                status="synced",
                downloaded_file_count=len(synced),
                synced_file_count=len(synced),
                message=f"已同步 {len(synced)} 个新文件；目标目录：{default_meituan_sync_config(Path.cwd()).target_root}",
                executed_by=user.username,
            )
            self._send_html(handler, self._automation_runs_page(user, f"下载目录同步完成：新增 {len(synced)} 个文件。", ""))
        except (TypeError, ValueError, IndexError, OSError) as exc:
            self._send_html(handler, self._automation_runs_page(user, "", str(exc)), status=400)

    def _handle_automation_execute(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        fields = self._read_urlencoded(handler)
        try:
            run_date = self._selected_meituan_report_date({"report_date": fields.get("run_date", [""])})
            synced = self._sync_meituan_download_sources()
            imported_count = self._ingest_meituan_plugin_files_to_foundation(user.username)
            enabled_tasks = [task for task in self.storage.list_automation_tasks() if task.enabled]
            if not enabled_tasks:
                raise ValidationError("没有启用的自动化任务。")
            success_count = 0
            missing_count = 0
            for task in enabled_tasks:
                if task.platform == "meituan":
                    has_data = self._foundation_has_task_data(task, run_date)
                    status = "foundation_ready" if has_data else "missing_source"
                    if has_data:
                        success_count += 1
                        message = f"{run_date} 已完成插件同步、自动入库，并在基础数据层找到 {self._file_type_label(task.file_type)}。"
                    else:
                        missing_count += 1
                        message = f"{run_date} 基础数据层缺少 {self._file_type_label(task.file_type)}。请在美团后台用插件导出该日期对应报表后再次执行。"
                else:
                    status = "unsupported"
                    message = f"{self._platform_label(task.platform)} 自动执行尚未接入插件。"
                self.storage.save_automation_run(
                    task_id=task.id,
                    run_date=f"{run_date[:4]}-{run_date[4:6]}-{run_date[6:8]}",
                    status=status,
                    downloaded_file_count=0,
                    synced_file_count=len(synced),
                    message=f"{message} 本次同步新增 {len(synced)} 个文件，自动入库 {imported_count} 个文件。",
                    executed_by=user.username,
                )
            if missing_count:
                success = f"执行完成：{success_count} 项基础数据就绪，{missing_count} 项缺少源数据。"
            else:
                success = f"执行成功：{success_count} 项任务基础数据已就绪。"
            self._send_html(handler, self._automation_runs_page(user, success, ""))
        except (ValidationError, ValueError, TypeError, OSError) as exc:
            self._send_html(handler, self._automation_runs_page(user, "", str(exc)), status=400)

    def _foundation_has_task_data(self, task: AutomationTaskRecord, compact_date: str) -> bool:
        if not isinstance(task, AutomationTaskRecord):
            raise TypeError("task must be AutomationTaskRecord")
        if len(compact_date) != 8 or not compact_date.isdigit():
            raise ValidationError("compact_date must be YYYYMMDD")
        rows = self.storage.load_meituan_foundation_rows(task.brand_id, task.platform, task.channel, task.file_type)
        if task.file_type == "product_order":
            return any(_compact_date_from_source(row.get("下单时间", "")) == compact_date for row in rows)
        if task.file_type in {"store_finance", "store_traffic"}:
            return any(
                _compact_date_from_source(row.get("开始时间", "")) <= compact_date <= _compact_date_from_source(row.get("结束时间", ""))
                for row in rows
                if _compact_date_from_source(row.get("开始时间", "")) and _compact_date_from_source(row.get("结束时间", ""))
            )
        if task.file_type == "service_review":
            return any(_compact_date_from_source(row.get("评价提交日期", "")) == compact_date for row in rows)
        return False

    def _handle_data_foundation_check(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            form_fields, uploaded_file = self._read_multipart(handler)
            metadata = UploadMetadata(
                business_unit=form_fields.get("business_unit", "").strip(),
                brand_id=form_fields.get("brand_id", "").strip(),
                brand_name=form_fields.get("brand_name", "").strip(),
                platform=form_fields.get("platform", "").strip(),
                channel=form_fields.get("channel", "").strip(),
                project_code=form_fields.get("project_code", "").strip(),
                declared_file_type=form_fields.get("declared_file_type", "").strip(),
                data_start_date=form_fields.get("data_start_date", "").strip(),
                data_end_date=form_fields.get("data_end_date", "").strip(),
                uploaded_by=user.username,
            )
            rows = read_table(uploaded_file.file_name, _BytesReader(uploaded_file.content))
            known_store_ids = self.storage.list_known_store_ids(metadata.brand_id, metadata.platform, metadata.channel)
            known_product_codes = self.storage.list_known_product_codes(metadata.brand_id, metadata.platform, metadata.channel)
            plan = build_ingestion_plan(metadata, rows, known_store_ids, known_product_codes)
            serial = _safe_serial()
            import_batch_id = f"batch_{serial}"
            foundation_dir = self.config.upload_dir / "data_foundation" / import_batch_id
            foundation_dir.mkdir(parents=True, exist_ok=True)
            stored_file_path = foundation_dir / _safe_name(uploaded_file.file_name)
            stored_file_path.write_bytes(uploaded_file.content)
            file_sha256 = hashlib.sha256(uploaded_file.content).hexdigest()
            status = "ready_for_import" if plan.validation.passed and plan.brand_match.decision == "auto_pass" else plan.brand_match.decision
            if not plan.validation.passed:
                status = "validation_failed"
            self.storage.save_foundation_check(
                import_batch_id=import_batch_id,
                metadata=metadata,
                original_file_name=uploaded_file.file_name,
                stored_file_path=stored_file_path,
                file_sha256=file_sha256,
                recognized_file_type=plan.recognition.file_type,
                row_count=len(plan.normalized_rows),
                status=status,
                brand_match_score=plan.brand_match.total_score,
                validation_errors=plan.validation.errors,
                validation_warnings=plan.validation.warnings + plan.brand_match.warnings,
            )
            if status == "ready_for_import":
                self.storage.save_foundation_fact_rows(import_batch_id, plan)
            self._send_html(handler, self._data_foundation_result_page(user, import_batch_id, uploaded_file.file_name, plan))
        except (ValidationError, ValueError, TypeError) as exc:
            logging.error("data foundation check failed: %s", exc)
            self._send_html(handler, self._data_foundation_page(user, "", str(exc)), status=400)

    def _handle_ai_key_save(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if user.role != "管理员":
            self._send_html(handler, self._page("无权访问", "<p>只有管理员可以配置AI接口。</p>"), status=403)
            return
        if not self._is_loopback_request(handler):
            self._send_html(handler, self._page("仅限本机", "<p>API Key只能在运行工作台的电脑上配置。</p>"), status=403)
            return
        fields = self._read_urlencoded(handler)
        try:
            save_bailian_api_key(fields.get("api_key", [""])[0])
            self._send_html(handler, self._ai_settings_page(user, "API Key已保存，可以测试连接。", "", True))
        except (TypeError, ValueError) as exc:
            self._send_html(handler, self._ai_settings_page(user, "", str(exc), True), status=400)

    def _handle_ai_connection_test(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if user.role != "管理员":
            self._send_html(handler, self._page("无权访问", "<p>只有管理员可以测试AI接口。</p>"), status=403)
            return
        try:
            result = BailianClient(BailianSettings.from_environment()).test_connection()
            message = f"连接成功：{result.provider} · {result.model} · 返回“{result.message}”"
            self._send_html(handler, self._ai_settings_page(user, message, "", self._is_loopback_request(handler)))
        except AiGatewayError as exc:
            self._send_html(handler, self._ai_settings_page(user, "", str(exc), self._is_loopback_request(handler)), status=400)

    def _handle_login(self, handler: BaseHTTPRequestHandler) -> None:
        fields = self._read_urlencoded(handler)
        username = fields.get("username", [""])[0].strip()
        password = fields.get("password", [""])[0]
        if not username or not password:
            self._send_html(handler, self._login_page("请输入账号和密码。"), status=400)
            return
        user = self.storage.get_user(username)
        if user is None or not verify_password(password, user.password_hash):
            logging.info("login failed for %s", username)
            self._send_html(handler, self._login_page("账号或密码不正确。"), status=403)
            return
        token = new_session_token()
        self.storage.create_session(token, user.username)
        self._redirect(handler, "/", cookie_value=token)

    def _handle_run(self, handler: BaseHTTPRequestHandler, user: UserRecord, scenario_key: str) -> None:
        if scenario_key not in self.scenarios:
            self._send_html(handler, self._page("场景不存在", "<p>未找到对应处理入口。</p>"), status=404)
            return
        if scenario_key not in PROCESSORS:
            self._send_html(handler, self._page("入口不可上传", "<p>该项目请从项目页面进入对应网页工具。</p>"), status=400)
            return
        try:
            form_fields, uploaded_file = self._read_multipart(handler)
            scenario = self.scenarios[scenario_key]
            rows = read_table(uploaded_file.file_name, _BytesReader(uploaded_file.content))
            if scenario_key == "ai_selection":
                project = form_fields.get("project", "通用选品").strip() or "通用选品"
                rows = [dict(row, 项目=project) for row in rows]
            processor = PROCESSORS[scenario_key]
            result = processor(rows)
            title = form_fields.get("title", scenario.name).strip() or scenario.name
            serial = _safe_serial()
            input_path = self.config.upload_dir / f"{serial}_{_safe_name(uploaded_file.file_name)}"
            result_path = self.config.result_dir / f"{serial}_{scenario_key}_result.csv"
            input_path.write_bytes(uploaded_file.content)
            write_csv(result_path, result.output_rows)
            job_id = self.storage.save_job(
                module=scenario.key,
                title=title,
                brand=scenario.brand,
                business_type=scenario.business_type,
                created_by=user.username,
                input_file=input_path,
                result_file=result_path,
                summary=result.summary,
                warnings=result.warnings,
            )
            self._send_html(handler, self._result_page(user, job_id, result))
        except (ValidationError, ValueError, TypeError) as exc:
            logging.error("processing failed: %s", exc)
            self._send_html(handler, self._scenario_page(user, scenario_key, str(exc)), status=400)

    def _handle_anta_reporting_run(self, handler: BaseHTTPRequestHandler, user: UserRecord, report_type: str) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if report_type not in ("weekly", "monthly"):
            raise ValueError("report_type must be weekly or monthly")
        try:
            if report_type == "weekly":
                result, input_path = self._build_anta_weekly_report()
                title = "安踏周报初稿"
            else:
                result, input_path = self._build_anta_monthly_report()
                title = "安踏月报初稿"
            serial = _safe_serial()
            result_path = self.config.result_dir / f"{serial}_{report_type}_anta_report.csv"
            write_csv(result_path, result.output_rows)
            scenario = self.scenarios[anta_reporting.MODULE_KEY]
            job_id = self.storage.save_job(
                module=scenario.key,
                title=title,
                brand=scenario.brand,
                business_type=scenario.business_type,
                created_by=user.username,
                input_file=input_path,
                result_file=result_path,
                summary=result.summary,
                warnings=result.warnings,
            )
            self._send_html(handler, self._result_page(user, job_id, result))
        except (ValidationError, ValueError, TypeError, FileNotFoundError) as exc:
            logging.error("anta reporting failed: %s", exc)
            self._send_html(handler, self._anta_reporting_page(user, str(exc)), status=400)

    def _handle_anta_meituan_reporting_run(self, handler: BaseHTTPRequestHandler, user: UserRecord, report_type: str) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if report_type not in {"daily", "weekly"}:
            raise ValueError("report_type must be daily or weekly")
        try:
            fields = self._read_urlencoded(handler)
            selected_report_date = self._selected_meituan_report_date(fields) if report_type == "daily" else ""
            if report_type == "daily" and _report_task_mode() == "task":
                task_result = self._submit_anta_meituan_daily_report_task(selected_report_date, user)
                self._send_html(handler, self._task_result_page(user, task_result))
                return
            synced_files = self._sync_meituan_download_sources()
            logging.info("synced %s meituan plugin files before report generation", len(synced_files))
            imported_count = self._ingest_meituan_plugin_files_to_foundation(user.username)
            logging.info("imported %s meituan plugin files into foundation before report generation", imported_count)
            sources, selected_files = self._load_anta_meituan_sources_from_foundation(report_type, selected_report_date)
            if report_type == "daily":
                result = anta_meituan_reporting.build_meituan_daily_report(sources, selected_files["product"].end_date)
                title = "安踏美团日报"
            else:
                result = anta_meituan_reporting.build_meituan_weekly_report(
                    sources,
                    selected_files["product"].start_date,
                    selected_files["product"].end_date,
                )
                title = "安踏美团周报"
            serial = _safe_serial()
            result_path = self.config.result_dir / f"{serial}_anta_meituan_{report_type}_report.csv"
            source_manifest_path = self.config.upload_dir / f"{serial}_anta_meituan_{report_type}_sources.json"
            source_manifest_path.write_text(
                json.dumps(
                    {
                        key: {
                            "path": str(value.path),
                            "start_date": value.start_date,
                            "end_date": value.end_date,
                            "rows": len(value.rows),
                        }
                        for key, value in selected_files.items()
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            write_csv(result_path, result.output_rows)
            job_id = self.storage.save_job(
                module=anta_meituan_reporting.MODULE_KEY,
                title=title,
                brand="安踏儿童",
                business_type="美团即时零售报表",
                created_by=user.username,
                input_file=source_manifest_path,
                result_file=result_path,
                summary=result.summary,
                warnings=result.warnings,
            )
            self._send_html(handler, self._result_page(user, job_id, result))
        except (ValidationError, ValueError, TypeError, FileNotFoundError, OSError) as exc:
            logging.error("anta meituan reporting failed: %s", exc)
            self._send_html(handler, self._anta_reporting_page(user, str(exc)), status=400)

    def _submit_anta_meituan_daily_report_task(self, report_date: str, user: UserRecord) -> TaskResult:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        payload = build_daily_report_task_payload(report_date, user)
        result = self._task_submitter().submit(TaskType.REPORT_GENERATE, payload, user.username)
        assert isinstance(result, TaskResult)
        return result

    def _handle_console_dashboard_api(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            self._send_json(handler, self._dashboard_service().get_dashboard(user))
        except PermissionError:
            self._send_json(handler, {"error": "forbidden"}, status=403)
        except (ValueError, TypeError, RuntimeError) as exc:
            logging.error("console dashboard api failed: %s", exc)
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_system_health_api(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            self._send_json(handler, self._system_status_service().get_health_status(user))
        except PermissionError:
            self._send_json(handler, {"error": "forbidden"}, status=403)
        except (ValueError, TypeError, RuntimeError) as exc:
            logging.error("system health api failed: %s", exc)
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_system_config_status_api(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            self._send_json(handler, self._system_status_service().get_config_status(user))
        except PermissionError:
            self._send_json(handler, {"error": "forbidden"}, status=403)
        except (ValueError, TypeError, RuntimeError) as exc:
            logging.error("system config status api failed: %s", exc)
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_task_api_submit(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            request = self._read_json(handler)
            task_type = _required_json_text(request, "task_type")
            payload = _required_json_object(request, "payload")
            created_by = _required_json_text(request, "created_by")
            normalized_payload = dict(payload)
            normalized_payload["output_folder"] = str(self.config.result_dir)
            if not self._permission_service().can_submit_task(user, task_type, normalized_payload):
                self._send_json(handler, {"error": "forbidden"}, status=403)
                return
            result = self._task_submitter().submit(task_type, normalized_payload, created_by)
            self._send_json(handler, {"task_id": result.task_id, "status": result.status.value})
        except (json.JSONDecodeError, ValueError, TypeError, RuntimeError) as exc:
            logging.error("task api submit failed: %s", exc)
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_task_api_list(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            filters = _task_console_filters_from_query(handler.path)
            self._send_json(handler, self._task_console_service().list_visible_tasks(user, filters))
        except PermissionError:
            self._send_json(handler, {"error": "forbidden"}, status=403)
        except (ValueError, TypeError) as exc:
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_task_api_get(self, handler: BaseHTTPRequestHandler, path: str, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            task_id = _task_id_from_api_path(path, expected_parts=3)
            self._send_json(handler, self._task_console_service().get_task_detail(user, task_id))
        except FileNotFoundError as exc:
            self._send_json(handler, {"error": str(exc)}, status=404)
        except PermissionError:
            self._send_json(handler, {"error": "forbidden"}, status=403)
        except (ValueError, TypeError) as exc:
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _handle_task_api_download(self, handler: BaseHTTPRequestHandler, path: str, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            task_id = _task_id_from_api_path(path, expected_parts=4)
            task = self._task_query_service().get_task(task_id)
            if task is None:
                raise FileNotFoundError(str(task_id))
            if not self._permission_service().can_download_task(user, task):
                self._send_json(handler, {"error": "forbidden"}, status=403)
                return
            info = self._task_result_service().get_download_info(task_id)
            self._send_file(handler, info.path, download_name=info.filename)
        except FileNotFoundError as exc:
            self._send_json(handler, {"error": str(exc)}, status=404)
        except (ValueError, TypeError, PermissionError) as exc:
            self._send_json(handler, {"error": str(exc)}, status=400)

    def _dashboard_service(self) -> DashboardService:
        service = DashboardService(self._system_status_service(), self._task_query_service(), self._permission_service())
        assert isinstance(service, DashboardService)
        return service

    def _system_status_service(self) -> SystemStatusService:
        container = getattr(self, "container", None)
        if not isinstance(container, ApplicationContainer):
            raise RuntimeError("system status service requires ApplicationContainer")
        service = SystemStatusService(container, self._permission_service())
        assert isinstance(service, SystemStatusService)
        return service

    def _task_console_service(self) -> TaskConsoleService:
        service = TaskConsoleService(self._task_query_service(), self._task_result_service(), self._permission_service())
        assert isinstance(service, TaskConsoleService)
        return service

    def _task_submitter(self) -> TaskSubmitter:
        services = self._container_services()
        if services is not None and services.task_submitter is not None:
            submitter = services.task_submitter
            assert isinstance(submitter, TaskSubmitter)
            return submitter
        foundation_repository = SQLiteFoundationRepository(self.storage)
        report_repository = SQLiteReportRepository(self.storage)
        task_repository = SQLiteTaskRepository(self.storage)
        data_foundation_service = DataFoundationService(foundation_repository)
        report_service = ReportService(foundation_repository, report_repository)
        ai_content_service = AIContentService(foundation_repository, AIService(), report_repository)
        result_asset_service = ResultAssetService(LocalStorageProvider(self.config.result_dir))
        task_runner = TaskRunner(
            {
                TaskType.DATA_IMPORT: DataImportExecutor(data_foundation_service),
                TaskType.REPORT_GENERATE: ReportExecutor(report_service, result_asset_service),
                TaskType.AI_CONTENT_GENERATE: AIContentExecutor(ai_content_service),
            }
        )
        submitter = TaskSubmitter(TaskService(task_repository), task_runner)
        assert isinstance(submitter, TaskSubmitter)
        return submitter

    def _task_result_service(self) -> TaskResultService:
        services = self._container_services()
        if services is not None:
            service = services.task_result
        else:
            service = TaskResultService(self._task_query_service(), self.config.result_dir)
        assert isinstance(service, TaskResultService)
        return service

    def _task_query_service(self) -> TaskQueryService:
        services = self._container_services()
        if services is not None:
            service = services.task_query
        else:
            task_repository = SQLiteTaskRepository(self.storage)
            service = TaskQueryService(task_repository)
        assert isinstance(service, TaskQueryService)
        return service

    def _permission_service(self) -> PermissionService:
        services = self._container_services()
        if services is not None:
            service = services.permissions
        else:
            service = PermissionService()
        assert isinstance(service, PermissionService)
        return service

    def _container_services(self):
        container = getattr(self, "container", None)
        if isinstance(container, ApplicationContainer):
            return container.services
        return None

    def _handle_p2_content_center_run(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            fields = self._read_urlencoded(handler)
            request = self._p2_content_request_from_fields(fields)
            synced_files = self._sync_meituan_download_sources()
            logging.info("synced %s meituan plugin files before P2 generation", len(synced_files))
            imported_count = self._ingest_meituan_plugin_files_to_foundation(user.username)
            logging.info("imported %s meituan plugin files into foundation before P2 generation", imported_count)
            product_rows = self.storage.load_meituan_foundation_rows(request.brand_id, request.platform, request.channel, "product_order")
            review_rows = self.storage.load_meituan_foundation_rows(request.brand_id, request.platform, request.channel, "service_review")
            ai_client = BailianClient(BailianSettings.from_environment())
            result = build_p2_content_pack(request, product_rows, review_rows, ai_client.generate_text)
            serial = _safe_serial()
            source_manifest_path = self.config.upload_dir / f"{serial}_p2_content_sources.json"
            result_path = self.config.result_dir / f"{serial}_p2_content_pack.csv"
            source_manifest_path.write_text(
                json.dumps(
                    {
                        "brand_id": request.brand_id,
                        "brand_name": request.brand_name,
                        "platform": request.platform,
                        "channel": request.channel,
                        "start_date": request.start_date,
                        "end_date": request.end_date,
                        "product_rows_from_foundation": len(product_rows),
                        "review_rows_from_foundation": len(review_rows),
                        "synced_plugin_files": len(synced_files),
                        "imported_foundation_files": imported_count,
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            write_csv(result_path, result.output_rows)
            job_id = self.storage.save_job(
                module=P2_CONTENT_MODULE_KEY,
                title=f"{request.brand_name}P2内容生产交付包",
                brand=request.brand_name,
                business_type="P2内容生产中心",
                created_by=user.username,
                input_file=source_manifest_path,
                result_file=result_path,
                summary=result.summary,
                warnings=result.warnings,
            )
            self._send_html(handler, self._result_page(user, job_id, result))
        except (AiGatewayError, ValidationError, ValueError, TypeError, FileNotFoundError, OSError) as exc:
            logging.error("P2 content center failed: %s", exc)
            self._send_html(handler, self._p2_content_center_page(user, str(exc)), status=400)

    def _p2_content_request_from_fields(self, fields: dict[str, list[str]]) -> P2ContentRequest:
        if not isinstance(fields, dict):
            raise TypeError("fields must be dict")
        brand_id = _form_value(fields, "brand_id", "anta_kids")
        if brand_id != "anta_kids":
            raise ValidationError("第一阶段仅接入安踏儿童，其他品牌会在基础数据稳定后扩展。")
        platform = _form_value(fields, "platform", "meituan")
        channel = _form_value(fields, "channel", "instant_retail")
        if platform != "meituan" or channel != "instant_retail":
            raise ValidationError("第一阶段仅接入美团即时零售。")
        output_count_text = _form_value(fields, "output_count", "5")
        try:
            output_count = int(output_count_text)
        except ValueError as exc:
            raise ValidationError("输出数量必须是整数。") from exc
        request = P2ContentRequest(
            brand_id=brand_id,
            brand_name="安踏儿童",
            platform=platform,
            channel=channel,
            start_date=_compact_form_date(_form_value(fields, "start_date", "")),
            end_date=_compact_form_date(_form_value(fields, "end_date", "")),
            task_type=_form_value(fields, "task_type", "social_copy"),
            output_count=output_count,
            brand_profile=ANTA_DEFAULT_BRAND_PROFILE,
            forbidden_words=DEFAULT_FORBIDDEN_WORDS,
        )
        assert request.brand_id == "anta_kids"
        return request

    def _selected_meituan_report_date(self, fields: dict[str, list[str]]) -> str:
        if not isinstance(fields, dict):
            raise TypeError("fields must be dict")
        raw_value = fields.get("report_date", [""])[0].strip()
        if not raw_value:
            raise ValidationError("请选择要生成的美团日报日期。")
        compact = raw_value.replace("-", "")
        if len(compact) != 8 or not compact.isdigit():
            raise ValidationError("美团日报日期必须是 YYYY-MM-DD 或 YYYYMMDD。")
        try:
            date(int(compact[:4]), int(compact[4:6]), int(compact[6:8]))
        except ValueError as exc:
            raise ValidationError("请选择有效的美团日报日期。") from exc
        return compact

    def _ingest_meituan_plugin_files_to_foundation(self, uploaded_by: str) -> int:
        if not isinstance(uploaded_by, str) or not uploaded_by.strip():
            raise ValueError("uploaded_by must not be empty")
        config = default_meituan_sync_config(Path.cwd())
        if not config.target_root.exists():
            logging.info("meituan plugin intake root does not exist: %s", config.target_root)
            return 0
        file_specs = {
            "product_order": ("日期", ("product_order", "商品数据")),
            "store_finance": ("开始时间", ("store_finance", "门店财务明细")),
            "store_traffic": ("开始时间", ("store_traffic", "门店流量明细")),
            "service_review": ("评价提交日期", ("service_review", "评价分析明细")),
        }
        imported_count = 0
        for path in sorted(config.target_root.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
                continue
            file_type = self._meituan_file_type_from_path(path, file_specs)
            if file_type == "":
                continue
            date_field = file_specs[file_type][0]
            try:
                rows = read_table(path.name, _BytesReader(path.read_bytes()))
                start_date, end_date = _source_date_range(rows, date_field)
                metadata = UploadMetadata(
                    business_unit="anta_retail_team",
                    brand_id="anta_kids",
                    brand_name="安踏儿童",
                    platform="meituan",
                    channel="instant_retail",
                    project_code="p1_p2_anta_meituan",
                    declared_file_type=file_type,
                    data_start_date=start_date,
                    data_end_date=end_date,
                    uploaded_by=uploaded_by,
                )
                plan = build_ingestion_plan(
                    metadata,
                    rows,
                    known_store_ids=self._known_store_ids_from_raw_rows(rows),
                    known_product_codes=self._known_product_codes_from_raw_rows(rows),
                )
                file_bytes = path.read_bytes()
                file_sha256 = hashlib.sha256(file_bytes).hexdigest()
                import_batch_id = f"plugin_{file_sha256[:24]}"
                status = "ready_for_import" if plan.validation.passed and plan.brand_match.decision == "auto_pass" else plan.brand_match.decision
                if not plan.validation.passed:
                    status = "validation_failed"
                self.storage.save_foundation_check(
                    import_batch_id=import_batch_id,
                    metadata=metadata,
                    original_file_name=path.name,
                    stored_file_path=path,
                    file_sha256=file_sha256,
                    recognized_file_type=plan.recognition.file_type,
                    row_count=len(plan.normalized_rows),
                    status=status,
                    brand_match_score=plan.brand_match.total_score,
                    validation_errors=plan.validation.errors,
                    validation_warnings=plan.validation.warnings + plan.brand_match.warnings,
                )
                if status == "ready_for_import":
                    self.storage.save_foundation_fact_rows(import_batch_id, plan)
                    imported_count += 1
            except (ValidationError, ValueError, TypeError, OSError) as exc:
                logging.info("skip meituan plugin foundation import %s: %s", path, exc)
        return imported_count

    def _sync_meituan_download_sources(self) -> list[SyncedFile]:
        configs = self._meituan_download_sync_configs(Path.cwd())
        synced: list[SyncedFile] = []
        for config in configs:
            synced.extend(copy_new_files(config))
        logging.info("synced %s files from %s meituan download sources", len(synced), len(configs))
        return synced

    @staticmethod
    def _meituan_download_sync_configs(project_root: Path) -> tuple[SyncConfig, ...]:
        if not isinstance(project_root, Path):
            raise TypeError("project_root must be Path")
        plugin_config = default_meituan_sync_config(project_root)
        downloads_config = SyncConfig(
            source_root=Path.home() / "Downloads",
            target_root=plugin_config.target_root,
            index_path=plugin_config.index_path,
            file_name_keywords=MEITUAN_REPORT_KEYWORDS,
            excluded_dir_names=("meituan_auto_download",),
            structure_meituan_reports=True,
        )
        result = (plugin_config, downloads_config)
        assert len(result) == 2
        return result

    def _meituan_file_type_from_path(self, path: Path, file_specs: dict[str, tuple[str, tuple[str, ...]]]) -> str:
        if not isinstance(path, Path):
            raise TypeError("path must be Path")
        normalized_path = str(path).lower()
        for file_type, (_date_field, tokens) in file_specs.items():
            if any(token.lower() in normalized_path for token in tokens):
                return file_type
        return ""

    def _known_store_ids_from_raw_rows(self, rows: list[dict[str, str]]) -> tuple[str, ...]:
        if not isinstance(rows, list):
            raise TypeError("rows must be list")
        values: list[str] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            for field_name in ("店铺ID", "商家ID"):
                value = str(row.get(field_name, "")).strip()
                if value:
                    values.append(value)
        return tuple(dict.fromkeys(values))

    def _known_product_codes_from_raw_rows(self, rows: list[dict[str, str]]) -> tuple[str, ...]:
        if not isinstance(rows, list):
            raise TypeError("rows must be list")
        values: list[str] = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            for field_name in ("商品SKU码", "UPC码"):
                value = str(row.get(field_name, "")).strip()
                if value:
                    values.append(value)
        return tuple(dict.fromkeys(values))

    def _handle_anta_blacklist_run(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        scenario = self.scenarios[ANTA_RETAIL_KEY]
        try:
            form_fields, uploaded_files = self._read_multipart_files(handler)
            required_names = {
                "blacklist_file": "黑名单明细",
                "meituan_product_file": "美团商品表",
                "jd_product_file": "京东商品表",
                "store_summary_file": "门店信息汇总",
            }
            missing_names = [label for field_name, label in required_names.items() if field_name not in uploaded_files]
            if missing_names:
                raise ValidationError("请上传：" + "、".join(missing_names))
            serial = _safe_serial()
            input_dir = self.config.upload_dir / f"{serial}_anta_blacklist"
            output_dir = self.config.result_dir / f"{serial}_anta_blacklist"
            input_dir.mkdir(parents=True, exist_ok=True)
            saved_paths = {
                field_name: self._save_uploaded_file(input_dir, field_name, uploaded_file)
                for field_name, uploaded_file in uploaded_files.items()
            }
            result = self._run_external_anta_blacklist(saved_paths, output_dir)
            summary = {str(row["指标"]): str(row["数值"]) for row in result.summary}
            warnings = self._anta_blacklist_warnings(result.tables, "old_meituan_blacklist_file" in uploaded_files)
            title = form_fields.get("title", "安踏即时零售黑名单筛选").strip() or "安踏即时零售黑名单筛选"
            job_id = self.storage.save_job(
                module=scenario.key,
                title=title,
                brand=scenario.brand,
                business_type=scenario.business_type,
                created_by=user.username,
                input_file=input_dir,
                result_file=result.workbook_path,
                summary=summary,
                warnings=warnings,
            )
            processing_result = ProcessingResult(module=scenario.key, output_rows=[], summary=summary, warnings=warnings)
            self._send_html(handler, self._result_page(user, job_id, processing_result))
        except (ValidationError, ValueError, TypeError, FileNotFoundError, ImportError) as exc:
            logging.error("anta blacklist processing failed: %s", exc)
            self._send_html(handler, self._anta_retail_page(user, str(exc), self._anta_retail_url(handler)), status=400)

    def _handle_archive_intake_run(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            result = run_archive_intake(ArchiveIntakeConfig(self.config.template_root))
            self._send_html(handler, self._archive_intake_result_page(user, result))
        except (ValueError, TypeError, OSError) as exc:
            logging.error("archive intake failed: %s", exc)
            self._send_html(handler, self._archive_intake_page(user, str(exc)), status=400)

    def _handle_archive_catalog_rebuild(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            result = rebuild_archive_catalog(ArchiveIntakeConfig(self.config.template_root))
            self._send_html(handler, self._archive_intake_result_page(user, result, "本地资料台账已刷新"))
        except (ValueError, TypeError, OSError) as exc:
            logging.error("archive catalog rebuild failed: %s", exc)
            self._send_html(handler, self._archive_intake_page(user, str(exc)), status=400)

    def _handle_archive_intake_upload(self, handler: BaseHTTPRequestHandler, user: UserRecord) -> None:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            uploaded_files = self._read_file_list(handler)
            if not uploaded_files:
                raise ValidationError("请至少选择一个文件")
            pending_dir = self.config.template_root / "00_intake" / "01_pending"
            pending_dir.mkdir(parents=True, exist_ok=True)
            for uploaded_file in uploaded_files:
                self._save_intake_file(pending_dir, uploaded_file)
            result = run_archive_intake(ArchiveIntakeConfig(self.config.template_root))
            self._send_html(handler, self._archive_intake_result_page(user, result))
        except (ValidationError, ValueError, TypeError, OSError) as exc:
            logging.error("archive upload failed: %s", exc)
            self._send_html(handler, self._archive_intake_page(user, str(exc)), status=400)

    def _build_anta_weekly_report(self) -> tuple[ProcessingResult, Path]:
        raw_dir = self.config.template_root / "01_data_processing" / "01-3_weekly_report" / "anta_weekly_report" / "01_raw_data"
        meituan_path = _latest_matching_file(raw_dir, ("meituan",), (".xlsx", ".csv"))
        jd_path = _latest_matching_file(raw_dir, ("jd",), (".xlsx", ".csv"))
        meituan_rows = read_table(meituan_path.name, _BytesReader(meituan_path.read_bytes()))
        jd_rows = read_table(jd_path.name, _BytesReader(jd_path.read_bytes()))
        result = anta_reporting.build_weekly_report(
            anta_reporting.ReportSource(name=meituan_path.name, rows=meituan_rows),
            anta_reporting.ReportSource(name=jd_path.name, rows=jd_rows),
        )
        return result, raw_dir

    def _build_anta_monthly_report(self) -> tuple[ProcessingResult, Path]:
        raw_dir = self.config.template_root / "01_data_processing" / "01-4_monthly_report" / "anta_monthly_report" / "01_raw_data"
        product_path = _latest_matching_file(raw_dir, ("product_data",), (".xlsx", ".csv"))
        store_path = _latest_matching_file(raw_dir, ("store_info_summary",), (".xlsx", ".csv"))
        finance_path = _latest_matching_file(raw_dir, ("store_finance_details",), (".xlsx", ".csv"))
        product_rows = read_table(product_path.name, _BytesReader(product_path.read_bytes()))
        store_rows = read_table(store_path.name, _BytesReader(store_path.read_bytes()))
        finance_rows = read_table(finance_path.name, _BytesReader(finance_path.read_bytes()))
        result = anta_reporting.build_monthly_report(
            anta_reporting.ReportSource(name=product_path.name, rows=product_rows),
            anta_reporting.ReportSource(name=store_path.name, rows=store_rows),
            anta_reporting.ReportSource(name=finance_path.name, rows=finance_rows),
        )
        return result, raw_dir

    def _load_anta_meituan_sources(
        self,
        report_type: str,
    ) -> tuple[anta_meituan_reporting.MeituanReportSources, dict[str, LocalReportSource]]:
        if report_type not in {"daily", "weekly"}:
            raise ValueError("report_type must be daily or weekly")
        source_roots = (
            self.config.upload_dir / "meituan_auto_download",
            self.config.upload_dir / "data_foundation",
            Path.cwd() / "intranet_app" / "runtime" / "intake" / "meituan_auto_download",
            Path.home() / "Downloads",
            self.config.template_root / "01_data_processing" / "01-2_daily_report" / "anta" / "01_raw_data",
            self.config.template_root / "01_data_processing" / "01-3_weekly_report" / "anta_weekly_report" / "01_raw_data",
        )
        selected = {
            "product": self._select_meituan_source("product", ("商品数据", "product_order"), "日期", report_type, source_roots),
            "finance": self._select_meituan_source("finance", ("门店财务明细", "store_finance"), "开始时间", report_type, source_roots),
            "traffic": self._select_meituan_source("traffic", ("门店流量明细", "store_traffic"), "开始时间", report_type, source_roots),
            "review": self._select_meituan_source("review", ("评价分析明细", "service_review"), "评价提交日期", report_type, source_roots),
        }
        product_window = selected["product"]
        for kind in ("finance", "traffic", "review"):
            current = selected[kind]
            if current.end_date < product_window.start_date or current.start_date > product_window.end_date:
                raise ValidationError(f"{kind} 文件日期范围与商品数据不重叠：{current.path.name}")
        source_for_calculation = selected
        if report_type == "daily":
            source_for_calculation = self._daily_sources_with_lookback(selected, source_roots)
        sources = anta_meituan_reporting.MeituanReportSources(
            product_rows=source_for_calculation["product"].rows,
            finance_rows=source_for_calculation["finance"].rows,
            traffic_rows=source_for_calculation["traffic"].rows,
            review_rows=source_for_calculation["review"].rows,
        )
        manifest_sources = dict(selected)
        for key, value in source_for_calculation.items():
            if value.path != selected[key].path:
                manifest_sources[f"{key}_lookback"] = value
        return sources, manifest_sources

    def _load_anta_meituan_sources_from_foundation(
        self,
        report_type: str,
        selected_report_date: str = "",
    ) -> tuple[anta_meituan_reporting.MeituanReportSources, dict[str, LocalReportSource]]:
        if report_type not in {"daily", "weekly"}:
            raise ValueError("report_type must be daily or weekly")
        brand_id = "anta_kids"
        platform = "meituan"
        channel = "instant_retail"
        product_rows = self.storage.load_meituan_foundation_rows(brand_id, platform, channel, "product_order")
        if not product_rows:
            raise ValidationError("基础数据层缺少安踏美团商品订单数据。请先在数据入库中心完成入库，日报/周报不允许直接读取下载原始文件。")
        finance_rows = self.storage.load_meituan_foundation_rows(brand_id, platform, channel, "store_finance")
        traffic_rows = self.storage.load_meituan_foundation_rows(brand_id, platform, channel, "store_traffic")
        review_rows = self.storage.load_meituan_foundation_rows(brand_id, platform, channel, "service_review")
        product_dates = sorted(
            {
                _compact_date_from_source(row.get("下单时间", ""))
                for row in product_rows
                if _compact_date_from_source(row.get("下单时间", ""))
            }
        )
        if not product_dates:
            raise ValidationError("基础数据层商品订单数据缺少有效下单日期，无法生成安踏美团日报/周报。")
        if report_type == "daily":
            target_date = selected_report_date.strip()
            if len(target_date) != 8 or not target_date.isdigit():
                raise ValidationError("请选择要生成的美团日报日期。")
            if target_date not in product_dates:
                raise ValidationError(
                    f"基础数据层缺少安踏美团 {target_date} 商品订单数据。请先在美团后台打开报表页，用浏览器插件导出 {target_date} 的商品、财务、流量数据；导出后再选择该日期生成日报，系统会自动同步、入库并生成。"
                )
            product_start_date = target_date
            product_end_date = target_date
        else:
            product_start_date = product_dates[0]
            product_end_date = product_dates[-1]
        sources = anta_meituan_reporting.MeituanReportSources(
            product_rows=product_rows,
            finance_rows=finance_rows,
            traffic_rows=traffic_rows,
            review_rows=review_rows,
        )
        selected_files = {
            "product": LocalReportSource("product", Path("foundation_fact_order_product"), product_rows, product_start_date, product_end_date),
            "finance": self._foundation_local_source("finance", "foundation_fact_store_finance", finance_rows, "开始时间", product_start_date, product_end_date),
            "traffic": self._foundation_local_source("traffic", "foundation_fact_store_traffic", traffic_rows, "开始时间", product_start_date, product_end_date),
            "review": self._foundation_local_source("review", "foundation_fact_service_review", review_rows, "评价提交日期", product_start_date, product_end_date),
        }
        return sources, selected_files

    def _foundation_local_source(
        self,
        kind: str,
        source_name: str,
        rows: list[dict[str, str]],
        date_field: str,
        fallback_start_date: str,
        fallback_end_date: str,
    ) -> LocalReportSource:
        for field_name, field_value in (
            ("kind", kind),
            ("source_name", source_name),
            ("date_field", date_field),
            ("fallback_start_date", fallback_start_date),
            ("fallback_end_date", fallback_end_date),
        ):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        if not isinstance(rows, list):
            raise TypeError("rows must be list")
        dates = sorted(
            {
                _compact_date_from_source(row.get(date_field, ""))
                for row in rows
                if _compact_date_from_source(row.get(date_field, ""))
            }
        )
        start_date = dates[0] if dates else fallback_start_date
        end_date = dates[-1] if dates else fallback_end_date
        return LocalReportSource(kind, Path(source_name), rows, start_date, end_date)

    def _daily_sources_with_lookback(
        self,
        selected_daily_sources: dict[str, LocalReportSource],
        source_roots: tuple[Path, ...],
    ) -> dict[str, LocalReportSource]:
        if not isinstance(selected_daily_sources, dict) or "product" not in selected_daily_sources:
            raise ValueError("selected_daily_sources must include product")
        report_date = selected_daily_sources["product"].end_date
        lookback_specs = {
            "product": (("商品数据", "product_order"), "日期"),
            "finance": (("门店财务明细", "store_finance"), "开始时间"),
            "traffic": (("门店流量明细", "store_traffic"), "开始时间"),
            "review": (("评价分析明细", "service_review"), "评价提交日期"),
        }
        result = dict(selected_daily_sources)
        for kind, (tokens, date_field) in lookback_specs.items():
            try:
                candidate = self._select_meituan_source(kind, tokens, date_field, "weekly", source_roots)
            except (FileNotFoundError, ValidationError, ValueError):
                continue
            if candidate.start_date <= report_date <= candidate.end_date and candidate.end_date <= report_date:
                result[kind] = candidate
        return result

    def _select_meituan_source(
        self,
        kind: str,
        name_tokens: tuple[str, ...],
        date_field: str,
        report_type: str,
        source_roots: tuple[Path, ...],
    ) -> LocalReportSource:
        if report_type not in {"daily", "weekly"}:
            raise ValueError("report_type must be daily or weekly")
        if not isinstance(source_roots, tuple) or not source_roots:
            raise ValueError("source_roots must not be empty")
        candidates: list[LocalReportSource] = []
        for root in source_roots:
            if not root.exists():
                continue
            for path in root.rglob("*"):
                if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
                    continue
                if not any(token.lower() in path.name.lower() for token in name_tokens):
                    continue
                try:
                    rows = read_table(path.name, _BytesReader(path.read_bytes()))
                    start_date, end_date = _source_date_range(rows, date_field)
                    candidates.append(LocalReportSource(kind=kind, path=path, rows=rows, start_date=start_date, end_date=end_date))
                except (ValidationError, ValueError, TypeError, OSError) as exc:
                    logging.info("skip unusable meituan source %s: %s", path, exc)
        if not candidates:
            raise FileNotFoundError(f"未找到美团{kind}源文件，请先用插件下载或同步文件。")
        if report_type == "daily":
            eligible = [item for item in candidates if item.start_date == item.end_date]
        else:
            eligible = [item for item in candidates if item.start_date < item.end_date]
        if not eligible:
            eligible = candidates
        selected = sorted(eligible, key=lambda item: (item.end_date, item.path.stat().st_mtime), reverse=True)[0]
        logging.info("selected meituan %s source: %s", kind, selected.path)
        return selected

    def _download_job_result(self, handler: BaseHTTPRequestHandler, path: str) -> None:
        parts = [part for part in path.split("/") if part]
        if len(parts) != 3:
            self._send_html(handler, self._page("下载失败", "<p>任务地址不正确。</p>"), status=400)
            return
        try:
            job_id = int(parts[1])
        except ValueError:
            self._send_html(handler, self._page("下载失败", "<p>任务编号不正确。</p>"), status=400)
            return
        job = self.storage.get_job(job_id)
        if job is None:
            self._send_html(handler, self._page("下载失败", "<p>任务不存在。</p>"), status=404)
            return
        result_path = Path(job.result_file)
        if not result_path.exists():
            self._send_html(handler, self._page("下载失败", "<p>结果文件不存在。</p>"), status=404)
            return
        self._send_file(handler, result_path, download_name=result_path.name)

    def _context(self, handler: BaseHTTPRequestHandler) -> RequestContext:
        raw_cookie = handler.headers.get("Cookie", "")
        parsed = cookies.SimpleCookie(raw_cookie)
        token = parsed.get("intranet_session")
        if token is None:
            return RequestContext(user=None, token=None)
        value = token.value
        user = self.storage.get_session_user(value)
        return RequestContext(user=user, token=value)

    def _read_urlencoded(self, handler: BaseHTTPRequestHandler) -> dict[str, list[str]]:
        length = int(handler.headers.get("Content-Length", "0"))
        body = handler.rfile.read(length).decode("utf-8")
        return parse_qs(body)

    def _read_json(self, handler: BaseHTTPRequestHandler) -> dict[str, object]:
        length = int(handler.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValueError("request body must not be empty")
        body = handler.rfile.read(length).decode("utf-8")
        payload = json.loads(body)
        if not isinstance(payload, dict):
            raise ValueError("JSON body must be object")
        return payload

    def _read_multipart(self, handler: BaseHTTPRequestHandler) -> tuple[dict[str, str], UploadedFile]:
        content_type = handler.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            raise ValidationError("请使用文件上传表单提交")
        length = int(handler.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValidationError("上传内容不能为空")
        body = handler.rfile.read(length)
        message = BytesParser(policy=policy.default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
        )
        fields: dict[str, str] = {}
        uploaded_file: UploadedFile | None = None
        for part in message.iter_parts():
            name = part.get_param("name", header="content-disposition")
            if not name:
                continue
            file_name = part.get_filename()
            payload = part.get_payload(decode=True) or b""
            if file_name:
                uploaded_file = UploadedFile(file_name=file_name, content=payload)
            else:
                fields[str(name)] = payload.decode("utf-8", errors="replace").strip()
        if uploaded_file is None:
            raise ValidationError("请上传业务数据文件")
        return fields, uploaded_file

    def _read_multipart_files(self, handler: BaseHTTPRequestHandler) -> tuple[dict[str, str], dict[str, UploadedFile]]:
        content_type = handler.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            raise ValidationError("请使用文件上传表单提交")
        length = int(handler.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValidationError("上传内容不能为空")
        body = handler.rfile.read(length)
        message = BytesParser(policy=policy.default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
        )
        fields: dict[str, str] = {}
        uploaded_files: dict[str, UploadedFile] = {}
        for part in message.iter_parts():
            name = part.get_param("name", header="content-disposition")
            if not name:
                continue
            payload = part.get_payload(decode=True) or b""
            file_name = part.get_filename()
            if file_name and payload:
                uploaded_files[str(name)] = UploadedFile(file_name=file_name, content=payload)
            elif not file_name:
                fields[str(name)] = payload.decode("utf-8", errors="replace").strip()
        if not isinstance(uploaded_files, dict):
            raise AssertionError("uploaded_files must be dict")
        return fields, uploaded_files

    def _read_file_list(self, handler: BaseHTTPRequestHandler) -> list[UploadedFile]:
        content_type = handler.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            raise ValidationError("请使用文件上传表单提交")
        length = int(handler.headers.get("Content-Length", "0"))
        if length <= 0:
            raise ValidationError("上传内容不能为空")
        body = handler.rfile.read(length)
        message = BytesParser(policy=policy.default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode("utf-8") + body
        )
        uploaded_files: list[UploadedFile] = []
        for part in message.iter_parts():
            file_name = part.get_filename()
            payload = part.get_payload(decode=True) or b""
            if file_name and payload:
                uploaded_files.append(UploadedFile(file_name=file_name, content=payload))
        return uploaded_files

    def _save_uploaded_file(self, input_dir: Path, field_name: str, uploaded_file: UploadedFile) -> Path:
        if not isinstance(input_dir, Path):
            raise TypeError("input_dir must be pathlib.Path")
        if not field_name.strip():
            raise ValueError("field_name must not be empty")
        if not isinstance(uploaded_file, UploadedFile):
            raise TypeError("uploaded_file must be UploadedFile")
        file_path = input_dir / f"{field_name}_{_safe_name(uploaded_file.file_name)}"
        file_path.write_bytes(uploaded_file.content)
        if not file_path.exists():
            raise AssertionError(f"failed to save uploaded file: {file_path}")
        return file_path

    def _save_intake_file(self, pending_dir: Path, uploaded_file: UploadedFile) -> Path:
        if not isinstance(pending_dir, Path):
            raise TypeError("pending_dir must be pathlib.Path")
        if not isinstance(uploaded_file, UploadedFile):
            raise TypeError("uploaded_file must be UploadedFile")
        file_path = _unique_upload_path(pending_dir / _safe_name(uploaded_file.file_name))
        file_path.write_bytes(uploaded_file.content)
        if not file_path.exists():
            raise AssertionError(f"failed to save intake file: {file_path}")
        return file_path

    def _run_external_anta_blacklist(self, saved_paths: dict[str, Path], output_dir: Path) -> object:
        if not isinstance(saved_paths, dict):
            raise TypeError("saved_paths must be dict[str, Path]")
        if not isinstance(output_dir, Path):
            raise TypeError("output_dir must be pathlib.Path")
        from .anta_retail_launcher import DEFAULT_PROJECT_ROOT

        external_src = DEFAULT_PROJECT_ROOT / "src"
        if not external_src.exists():
            raise FileNotFoundError(f"外部安踏项目源码不存在：{external_src}")
        sys.path.insert(0, str(external_src))
        from anta_listing_checker.blacklist_matcher import BlacklistMatchConfig, BlacklistSourceFiles, run_blacklist_match

        source_files = BlacklistSourceFiles(
            blacklist_path=saved_paths["blacklist_file"],
            meituan_product_path=saved_paths["meituan_product_file"],
            jd_product_path=saved_paths["jd_product_file"],
            store_summary_path=saved_paths["store_summary_file"],
            old_meituan_blacklist_path=saved_paths.get("old_meituan_blacklist_file"),
            delivery_reference_root=None,
        )
        result = run_blacklist_match(source_files, BlacklistMatchConfig(), output_dir)
        result.validate()
        return result

    def _anta_blacklist_warnings(self, tables: object, has_old_meituan_file: bool) -> list[str]:
        if not isinstance(tables, dict):
            raise TypeError("tables must be dict")
        if not isinstance(has_old_meituan_file, bool):
            raise TypeError("has_old_meituan_file must be bool")
        warnings: list[str] = []
        anomaly_count = len(tables.get("异常数据", ()))
        if anomaly_count > 0:
            warnings.append(f"存在 {anomaly_count} 条异常数据，请打开结果 Excel 的“异常数据”页复核。")
        if not has_old_meituan_file:
            warnings.append("未上传旧版美团黑名单，系统会生成新增结果，但无法准确判断美团释放清单。")
        if not warnings:
            warnings.append("无")
        return warnings

    def _dashboard(self, user: UserRecord) -> str:
        priority_cards = "".join(self._priority_nav_card(priority, title, description) for priority, title, description in PRIORITY_SECTIONS)
        group_development_tree = self._group_development_tree_panel()
        developable_total = self._developable_project_count()
        job_rows = "".join(self._job_row(job) for job in self.storage.list_jobs())
        if not job_rows:
            job_rows = "<tr><td colspan='7'>暂无处理记录</td></tr>"
        body = f"""
        <section class="toolbar">
          <div>
            <h1>内网自动化工作台</h1>
            <p>当前登录：{_e(user.display_name)} · { _e(user.role) }</p>
          </div>
          <a class="button secondary" href="/logout">退出</a>
        </section>
        <section class="dashboard-hero">
          <div>
            <h2>数据看板</h2>
            <p>集中查看项目覆盖、品牌接入、资料沉淀和处理记录，再按 P1-P4 进入分级页。</p>
          </div>
          <div class="button-row">
            <a class="button" href="/data-foundation">数据入库中心</a>
            <a class="button" href="/automation-runs">自动化数据执行</a>
            <a class="button" href="/archive-intake">投递资料</a>
            <a class="button" href="/work-item-planning">{developable_total}个可提效项目明细</a>
          </div>
        </section>
        <h2 class="section-title">P1-P4 分级入口</h2>
        <section class="priority-nav priority-nav-lead">{priority_cards}</section>
        {group_development_tree}
        <section>
          <h2>最近处理记录</h2>
          <table>
            <thead><tr><th>编号</th><th>模块</th><th>品牌</th><th>类型</th><th>提交人</th><th>时间</th><th>结果</th></tr></thead>
            <tbody>{job_rows}</tbody>
          </table>
        </section>
        """
        return self._page("内网自动化工作台", body)

    def _group_development_tree_panel(self) -> str:
        dashboard_metrics = self._dashboard_metrics()
        metric_cards = "".join(
            f"""
            <article>
              <span class="summary-label">{_e(label)}</span>
              <strong>{_e(value)}</strong>
            </article>
            """
            for label, value in dashboard_metrics
        )
        tree_sections = "".join(
            self._group_priority_tree(priority, feedback_by_project)
            for priority, _, _ in PRIORITY_SECTIONS
        )
        feedback_count = sum(
            1
            for item in GROUP_PROJECT_TREE_ITEMS
            if item.project in feedback_by_project and feedback_by_project[item.project].business_feedback.strip()
        )
        filled_current_time_count = sum(
            1
            for item in GROUP_PROJECT_TREE_ITEMS
            if item.project in feedback_by_project and feedback_by_project[item.project].current_processing_time.strip()
        )
        total_original_hours = sum((item.original_hours for item in GROUP_PROJECT_TREE_ITEMS), Decimal("0"))
        summary_html = f"""
        <div class="development-stat neutral"><span>全组项目</span><strong>{len(GROUP_PROJECT_TREE_ITEMS)}</strong></div>
        <div class="development-stat complete"><span>原人工耗时</span><strong>{_format_decimal(total_original_hours)} 小时/月</strong></div>
        <div class="development-stat time"><span>已填处理耗时（已填现处理时间）</span><strong>{filled_current_time_count}</strong></div>
        <div class="development-stat feedback"><span>已收业务反馈</span><strong>{feedback_count}</strong></div>
        """
        body = f"""
        <section class="dashboard-panel group-development-panel">
          <div class="dashboard-panel-header">
            <div>
              <h2>开发覆盖总览</h2>
              <p class="note">全组项目开发与覆盖总览</p>
              <p>按 P1-P4 展开全组工作项目，在同一处维护优先级、原人工耗时、现处理时间、提升比例和业务反馈。</p>
            </div>
          </div>
          <section class="ledger-summary">{metric_cards}</section>
          <section class="development-stat-strip group-stat-strip">{summary_html}</section>
          <section class="group-tree">{tree_sections}</section>
        </section>
        """
        assert "全组项目开发与覆盖总览" in body
        assert "group-tree" in body
        return body

    def _group_priority_tree(self, priority: str, feedback_by_project: dict[str, object]) -> str:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        items = [item for item in GROUP_PROJECT_TREE_ITEMS if item.priority == priority]
        if not items:
            return ""
        priority_meta = next(item for item in PRIORITY_SECTIONS if item[0] == priority)
        branches = tuple(dict.fromkeys(item.branch for item in items))
        branch_html = "".join(
            self._group_branch_tree(priority, branch, [item for item in items if item.branch == branch], feedback_by_project)
            for branch in branches
        )
        hours = sum((item.original_hours for item in items), Decimal("0"))
        open_attribute = " open" if priority in {"P1", "P2"} else ""
        result = f"""
        <details class="group-priority-node priority-border-{_e(priority.lower())}"{open_attribute}>
          <summary>
            <span class="priority-badge priority-{_e(priority.lower())}">{_e(priority)}</span>
            <strong>{_e(priority_meta[1])}</strong>
            <em>{len(items)} 个项目 · {_format_decimal(hours)} 小时/月</em>
          </summary>
          <div class="group-branch-list">{branch_html}</div>
        </details>
        """
        assert priority in result
        return result

    def _group_branch_tree(
        self,
        priority: str,
        branch: str,
        items: list[GroupProjectTreeItem],
        feedback_by_project: dict[str, object],
    ) -> str:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        if not isinstance(branch, str) or not branch.strip():
            raise ValueError("branch must not be empty")
        if not isinstance(items, list) or not items:
            raise ValueError("items must not be empty")
        rows = "".join(self._group_project_feedback_form(item, feedback_by_project) for item in items)
        hours = sum((item.original_hours for item in items), Decimal("0"))
        result = f"""
        <details class="group-branch-node" open>
          <summary>
            <span>{_e(branch)}</span>
            <em>{len(items)} 项 · {_format_decimal(hours)} 小时/月</em>
          </summary>
          <div class="group-project-list">{rows}</div>
        </details>
        """
        assert branch in result
        return result

    def _group_project_feedback_form(
        self,
        item: GroupProjectTreeItem,
        feedback_by_project: dict[str, object],
    ) -> str:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        record = feedback_by_project.get(item.project)
        current_time = record.current_processing_time if record is not None else ""
        business_feedback = record.business_feedback if record is not None else ""
        iteration_need = record.iteration_need if record is not None else ""
        improvement = self._format_efficiency_gain(item.original_hours, current_time)
        feedback_status = "已反馈" if business_feedback.strip() else "待反馈"
        feedback_class = "complete" if business_feedback.strip() else "muted"
        result = f"""
        <form method="post" action="/project-stages/feedback" class="group-project-row">
          <input type="hidden" name="return_to" value="/">
          <input type="hidden" name="project" value="{_e(item.project)}">
          <div class="tree-cell project-cell">
            <span class="tree-label">项目</span>
            <strong>{_e(item.project)}</strong>
            <small>{_e(item.brand)} · {_e(item.business_type)}</small>
          </div>
          <div class="tree-cell">
            <span class="tree-label">优先级</span>
            <span class="priority-badge priority-{_e(item.priority.lower())}">{_e(item.priority)}</span>
          </div>
          <div class="tree-cell">
            <span class="tree-label">原人工耗时</span>
            <strong>{_format_decimal(item.original_hours)} 小时/月</strong>
            <small>{_e(item.source_detail)}</small>
          </div>
          <label class="tree-cell">
            <span class="tree-label">现处理时间</span>
            <input name="current_processing_time" maxlength="100" value="{_e(current_time)}" placeholder="如：40分钟/次 或 6小时/月">
          </label>
          <div class="tree-cell">
            <span class="tree-label">提升比例</span>
            <strong>{_e(improvement)}</strong>
          </div>
          <label class="tree-cell feedback-tree-cell">
            <span class="tree-label">业务反馈</span>
            <textarea name="business_feedback" maxlength="2000" placeholder="业务试用后的效果、问题或建议">{_e(business_feedback)}</textarea>
          </label>
          <input type="hidden" name="iteration_need" value="{_e(iteration_need)}">
          <div class="tree-cell tree-actions">
            <span class="status-pill {feedback_class}">{feedback_status}</span>
            <button class="button save-feedback-button" type="submit">保存</button>
          </div>
        </form>
        """
        assert item.project in result
        return result

    def _project_feedback_names(self) -> set[str]:
        result = {row[2] for row in PROJECT_STAGE_ROWS}
        result.update(item.project for item in GROUP_PROJECT_TREE_ITEMS)
        result.update(item.feedback_key for item in self._completed_feedback_items_for_dashboard())
        assert result
        return result

    def _format_efficiency_gain(self, original_hours: Decimal, current_processing_time: str) -> str:
        if not isinstance(original_hours, Decimal):
            raise TypeError("original_hours must be Decimal")
        if not isinstance(current_processing_time, str):
            raise TypeError("current_processing_time must be str")
        if original_hours <= 0 or not current_processing_time.strip():
            return "待计算"
        current_hours = _parse_duration_hours(current_processing_time)
        if current_hours is None or current_hours < 0:
            return "待计算"
        if current_hours == 0:
            return "100%"
        gain = (Decimal("1") - (current_hours / original_hours)) * Decimal("100")
        if gain < Decimal("0"):
            return "未提升"
        return f"{gain.quantize(Decimal('0.1'))}%"

    @staticmethod
    def _format_time_saved(original_hours: Decimal, current_processing_time: str) -> str:
        if not isinstance(original_hours, Decimal):
            raise TypeError("original_hours must be Decimal")
        if not isinstance(current_processing_time, str):
            raise TypeError("current_processing_time must be str")
        if original_hours <= 0 or not current_processing_time.strip():
            return "待计算"
        current_hours = _parse_duration_hours(current_processing_time)
        if current_hours is None or current_hours < 0:
            return "待计算"
        saved_hours = original_hours - current_hours
        if saved_hours < 0:
            return "未提效"
        rounded = saved_hours.quantize(Decimal("0.01")).normalize()
        return f"{rounded}小时"

    def _group_development_tree_panel(self) -> str:
        feedback_by_project = self.storage.list_project_feedback()
        dashboard_metrics = self._dashboard_metrics()
        metric_cards = "".join(
            f"""
            <article>
              <span class="summary-label">{_e(label)}</span>
              <strong>{_e(value)}</strong>
            </article>
            """
            for label, value in dashboard_metrics
        )
        completed_feedback_table = self._completed_feedback_dashboard_table(feedback_by_project)
        efficiency_mapping = self._improved_efficiency_panel()
        body = f"""
        <section class="dashboard-panel group-development-panel">
          <div class="dashboard-panel-header project-map-header">
            <div>
              <h2>开发覆盖总览</h2>
              <p class="note">全组项目开发与覆盖总览</p>
              <p>合并展示可提效项目覆盖和已开发反馈汇总；置灰 C 类人工协同项目不纳入首页总数，反馈数据来自二级页面“已开发完整业务反馈”。</p>
            </div>
          </div>
          <section class="ledger-summary">{metric_cards}</section>
          {completed_feedback_table}
          {efficiency_mapping}
        </section>
        """
        assert "全组项目开发与覆盖总览" in body
        assert "已开发反馈汇总" in body
        assert "已提效项目" in body
        return body

    def _work_item_planning_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        map_lanes = "".join(self._project_map_priority_lane(priority) for priority, _, _ in PRIORITY_SECTIONS)
        modals = "".join(
            self._project_detail_modal(item, index)
            for index, item in enumerate(GROUP_PROJECT_TREE_ITEMS, start=1)
        )
        developable_items = self._developable_group_items()
        total_original_hours = sum((item.original_hours for item in developable_items), Decimal("0"))
        developable_total = len(developable_items)
        body = f"""
        <section class="toolbar">
          <div>
            <h1>{developable_total}个可提效项目明细</h1>
            <p>来自《内容任务耗时统计.xlsx》的全组工作项目；A/B 计入可提效项目，C 类置灰展示但不归入首页统计。</p>
          </div>
          <div class="toolbar-actions">
            <a class="button" href="/efficiency-mapping">优先开发</a>
            <a class="button secondary" href="/">返回数据看板</a>
          </div>
        </section>
        <section class="dashboard-panel group-development-panel">
          <section class="project-map-summary">
            <article><span>可提效项目</span><strong>{developable_total}</strong><small>排除 C 类置灰项目</small></article>
            <article><span>原人工耗时</span><strong>{_format_decimal(total_original_hours)}</strong><small>小时/月</small></article>
            <article><span>展示层级</span><strong>P1-P4</strong><small>按优先级和业务工作流聚合</small></article>
            <article><span>入口层级</span><strong>二级页</strong><small>首页只保留渠道规划图</small></article>
          </section>
          <section class="project-map-grid">{map_lanes}</section>
          {modals}
        </section>
        """
        assert "可提效项目明细" in body
        assert "project-map-grid" in body
        return self._page("可提效项目明细", body)

    def _efficiency_mapping_page(self, user: UserRecord, message: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(message, str):
            raise TypeError("message must be str")
        alert = f"<div class='success'>{_e(message)}</div>" if message and "已" in message else f"<div class='error'>{_e(message)}</div>" if message else ""
        body = f"""
        <section class="toolbar">
          <div>
            <h1>高耗时任务映射台账</h1>
            <p>完整维护待提效品牌、未提效原因、排期规划和手动补充品牌；首页只展示已提效项目。</p>
          </div>
          <div class="toolbar-actions">
            <a class="button secondary" href="/">返回数据看板</a>
          </div>
        </section>
        {alert}
        {self._manual_efficiency_brand_form()}
        {self._high_efficiency_mapping_panel()}
        """
        assert "高耗时任务映射台账" in body
        assert "手动增加品牌" in body
        return self._page("高耗时任务映射台账", body)

    def _improved_efficiency_panel(self) -> str:
        all_items = self._high_efficiency_mapping_items()
        improved_items = tuple(
            EfficiencyMappingTask(
                priority=item.priority,
                replacement_type=item.replacement_type,
                task_name=item.task_name,
                monthly_hours=item.monthly_hours,
                replacement_reason=item.replacement_reason,
                brands=tuple(brand for brand in item.brands if brand.is_improved),
            )
            for item in all_items
            if any(brand.is_improved for brand in item.brands)
        )
        cards = "".join(self._improved_efficiency_card(item) for item in improved_items)
        result = f"""
        <section class="efficiency-mapping-panel compact-efficiency-panel">
          <div class="section-heading split-heading">
            <div>
              <h2>已提效项目</h2>
              <p>只展示已提效任务与标绿品牌；在完整台账中把品牌标记为“已提效”后，这里会自动联动展示。</p>
            </div>
            <a class="button secondary" href="/efficiency-mapping">优先开发</a>
          </div>
          <div class="improved-efficiency-grid">{cards}</div>
        </section>
        """
        assert "已提效项目" in result
        assert "/efficiency-mapping" in result
        return result

    def _improved_efficiency_card(self, item: EfficiencyMappingTask) -> str:
        if not isinstance(item, EfficiencyMappingTask):
            raise TypeError("item must be EfficiencyMappingTask")
        brand_chips = "".join(self._brand_logo_chip(brand.brand_name, "completed-brand-chip") for brand in item.brands)
        result = f"""
        <article class="improved-efficiency-card priority-border-{_e(item.priority.lower())}">
          <header>
            <span class="priority-badge priority-{_e(item.priority.lower())}">{_e(item.priority)}</span>
            <strong>{_e(item.task_name)}</strong>
          </header>
          <div class="completed-brand-list">{brand_chips}</div>
          <small>{_format_decimal(item.monthly_hours)} 小时/月 · {len(item.brands)} 个已提效品牌</small>
        </article>
        """
        assert item.task_name in result
        return result

    def _manual_efficiency_brand_form(self) -> str:
        items = self._high_efficiency_mapping_items()
        options = "".join(f'<option value="{_e(item.task_name)}">{_e(item.priority)} - {_e(item.task_name)}</option>' for item in items)
        result = f"""
        <section class="manual-efficiency-brand-panel">
          <div class="section-heading">
            <h2>手动增加品牌</h2>
            <p>当参考表没有覆盖某个品牌时，可在这里补充到对应任务；勾选已提效后会自动出现在首页已提效项目里。</p>
          </div>
          <form method="post" action="/efficiency-mapping/add-brand" class="manual-efficiency-brand-form">
            <label><span>所属任务</span><select name="task_name" required>{options}</select></label>
            <label><span>品牌名称</span><input name="brand_name" maxlength="100" required placeholder="如：安踏、CK、博西"></label>
            <label class="checkbox-line"><input type="checkbox" name="is_improved"><span>已提效，首页标绿展示</span></label>
            <label><span>未提效原因/备注</span><textarea name="not_improved_reason" maxlength="2000" placeholder="可选"></textarea></label>
            <label><span>排期规划</span><textarea name="schedule_plan" maxlength="2000" placeholder="可选，留空则使用系统默认排期"></textarea></label>
            <button class="button" type="submit">新增品牌</button>
          </form>
        </section>
        """
        assert "手动增加品牌" in result
        return result

    def _high_efficiency_mapping_panel(self) -> str:
        items = self._high_efficiency_mapping_items()
        if not items:
            return """
            <section class="efficiency-mapping-panel">
              <div class="section-heading">
                <h2>高耗时可替代/可提效任务映射</h2>
                <p>参考规划排期文件暂未读取成功，请确认文件仍在项目根目录。</p>
              </div>
            </section>
            """
        priority_sections = "".join(
            self._efficiency_priority_section(priority, tuple(item for item in items if item.priority == priority))
            for priority, _, _ in PRIORITY_SECTIONS
            if any(item.priority == priority for item in items)
        )
        result = f"""
        <section class="efficiency-ledger-panel">
          <div class="section-heading">
            <h2>全部待提效和排期</h2>
            <p>按 P1-P4 排序，默认只显示品牌名称和状态；点击品牌后再维护已提效、未提效原因和排期规划。</p>
          </div>
          <div class="efficiency-priority-stack">{priority_sections}</div>
        </section>
        """
        assert "全部待提效和排期" in result
        assert "efficiency-priority-stack" in result
        return result

    def _efficiency_priority_section(self, priority: str, items: tuple[EfficiencyMappingTask, ...]) -> str:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        if not isinstance(items, tuple):
            raise TypeError("items must be tuple")
        priority_name = next(section[1] for section in PRIORITY_SECTIONS if section[0] == priority)
        task_cards = "".join(self._high_efficiency_mapping_card(item) for item in items)
        result = f"""
        <section class="efficiency-priority-section priority-border-{_e(priority.lower())}">
          <header>
            <span class="priority-badge priority-{_e(priority.lower())}">{_e(priority)}</span>
            <strong>{_e(priority_name)}</strong>
            <small>{len(items)} 个任务</small>
          </header>
          <div class="efficiency-task-stack">{task_cards}</div>
        </section>
        """
        assert priority in result
        return result

    def _high_efficiency_mapping_card(self, item: EfficiencyMappingTask) -> str:
        if not isinstance(item, EfficiencyMappingTask):
            raise TypeError("item must be EfficiencyMappingTask")
        improved_count = sum(1 for brand in item.brands if brand.is_improved)
        brand_forms = "".join(self._high_efficiency_brand_entry(brand) for brand in item.brands)
        result = f"""
        <article class="efficiency-task-card compact-ledger-task">
          <header>
            <div class="efficiency-task-title">
              <div>
                <h3>{_e(item.task_name)}</h3>
                <p>类型 {_e(item.replacement_type)} · {_format_decimal(item.monthly_hours)} 小时/月 · 已提效 {improved_count}/{len(item.brands)}</p>
              </div>
            </div>
            <p class="efficiency-reason"><strong>替代/提效理由</strong>{_e(item.replacement_reason)}</p>
          </header>
          <div class="efficiency-brand-list">{brand_forms}</div>
        </article>
        """
        assert item.task_name in result
        return result

    def _high_efficiency_brand_entry(self, brand: EfficiencyMappingBrand) -> str:
        if not isinstance(brand, EfficiencyMappingBrand):
            raise TypeError("brand must be EfficiencyMappingBrand")
        status_class = "is-improved" if brand.is_improved else "is-pending"
        status_label = "已提效" if brand.is_improved else "待提效"
        form = self._high_efficiency_brand_form(brand)
        result = f"""
        <details class="efficiency-brand-entry {status_class}">
          <summary>
            {self._brand_logo_chip(brand.brand_name, "brand-entry-chip")}
            <span class="brand-entry-status">{_e(status_label)}</span>
          </summary>
          {form}
        </details>
        """
        assert _e(brand.brand_name) in result
        return result

    def _high_efficiency_brand_form(self, brand: EfficiencyMappingBrand) -> str:
        if not isinstance(brand, EfficiencyMappingBrand):
            raise TypeError("brand must be EfficiencyMappingBrand")
        status_class = "is-improved" if brand.is_improved else "is-pending"
        status_label = "已提效" if brand.is_improved else "待提效"
        note = brand.note
        reason = note.not_improved_reason if note is not None else ""
        schedule = note.schedule_plan if note is not None and note.schedule_plan.strip() else self._default_efficiency_schedule(brand.task_name)
        is_manual_brand = note.is_manual_brand if note is not None else False
        reason_label = "备注" if brand.is_improved else "未提效原因"
        reason_placeholder = "可填写已提效后的补充反馈" if brand.is_improved else "填写暂未提效的原因，如：缺少源数据、字段未统一、业务排期未确认"
        checked = " checked" if brand.is_improved else ""
        result = f"""
        <form method="post" action="/efficiency-mapping/save" class="efficiency-brand-form {status_class}">
          <input type="hidden" name="task_name" value="{_e(brand.task_name)}">
          <input type="hidden" name="brand_name" value="{_e(brand.brand_name)}">
          <input type="hidden" name="is_manual_brand" value="{"1" if is_manual_brand else "0"}">
          <div class="efficiency-brand-head">
            <strong>{_e(brand.brand_name)}</strong>
            <span>{_e(status_label)}</span>
          </div>
          <label class="checkbox-line efficiency-status-check"><input type="checkbox" name="is_improved"{checked}><span>已提效，联动首页标绿</span></label>
          <label>
            <span>{_e(reason_label)}</span>
            <textarea name="not_improved_reason" maxlength="2000" placeholder="{_e(reason_placeholder)}">{_e(reason)}</textarea>
          </label>
          <label>
            <span>排期规划</span>
            <textarea name="schedule_plan" maxlength="2000" placeholder="填写计划周期、依赖资料和下一步动作">{_e(schedule)}</textarea>
          </label>
          <button class="button compact-save-button" type="submit">保存</button>
        </form>
        """
        assert _e(brand.brand_name) in result
        return result

    def _high_efficiency_mapping_items(self) -> tuple[EfficiencyMappingTask, ...]:
        workbook_path = DEVELOPMENT_PLAN_PATH
        if not workbook_path.exists():
            logging.error("development plan workbook missing: %s", workbook_path)
            return ()
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            task_sheet = workbook["高耗时任务映射"]
            brand_sheet = workbook["P1-P4品牌业务方"]
        except (ImportError, OSError, KeyError, ValueError) as exc:
            logging.error("failed to open development plan workbook: %s", exc)
            return ()

        notes = self.storage.list_efficiency_mapping_notes()
        brand_rows: dict[str, tuple[str, tuple[str, ...], tuple[str, ...]]] = {}
        for row_index in range(4, brand_sheet.max_row + 1):
            priority = brand_sheet.cell(row_index, 1).value
            task_name = brand_sheet.cell(row_index, 2).value
            brands_text = brand_sheet.cell(row_index, 3).value
            owners_text = brand_sheet.cell(row_index, 4).value
            if not isinstance(priority, str) or priority.strip() not in {"P1", "P2", "P3", "P4"}:
                continue
            if not isinstance(task_name, str) or not task_name.strip():
                continue
            brands = self._split_cn_list(str(brands_text or ""))
            owners = self._split_cn_list(str(owners_text or ""))
            brand_rows[task_name.strip()] = (priority.strip(), brands or ("全组多品牌",), owners)

        tasks: list[EfficiencyMappingTask] = []
        for row_index in range(4, task_sheet.max_row + 1):
            replacement_type = task_sheet.cell(row_index, 1).value
            task_name = task_sheet.cell(row_index, 2).value
            monthly_hours = task_sheet.cell(row_index, 3).value
            reason = task_sheet.cell(row_index, 4).value
            if not isinstance(replacement_type, str) or not replacement_type.strip():
                continue
            if not isinstance(task_name, str) or not task_name.strip():
                continue
            if not isinstance(reason, str) or not reason.strip():
                continue
            try:
                hours = Decimal(str(monthly_hours))
            except (InvalidOperation, TypeError):
                logging.error("invalid monthly hours in development plan: row=%s task=%s", row_index, task_name)
                continue
            priority, brands, owners = brand_rows.get(
                task_name.strip(),
                (self._priority_for_replacement_type(replacement_type.strip()), ("全组多品牌",), ()),
            )
            manual_brands = tuple(
                note.brand_name
                for (note_task, _), note in notes.items()
                if note_task == task_name.strip() and note.is_manual_brand and note.brand_name not in brands
            )
            merged_brands = tuple(dict.fromkeys((*brands, *manual_brands)))
            brand_items = tuple(
                EfficiencyMappingBrand(
                    task_name=task_name.strip(),
                    brand_name=self._normalize_brand(brand),
                    business_owners=owners,
                    is_improved=self._resolved_efficiency_improved_state(
                        task_name=task_name.strip(),
                        brand_name=brand,
                        notes=notes,
                    ),
                    note=notes.get((task_name.strip(), self._normalize_brand(brand))),
                )
                for brand in merged_brands
            )
            tasks.append(
                EfficiencyMappingTask(
                    priority=priority,
                    replacement_type=replacement_type.strip(),
                    task_name=task_name.strip(),
                    monthly_hours=hours,
                    replacement_reason=reason.strip(),
                    brands=brand_items,
                )
            )
        result = tuple(sorted(tasks, key=lambda item: item.monthly_hours, reverse=True))
        logging.info("built efficiency mapping items: tasks=%s", len(result))
        assert result
        return result

    def _efficiency_mapping_pairs(self) -> set[tuple[str, str]]:
        result = {
            (item.task_name, brand.brand_name)
            for item in self._high_efficiency_mapping_items()
            for brand in item.brands
        }
        assert isinstance(result, set)
        return result

    @staticmethod
    def _split_cn_list(value: str) -> tuple[str, ...]:
        if not isinstance(value, str):
            raise TypeError("value must be str")
        tokens = [token.strip() for token in re.split(r"[、\r\n]+", value) if token.strip()]
        result = tuple(dict.fromkeys(tokens))
        return result

    @staticmethod
    def _priority_for_replacement_type(replacement_type: str) -> str:
        if not isinstance(replacement_type, str) or not replacement_type.strip():
            raise ValueError("replacement_type must not be empty")
        mapping = {"A": "P1", "B": "P2", "C": "P3", "D": "P4"}
        return mapping.get(replacement_type.strip().upper(), "P4")

    @staticmethod
    def _default_efficiency_schedule(task_name: str) -> str:
        if not isinstance(task_name, str) or not task_name.strip():
            raise ValueError("task_name must not be empty")
        if any(keyword in task_name for keyword in ("日报", "周报", "月报", "短彩信数据")):
            return "第1-5周：确认源文件、字段口径、清洗规则和报表模板，完成自动化闭环。"
        if any(keyword in task_name for keyword in ("选品", "内容", "文案")):
            return "第6-10周：补齐商品资料、品牌规范和审核规则，接入AI内容生产流水线。"
        if any(keyword in task_name for keyword in ("配置", "上下架", "商品")):
            return "第8-12周：冻结配置模板和校验规则，先半自动输出，人工审核后执行。"
        return "待业务确认优先级、资料完整度和上线窗口后排期。"

    def _is_improved_task_brand(self, task_name: str, brand_name: str) -> bool:
        if not isinstance(task_name, str) or not task_name.strip():
            raise ValueError("task_name must not be empty")
        if not isinstance(brand_name, str) or not brand_name.strip():
            raise ValueError("brand_name must not be empty")
        normalized_brand = self._normalize_brand(brand_name).lower()
        if "短彩信数据" in task_name:
            return normalized_brand in {"博西", "ck", "armani", "tommy", "nes"}
        if any(keyword in task_name for keyword in ("日报", "周报", "月报", "上下架")):
            return normalized_brand == "安踏"
        return False

    def _resolved_efficiency_improved_state(
        self,
        task_name: str,
        brand_name: str,
        notes: dict[tuple[str, str], EfficiencyMappingRecord],
    ) -> bool:
        if not isinstance(task_name, str) or not task_name.strip():
            raise ValueError("task_name must not be empty")
        if not isinstance(brand_name, str) or not brand_name.strip():
            raise ValueError("brand_name must not be empty")
        if not isinstance(notes, dict):
            raise TypeError("notes must be dict")
        normalized_brand = self._normalize_brand(brand_name)
        note = notes.get((task_name.strip(), normalized_brand))
        if note is not None:
            return note.is_improved
        return self._is_improved_task_brand(task_name.strip(), normalized_brand)

    def _project_platform_architecture_panel(self) -> str:
        groups = self._project_platform_groups()
        platform_cards = "".join(self._project_platform_card(group) for group in groups)
        total_brands = len({brand.brand for group in groups for brand in group.brands})
        result = f"""
        <section class="platform-architecture">
          <div class="section-heading">
            <h2>渠道品牌架构图</h2>
            <p>按渠道归类品牌覆盖范围；跨渠道品牌会在对应渠道中同时出现，44 个项目明细放入二级页面查看。</p>
          </div>
          <div class="platform-architecture-board">
            <div class="platform-root-node">
              <span>全组品牌覆盖池</span>
              <strong>{total_brands}</strong>
              <small>渠道 → 品牌</small>
            </div>
            <div class="platform-node-grid">{platform_cards}</div>
          </div>
        </section>
        """
        assert "渠道品牌架构图" in result
        assert total_brands > 0
        return result

    def _project_platform_groups(self) -> tuple[ProjectPlatformGroup, ...]:
        distribution = self._channel_brand_distribution()
        if distribution:
            developed_brands = self._developed_brand_names()
            groups = tuple(
                ProjectPlatformGroup(
                    platform=platform,
                    description=description,
                    brands=tuple(
                        ChannelBrandNode(
                            brand=brand,
                            platform=platform,
                            is_developed=brand.lower() in developed_brands,
                            source_project_count=1,
                        )
                        for brand in sorted(
                            distribution[platform],
                            key=lambda brand: (brand.lower() not in developed_brands, brand.lower()),
                        )
                    ),
                    source_project_count=1,
                )
                for platform, description in PLATFORM_SECTIONS
                if platform in distribution and distribution[platform]
            )
            logging.info("built channel brand architecture from distribution workbook: groups=%s", len(groups))
            assert groups
            return groups

        platform_brand_rows: dict[str, dict[str, dict[str, object]]] = {platform: {} for platform, _ in PLATFORM_SECTIONS}
        platform_source_projects: dict[str, set[str]] = {platform: set() for platform, _ in PLATFORM_SECTIONS}
        coverage_by_item = self._work_item_coverage_map()
        for item in GROUP_PROJECT_TREE_ITEMS:
            coverage = coverage_by_item.get(item.project, WorkItemCoverage(channels=("多渠道通用",), brands=()))
            developed_projects = self._developed_projects_for_work_item(item)
            developed_brands = {brand.lower() for brand in developed_projects if brand != "待补充"}
            developed_brand_names = tuple(brand for brand in developed_projects if brand != "待补充")
            covered_brands = tuple(dict.fromkeys((*coverage.brands, *developed_brand_names))) or ("未识别",)
            for platform in coverage.channels:
                if platform not in platform_brand_rows:
                    platform = "多渠道通用"
                platform_source_projects[platform].add(item.project)
                for brand in covered_brands:
                    brand_row = platform_brand_rows[platform].setdefault(
                        brand,
                        {"is_developed": False, "source_projects": set()},
                    )
                    if not isinstance(brand_row["source_projects"], set):
                        raise TypeError("source_projects must be set")
                    brand_row["source_projects"].add(item.project)
                    if brand.lower() in developed_brands:
                        brand_row["is_developed"] = True
        groups = tuple(
            ProjectPlatformGroup(
                platform=platform,
                description=description,
                brands=tuple(
                    ChannelBrandNode(
                        brand=brand,
                        platform=platform,
                        is_developed=bool(values["is_developed"]),
                        source_project_count=len(values["source_projects"]),
                    )
                    for brand, values in sorted(
                        platform_brand_rows[platform].items(),
                        key=lambda row: (not bool(row[1]["is_developed"]), row[0].lower()),
                    )
                ),
                source_project_count=len(platform_source_projects[platform]),
            )
            for platform, description in PLATFORM_SECTIONS
            if platform_brand_rows[platform]
        )
        assigned_projects = set().union(*platform_source_projects.values()) if platform_source_projects else set()
        logging.info("built channel brand architecture: groups=%s unique_work_items=%s", len(groups), len(assigned_projects))
        assert len(assigned_projects) == len(GROUP_PROJECT_TREE_ITEMS)
        return groups

    def _channel_brand_distribution(self) -> dict[str, set[str]]:
        workbook_path = CHANNEL_BRAND_DISTRIBUTION_PATH
        if not workbook_path.exists():
            logging.error("channel brand distribution workbook missing: %s", workbook_path)
            return {}
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
        except (ImportError, OSError, KeyError, ValueError) as exc:
            logging.error("failed to open channel brand distribution workbook: %s", exc)
            return {}

        known_platforms = {platform for platform, _ in PLATFORM_SECTIONS}
        distribution: dict[str, set[str]] = {platform: set() for platform in known_platforms}
        current_channel = ""
        for column in range(1, worksheet.max_column + 1):
            raw_channel = worksheet.cell(1, column).value
            raw_brand = worksheet.cell(2, column).value
            if isinstance(raw_channel, str) and raw_channel.strip():
                current_channel = raw_channel.strip()
            if not isinstance(raw_brand, str) or not raw_brand.strip() or raw_brand.strip() == "品牌":
                continue
            if not current_channel or current_channel == "平台":
                continue
            brand = self._normalize_brand(raw_brand)
            for channel in self._normalize_channel(current_channel):
                platform = channel if channel in known_platforms else "多渠道通用"
                distribution.setdefault(platform, set()).add(brand)

        result = {platform: brands for platform, brands in distribution.items() if brands}
        assert isinstance(result, dict)
        return result

    @staticmethod
    def _developed_brand_names() -> set[str]:
        result = {
            item.brand.lower()
            for item in IntranetApp._completed_feedback_items()
            if item.brand != "待补充"
        }
        assert result
        return result

    @staticmethod
    def _project_platform_for_stage(project: str, brand: str, business_type: str) -> str:
        return "、".join(IntranetApp._project_channels_for_stage(project, brand, business_type))

    @staticmethod
    def _project_channels_for_stage(project: str, brand: str, business_type: str) -> tuple[str, ...]:
        for field_name, field_value in (("project", project), ("brand", brand), ("business_type", business_type)):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        text = f"{project} {brand} {business_type}"
        if "短彩信" in text or "CRM" in text:
            return ("CRM",)
        if "安踏即时零售" in text:
            return ("美团", "京东")
        if "安踏周报/月报" in text:
            return ("美团", "京东")
        if "页面巡检" in text:
            return ("美团", "京东", "天猫", "小程序", "官网")
        return ("多渠道通用",)

    @staticmethod
    def _normalize_channel(raw_channel: str) -> tuple[str, ...]:
        if not isinstance(raw_channel, str) or not raw_channel.strip():
            raise ValueError("raw_channel must not be empty")
        normalized: list[str] = []
        for token in re.split(r"[&/、,，]+", raw_channel):
            value = token.strip()
            if not value:
                continue
            if value in {"企微", "社群"}:
                value = "企微/社群"
            normalized.append(value)
        result = tuple(dict.fromkeys(normalized))
        assert result
        return result

    @staticmethod
    def _channels_for_work_item(project: str, channels: list[str]) -> tuple[str, ...]:
        if not isinstance(project, str) or not project.strip():
            raise ValueError("project must not be empty")
        if not isinstance(channels, list):
            raise TypeError("channels must be list")
        if any(not isinstance(channel, str) or not channel.strip() for channel in channels):
            raise ValueError("channels must only contain non-empty strings")

        unique_channels = tuple(dict.fromkeys(channels))
        crm_keywords = ("短彩信", "CRM", "触达", "会员")
        sales_report_keywords = ("日报", "周报", "月报", "上下架", "页面", "巡店")
        is_crm_project = any(keyword in project for keyword in crm_keywords)

        if is_crm_project and "CRM" not in unique_channels:
            unique_channels = ("CRM", *unique_channels)
        if not is_crm_project:
            unique_channels = tuple(channel for channel in unique_channels if channel != "CRM")
        if any(keyword in project for keyword in sales_report_keywords):
            unique_channels = tuple(dict.fromkeys((*unique_channels, "美团", "京东")))

        result = unique_channels or ("多渠道通用",)
        assert result
        return result

    @staticmethod
    def _normalize_brand(raw_brand: str) -> str:
        if not isinstance(raw_brand, str) or not raw_brand.strip():
            raise ValueError("raw_brand must not be empty")
        mapping = {
            "ANTA": "安踏",
            "armani": "Armani",
            "阿玛尼": "Armani",
            "TOMMY": "Tommy",
            "NES": "Nes",
            "Nespresso": "Nes",
            "bosch": "博西",
        }
        return mapping.get(raw_brand.strip(), raw_brand.strip())

    def _work_item_coverage_map(self) -> dict[str, WorkItemCoverage]:
        workbook_path = Path.home() / "Downloads" / "内容任务耗时统计.xlsx"
        if not workbook_path.exists():
            logging.error("work item coverage workbook missing: %s", workbook_path)
            return {}
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(workbook_path, data_only=True)
            worksheet = workbook[workbook.sheetnames[0]]
        except (ImportError, OSError, KeyError, ValueError) as exc:
            logging.error("failed to open work item coverage workbook: %s", exc)
            return {}
        platform_headers = [worksheet.cell(3, column).value for column in range(8, 69)]
        brand_headers = [worksheet.cell(4, column).value for column in range(8, 69)]
        coverage: dict[str, WorkItemCoverage] = {}
        for row_index in range(7, worksheet.max_row + 1):
            project = worksheet.cell(row_index, 3).value
            total_hours = worksheet.cell(row_index, 7).value
            if not isinstance(project, str) or not isinstance(total_hours, (int, float)):
                continue
            channels: list[str] = []
            brands: list[str] = []
            for offset, column in enumerate(range(8, 69)):
                value = worksheet.cell(row_index, column).value
                if not isinstance(value, (int, float)) or value <= 0:
                    continue
                raw_brand = brand_headers[offset]
                raw_channel = platform_headers[offset]
                if isinstance(raw_brand, str) and raw_brand.strip():
                    brands.append(self._normalize_brand(raw_brand))
                if isinstance(raw_channel, str) and raw_channel.strip():
                    channels.extend(self._normalize_channel(raw_channel))
            normalized_channels = self._channels_for_work_item(project, channels)
            normalized_brands = tuple(dict.fromkeys(brands))
            coverage[project] = WorkItemCoverage(channels=normalized_channels, brands=normalized_brands)
        assert len(coverage) >= 40
        return coverage

    @staticmethod
    def _developed_projects_for_work_item(item: GroupProjectTreeItem) -> tuple[str, ...]:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        if "短彩信" in item.project:
            return ("博西", "CK", "Armani", "Tommy", "Nes")
        if any(keyword in item.project for keyword in ("日报", "周报", "月报")):
            return ("安踏",)
        return ("待补充",)

    @staticmethod
    def _pending_projects_for_work_item(
        item: GroupProjectTreeItem,
        covered_brands: tuple[str, ...],
        developed_projects: tuple[str, ...],
    ) -> tuple[str, ...]:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        if not isinstance(covered_brands, tuple):
            raise TypeError("covered_brands must be tuple")
        if not isinstance(developed_projects, tuple):
            raise TypeError("developed_projects must be tuple")
        developed = {brand.lower() for brand in developed_projects if brand != "待补充"}
        pending = tuple(
            brand
            for brand in covered_brands
            if brand.lower() not in developed and brand not in {"标准模板包", "未识别"}
        )
        if pending:
            return pending
        return ("待补充",)

    @staticmethod
    def _original_hours_for_work_item(project: str) -> Decimal:
        if not isinstance(project, str) or not project.strip():
            raise ValueError("project must not be empty")
        for item in GROUP_PROJECT_TREE_ITEMS:
            if item.project == project:
                return item.original_hours
        logging.error("work item original hours missing: %s", project)
        return Decimal("0")

    def _project_platform_card(self, group: ProjectPlatformGroup) -> str:
        if not isinstance(group, ProjectPlatformGroup):
            raise TypeError("group must be ProjectPlatformGroup")
        brand_rows = "".join(self._project_channel_brand_row(brand) for brand in group.brands)
        result = f"""
        <article class="platform-node">
          <header>
            <span>渠道</span>
            <strong>{_e(group.platform)}</strong>
            <small>{len(group.brands)} 个品牌</small>
          </header>
          <p>{_e(group.description)}</p>
          <div class="platform-project-list">{brand_rows}</div>
        </article>
        """
        assert group.platform in result
        return result

    def _project_channel_brand_row(self, brand: ChannelBrandNode) -> str:
        if not isinstance(brand, ChannelBrandNode):
            raise TypeError("brand must be ChannelBrandNode")
        developed_class = " is-developed" if brand.is_developed else ""
        result = f"""
        <div class="platform-project-chip{developed_class}">
          <div class="platform-project-body">
            <strong>{_e(brand.brand)}</strong>
          </div>
        </div>
        """
        assert _e(brand.brand) in result
        return result

    def _project_map_priority_lane(self, priority: str) -> str:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        items = [item for item in GROUP_PROJECT_TREE_ITEMS if item.priority == priority]
        if not items:
            raise ValueError("priority must contain project items")
        priority_meta = next(item for item in PRIORITY_SECTIONS if item[0] == priority)
        branch_names = tuple(dict.fromkeys(item.branch for item in items))
        branch_blocks = "".join(
            self._project_map_branch_block(
                priority,
                branch,
                [item for item in items if item.branch == branch],
            )
            for branch in branch_names
        )
        hours = sum((item.original_hours for item in items), Decimal("0"))
        result = f"""
        <article class="project-map-lane priority-border-{_e(priority.lower())}">
          <header>
            <span class="priority-badge priority-{_e(priority.lower())}">{_e(priority)}</span>
            <strong>{_e(priority_meta[1])}</strong>
            <small>{len(items)} 个项目 · {_format_decimal(hours)} 小时/月</small>
          </header>
          <div class="project-map-branches">{branch_blocks}</div>
        </article>
        """
        assert priority in result
        return result

    def _project_map_branch_block(self, priority: str, branch: str, items: list[GroupProjectTreeItem]) -> str:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        if not isinstance(branch, str) or not branch.strip():
            raise ValueError("branch must not be empty")
        if not isinstance(items, list) or not items:
            raise ValueError("items must not be empty")
        cards = "".join(self._project_map_card(item, GROUP_PROJECT_TREE_ITEMS.index(item) + 1) for item in items)
        result = f"""
        <section class="project-map-branch">
          <div class="project-map-branch-title">
            <span>{_e(branch)}</span>
            <em>{len(items)}项</em>
          </div>
          <div class="project-map-cards">{cards}</div>
        </section>
        """
        assert branch in result
        return result

    def _project_map_card(self, item: GroupProjectTreeItem, index: int) -> str:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        if not isinstance(index, int) or index <= 0:
            raise ValueError("index must be positive")
        replacement_type = self._work_item_replacement_type(item.project)
        can_improve = replacement_type != "C"
        disabled_class = "" if can_improve else " is-not-efficiency-fit"
        status_label = f"{replacement_type} · 可提效" if can_improve else "C · 暂不提效"
        result = f"""
        <a class="project-map-card priority-card-{_e(item.priority.lower())}{disabled_class}" href="#project-detail-{index}">
          <strong>{_e(item.project)}</strong>
          <span>{_e(item.business_type)}</span>
          <small class="project-ai-status">{_e(status_label)}</small>
          <em>{_format_decimal(item.original_hours)} 小时/月</em>
        </a>
        """
        assert item.project in result
        return result

    def _project_detail_modal(self, item: GroupProjectTreeItem, index: int) -> str:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        if not isinstance(index, int) or index <= 0:
            raise ValueError("index must be positive")
        replacement_type = self._work_item_replacement_type(item.project)
        replacement_status = "有机会 AI 提效" if replacement_type != "C" else "无法 AI 提效，保留人工协同"
        result = f"""
        <div id="project-detail-{index}" class="project-modal" role="dialog" aria-modal="true">
          <a class="project-modal-backdrop" href="#"></a>
          <section class="project-modal-panel">
            <div class="project-modal-heading">
              <span class="priority-badge priority-{_e(item.priority.lower())}">{_e(item.priority)}</span>
              <div>
                <h2>{_e(item.project)}</h2>
                <p>{_e(item.branch)} · {_e(item.brand)}</p>
              </div>
              <a class="project-modal-close" href="#">关闭</a>
            </div>
            <dl class="project-detail-list">
              <div><dt>具体工作内容</dt><dd>{_e(item.project)}</dd></div>
              <div><dt>渠道归属</dt><dd>{_e(self._project_platform_for_stage(item.project, item.brand, item.business_type))}</dd></div>
              <div><dt>业务类型</dt><dd>{_e(item.business_type)}</dd></div>
              <div><dt>原人工耗时</dt><dd>{_format_decimal(item.original_hours)} 小时/月</dd></div>
              <div><dt>可替代类型</dt><dd>{_e(replacement_type)} · {_e(replacement_status)}</dd></div>
              <div><dt>数据来源</dt><dd>{_e(item.source_detail)}</dd></div>
              <div><dt>当前建议</dt><dd>{_e(self._project_ai_replacement_note(item))}</dd></div>
            </dl>
          </section>
        </div>
        """
        assert item.project in result
        return result

    @staticmethod
    def _work_item_replacement_type(project: str) -> str:
        if not isinstance(project, str) or not project.strip():
            raise ValueError("project must not be empty")
        result = "C" if project.strip() in NON_AI_EFFICIENCY_WORK_ITEMS else "A/B"
        assert result in {"A/B", "C"}
        return result

    def _developable_group_items(self, priority: str | None = None) -> tuple[GroupProjectTreeItem, ...]:
        if priority is not None and priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4 or None")
        result = tuple(
            item
            for item in GROUP_PROJECT_TREE_ITEMS
            if (priority is None or item.priority == priority) and self._work_item_replacement_type(item.project) != "C"
        )
        assert all(self._work_item_replacement_type(item.project) != "C" for item in result)
        return result

    def _developable_project_count(self) -> int:
        result = len(self._developable_group_items())
        if result <= 0:
            raise AssertionError("developable project count must be positive")
        return result

    def _priority_development_stats(self, priority: str) -> PriorityDevelopmentStats:
        if priority not in {"P1", "P2", "P3", "P4"}:
            raise ValueError("priority must be P1-P4")
        total_count = len(self._developable_group_items(priority))
        developed_count = sum(
            1
            for row_priority, stage, _, _, _ in PROJECT_STAGE_ROWS
            if row_priority == priority and stage in {"开发完成", "已经开发"}
        )
        developed_count = min(developed_count, total_count)
        result = PriorityDevelopmentStats(
            priority=priority,
            total_count=total_count,
            developed_count=developed_count,
            pending_count=total_count - developed_count,
        )
        logging.info(
            "built priority development stats: priority=%s total=%s developed=%s pending=%s",
            result.priority,
            result.total_count,
            result.developed_count,
            result.pending_count,
        )
        assert result.pending_count >= 0
        return result

    def _completed_business_feedback_panel(self, feedback_by_project: dict[str, ProjectFeedbackRecord], return_to: str) -> str:
        if not isinstance(feedback_by_project, dict):
            raise TypeError("feedback_by_project must be dict")
        if return_to not in {"/", "/project-stages"}:
            raise ValueError("return_to must be supported route")
        manual_time_by_project = self._manual_time_by_project()
        feedback_items = self._completed_feedback_items_for_dashboard()
        cards = "".join(
            self._completed_feedback_card(item, manual_time_by_project, feedback_by_project, return_to)
            for item in feedback_items
        )
        result = f"""
        <section class="completed-feedback-panel">
          <div class="section-heading">
            <h2>已开发完整业务反馈</h2>
            <p>这里只展示已提效映射生成的反馈卡；核心只看原耗时数据（单次配置耗时）、现在耗时数据和提效时间。</p>
          </div>
          <div class="completed-feedback-grid">{cards}</div>
        </section>
        """
        assert "已开发完整业务反馈" in result
        return result

    def _completed_feedback_dashboard_table(self, feedback_by_project: dict[str, ProjectFeedbackRecord]) -> str:
        if not isinstance(feedback_by_project, dict):
            raise TypeError("feedback_by_project must be dict")
        manual_time_by_project = self._manual_time_by_project()
        feedback_items = self._completed_feedback_items_for_dashboard()
        rows: list[str] = []
        for item in feedback_items:
            if not isinstance(item, CompletedFeedbackItem):
                raise TypeError("feedback_items must contain CompletedFeedbackItem")
            default_original_time = manual_time_by_project.get(item.manual_time_project, {}).get(
                "原人工耗时",
                self._default_original_time_for_completed_item(item),
            )
            record = feedback_by_project.get(item.feedback_key) or feedback_by_project.get(item.legacy_project)
            original_time = (
                record.original_manual_time
                if record is not None and record.original_manual_time.strip()
                else default_original_time
            )
            current_time = record.current_processing_time if record is not None else ""
            business_feedback = record.business_feedback if record is not None else ""
            original_hours = _parse_duration_hours(original_time)
            saved_time = self._format_time_saved(original_hours, current_time) if original_hours is not None else "待计算"
            rows.append(
                f"""
                <tr>
                  <td><span class="priority-badge priority-{_e(item.priority.lower())}">{_e(item.priority)}</span></td>
                  <td><strong>{_e(item.project)}</strong></td>
                  <td>{self._brand_logo_chip(item.brand, "completed-brand-chip")}</td>
                  <td>{_e(_normalize_duration_hours_text(original_time))}</td>
                  <td>{_e(_normalize_duration_hours_text(current_time)) if current_time.strip() else "待填写"}</td>
                  <td>{_e(saved_time)}</td>
                  <td>{_e(business_feedback) if business_feedback.strip() else "待填写"}</td>
                </tr>
                """
            )
        result = f"""
        <section class="feedback-summary-panel merged-feedback-summary">
          <div class="section-heading split-heading">
            <div>
              <h2>已开发反馈汇总</h2>
              <p>只保留表格形式；每一行都来自二级页面“已开发完整业务反馈”的项目、品牌、耗时和业务反馈。</p>
            </div>
            <a class="button secondary" href="/project-stages">查看完整反馈</a>
          </div>
          <div class="table-panel dashboard-stage-table-panel">
            <table class="data-table dashboard-stage-table">
              <thead><tr><th>优先级</th><th>项目</th><th>已开发品牌</th><th>原耗时数据</th><th>现在耗时数据</th><th>提效时间</th><th>业务反馈</th></tr></thead>
              <tbody>{''.join(rows)}</tbody>
            </table>
          </div>
        </section>
        """
        assert "已开发反馈汇总" in result
        assert "已开发品牌" in result
        return result

    def _completed_feedback_card(
        self,
        item: CompletedFeedbackItem,
        manual_time_by_project: dict[str, dict[str, str]],
        feedback_by_project: dict[str, ProjectFeedbackRecord],
        return_to: str = "/project-stages",
    ) -> str:
        if not isinstance(item, CompletedFeedbackItem):
            raise TypeError("item must be CompletedFeedbackItem")
        if return_to not in {"/", "/project-stages"}:
            raise ValueError("return_to must be supported route")
        default_original_time = manual_time_by_project.get(item.manual_time_project, {}).get(
            "原人工耗时",
            self._default_original_time_for_completed_item(item),
        )
        record = feedback_by_project.get(item.feedback_key) or feedback_by_project.get(item.legacy_project)
        original_time = record.original_manual_time if record is not None and record.original_manual_time.strip() else default_original_time
        original_hours = _parse_duration_hours(original_time)
        current_time = record.current_processing_time if record is not None else ""
        business_feedback = record.business_feedback if record is not None else ""
        iteration_need = record.iteration_need if record is not None else ""
        saved_time = "待计算"
        if original_hours is not None:
            saved_time = self._format_time_saved(original_hours, current_time)
        result = f"""
        <form method="post" action="/project-stages/feedback" class="completed-feedback-card">
          <input type="hidden" name="return_to" value="{_e(return_to)}">
          <input type="hidden" name="project" value="{_e(item.feedback_key)}">
          <header>
            <span class="priority-badge priority-{_e(item.priority.lower())}">{_e(item.priority)}</span>
            <strong>{_e(item.project)}</strong>
          </header>
          <div class="completed-brand-row">
            <span>已开发品牌</span>
            <div class="completed-brand-list">{self._brand_logo_chip(item.brand, "completed-brand-chip")}</div>
          </div>
          <div class="feedback-metrics">
            <label><span>原耗时数据（单次配置耗时）</span><input class="efficiency-source-input" name="original_manual_time" maxlength="100" value="{_e(_normalize_duration_hours_text(original_time))}" placeholder="如：40小时"></label>
            <label><span>现在耗时数据</span><input class="efficiency-source-input" name="current_processing_time" maxlength="100" value="{_e(_normalize_duration_hours_text(current_time))}" placeholder="如：0.67小时"></label>
            <div class="saved-time-metric"><span>提效时间</span><strong data-time-saved-output>{_e(saved_time)}</strong></div>
          </div>
          <label class="completed-feedback-text">
            <span>业务反馈</span>
            <textarea name="business_feedback" maxlength="2000" placeholder="业务试用后的效果、问题或建议">{_e(business_feedback)}</textarea>
          </label>
          <input type="hidden" name="iteration_need" value="{_e(iteration_need)}">
          <button class="button save-feedback-button" type="submit">保存反馈</button>
        </form>
        """
        assert _e(item.project) in result
        assert _e(item.brand) in result
        assert _e(item.feedback_key) in result
        return result

    def _brand_logo_chip(self, brand_name: str, class_name: str) -> str:
        if not isinstance(brand_name, str) or not brand_name.strip():
            raise ValueError("brand_name must not be empty")
        if not isinstance(class_name, str) or not class_name.strip():
            raise ValueError("class_name must not be empty")
        normalized_brand = self._normalize_brand(brand_name)
        logo_domain = self._brand_logo_domain(normalized_brand)
        initials = self._brand_logo_initials(normalized_brand)
        image_html = ""
        if logo_domain:
            image_url = f"https://www.google.com/s2/favicons?domain_url={quote(logo_domain)}&sz=64"
            image_html = f'<img src="{_e(image_url)}" alt="" loading="lazy" onerror="this.style.display=\'none\'">'
        result = f"""
        <span class="{_e(class_name)} brand-logo-chip">
          <span class="brand-logo-mark">{image_html}<em>{_e(initials)}</em></span>
          <span>{_e(normalized_brand)}</span>
        </span>
        """
        assert _e(normalized_brand) in result
        return result

    @staticmethod
    def _brand_logo_initials(brand_name: str) -> str:
        if not isinstance(brand_name, str) or not brand_name.strip():
            raise ValueError("brand_name must not be empty")
        compact = re.sub(r"\s+", "", brand_name.strip())
        result = compact[:2].upper()
        assert result
        return result

    @staticmethod
    def _brand_logo_domain(brand_name: str) -> str:
        if not isinstance(brand_name, str) or not brand_name.strip():
            raise ValueError("brand_name must not be empty")
        normalized = brand_name.strip().lower()
        mapping = {
            "安踏": "anta.com",
            "anta": "anta.com",
            "ck": "calvinklein.us",
            "armani": "armani.com",
            "阿玛尼": "armani.com",
            "tommy": "tommy.com",
            "nes": "nespresso.com",
            "nespresso": "nespresso.com",
            "博西": "bosch-home.cn",
            "bosch": "bosch-home.cn",
            "crocs": "crocs.com",
            "nike": "nike.com",
            "kolon": "kolonsport.com",
            "ecco": "ecco.com",
            "rapido": "rapido.com",
            "rituals": "rituals.com",
            "swisse": "swisse.com",
            "vans": "vans.com",
            "tb": "toryburch.com",
            "aape": "aape.jp",
            "bape": "bape.com",
            "af&hco": "abercrombie.com",
            "pns": "pns.com",
            "哥伦比亚": "columbia.com",
            "妮维雅": "nivea.com",
            "华润": "crc.com.cn",
            "吉利/康倍信": "geely.com",
            "港迪": "gandour.com",
            "维密": "victoriassecret.com",
            "爱茉莉": "amorepacific.com",
            "碧柔": "kao.com",
        }
        result = mapping.get(normalized, "")
        assert isinstance(result, str)
        return result

    def _completed_feedback_items_for_dashboard(self) -> tuple[CompletedFeedbackItem, ...]:
        by_key: dict[str, CompletedFeedbackItem] = {}
        for item in self._completed_feedback_items_from_efficiency_mapping():
            by_key.setdefault(item.feedback_key, item)
        result = tuple(by_key.values())
        assert result
        return result

    def _completed_feedback_items_from_efficiency_mapping(self) -> tuple[CompletedFeedbackItem, ...]:
        items: list[CompletedFeedbackItem] = []
        for task in self._high_efficiency_mapping_items():
            for brand in task.brands:
                if not brand.is_improved:
                    continue
                feedback_key = f"{task.priority}-{task.task_name}-{brand.brand_name}"
                items.append(
                    CompletedFeedbackItem(
                        priority=task.priority,
                        project=task.task_name,
                        brand=brand.brand_name,
                        feedback_key=feedback_key,
                        legacy_project=feedback_key,
                        manual_time_project=task.task_name,
                    )
                )
        result = tuple(items)
        assert isinstance(result, tuple)
        return result

    @staticmethod
    def _completed_feedback_items() -> tuple[CompletedFeedbackItem, ...]:
        sms_brands = ("博西", "CK", "Armani", "Tommy", "Nes")
        items = [
            CompletedFeedbackItem(
                priority="P1",
                project="短彩信数据处理",
                brand=brand,
                feedback_key=f"P1-短彩信数据处理-{brand}",
                legacy_project="博西短彩信数据处理",
                manual_time_project="博西短彩信数据处理",
            )
            for brand in sms_brands
        ]
        items.append(
            CompletedFeedbackItem(
                priority="P3",
                project="即时零售",
                brand="安踏",
                feedback_key="P3-即时零售-安踏",
                legacy_project="安踏即时零售",
                manual_time_project="安踏即时零售",
            )
        )
        result = tuple(items)
        assert result
        return result

    @staticmethod
    def _default_original_time_for_completed_item(item: CompletedFeedbackItem) -> str:
        if not isinstance(item, CompletedFeedbackItem):
            raise TypeError("item must be CompletedFeedbackItem")
        if item.project == "短彩信数据处理":
            result = "40小时"
        elif item.project == "即时零售" and item.brand == "安踏":
            result = "74小时"
        else:
            original_hours = next(
                (project_item.original_hours for project_item in GROUP_PROJECT_TREE_ITEMS if project_item.project == item.project),
                Decimal("0"),
            )
            result = f"{_format_decimal(original_hours)}小时" if original_hours > 0 else "待补充"
        assert result
        return result

    @staticmethod
    def _project_ai_replacement_note(item: GroupProjectTreeItem) -> str:
        if not isinstance(item, GroupProjectTreeItem):
            raise TypeError("item must be GroupProjectTreeItem")
        if item.priority == "P1":
            return "优先沉淀为标准数据层，再自动生成日报、周报、月报和指标分析。"
        if item.priority == "P2":
            return "基于统一基础数据层和品牌资料，由AI辅助选品、策略、卖点和文案生产。"
        if item.priority == "P3":
            return "适合用规则引擎和浏览器插件/RPA执行重复配置，但必须保留人工确认。"
        return "先沉淀巡检标准、历史问题和人工判断口径，成熟后再进入自动化。"

    def _development_overview_panel(self) -> str:
        feedback_by_project = self.storage.list_project_feedback()
        manual_time_by_project = self._manual_time_by_project()
        stage_counts = {
            stage: sum(1 for _, row_stage, _, _, _ in PROJECT_STAGE_ROWS if row_stage == stage)
            for stage in ("开发完成", "已经开发", "正在开发")
        }
        processing_time_count = sum(
            1 for record in feedback_by_project.values() if record.current_processing_time.strip()
        )
        feedback_count = sum(
            1 for record in feedback_by_project.values() if record.business_feedback.strip()
        )
        summary_items = (
            ("项目总数", str(len(PROJECT_STAGE_ROWS)), "neutral"),
            ("开发完成", str(stage_counts["开发完成"]), "complete"),
            ("已经开发", str(stage_counts["已经开发"]), "developed"),
            ("正在开发", str(stage_counts["正在开发"]), "in-progress"),
            ("已填处理耗时", str(processing_time_count), "time"),
            ("已收业务反馈", str(feedback_count), "feedback"),
        )
        summary_html = "".join(
            f"<div class='development-stat {css_class}'><span>{_e(label)}</span><strong>{_e(value)}</strong></div>"
            for label, value, css_class in summary_items
        )
        rows: list[str] = []
        for priority, stage, project, _, _ in PROJECT_STAGE_ROWS:
            record = feedback_by_project.get(project)
            original_time = manual_time_by_project.get(project, {}).get("原人工耗时", "待补充")
            current_time = (
                record.current_processing_time
                if record is not None and record.current_processing_time.strip()
                else "待填写"
            )
            has_feedback = record is not None and bool(record.business_feedback.strip())
            feedback_status = "已反馈" if has_feedback else "待反馈"
            feedback_class = "complete" if has_feedback else "muted"
            rows.append(
                f"""
                <tr>
                  <td><strong>{_e(project)}</strong></td>
                  <td><span class="priority-badge priority-{_e(priority.lower())}">{_e(priority)}</span></td>
                  <td><span class="{_e(self._stage_class(stage))}">{_e(stage)}</span></td>
                  <td>{_e(original_time)}</td>
                  <td>{_e(current_time)}</td>
                  <td><span class="status-pill {feedback_class}">{feedback_status}</span></td>
                </tr>
                """
            )
        result = f"""
        <section class="development-overview">
          <div class="dashboard-panel-header">
            <div>
              <h2>开发汇总面板</h2>
              <p>一级页面查看整体进度，具体耗时、反馈和迭代需求在项目跟踪页维护。</p>
            </div>
            <div class="button-row">
              <a class="button accent" href="/development-roadmap">查看详细排期</a>
              <a class="button secondary" href="/project-stages">进入项目跟踪</a>
            </div>
          </div>
          <div class="development-stat-strip">{summary_html}</div>
          <div class="roadmap-snapshot">
            <div class="roadmap-snapshot-title">
              <span class="priority-badge priority-p1">专项</span>
              <div><strong>数据处理 + AI智能Brief</strong><p>2026-07-27 至 2026-10-02，共10周、50个工作日步骤。</p></div>
            </div>
            <div class="roadmap-track data-track"><span>P1 数据处理</span><strong>第1-5周</strong><em>标准数据层 → 日报 → 周报</em></div>
            <div class="roadmap-track brief-track"><span>P2 AI智能Brief</span><strong>第6-10周</strong><em>选品 → 文案/视觉Brief → 质检交付</em></div>
          </div>
          <div class="table-panel development-table-panel">
            <table class="development-table">
              <thead><tr><th>项目</th><th>优先级</th><th>阶段</th><th>原人工耗时</th><th>现在处理耗时</th><th>业务反馈</th></tr></thead>
              <tbody>{''.join(rows)}</tbody>
            </table>
          </div>
        </section>
        """
        assert "开发汇总面板" in result
        assert 'href="/development-roadmap"' in result
        return result

    def _development_roadmap_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        capability_rows = "".join(
            f"""
            <tr>
              <td>{_e(item.area)}</td>
              <td><strong>{_e(item.capability)}</strong></td>
              <td><span class="status-pill {_e(self._roadmap_status_class(item.status))}">{_e(item.status)}</span></td>
              <td>{_e(item.evidence)}</td>
              <td>{_e(item.next_action)}</td>
            </tr>
            """
            for item in CAPABILITY_STATUSES
        )
        week_sections: list[str] = []
        for week in ROADMAP_WEEKS:
            task_rows: list[str] = []
            for task in week.tasks:
                task_date = daily_task_date(week, task).isoformat()
                task_rows.append(
                    f"""
                    <div class="roadmap-day-row">
                      <div class="roadmap-day-label"><span>{_e(task_date)}</span><strong>步骤 {task.day_index} · {_e(task.title)}</strong></div>
                      <div><span class="roadmap-cell-label">你需要做</span><p>{_e(task.business_action)}</p></div>
                      <div><span class="roadmap-cell-label">我负责开发</span><p>{_e(task.developer_action)}</p></div>
                      <div><span class="roadmap-cell-label">当天看得到</span><p>{_e(task.deliverable)}</p></div>
                      <div><span class="roadmap-cell-label">通过标准</span><p>{_e(task.acceptance)}</p></div>
                    </div>
                    """
                )
            open_attribute = " open" if week.week_number == 1 else ""
            week_sections.append(
                f"""
                <details class="roadmap-week"{open_attribute}>
                  <summary>
                    <span class="roadmap-week-number">第{week.week_number}周</span>
                    <span class="roadmap-week-title">{_e(week.objective)}</span>
                    <span class="roadmap-week-date">{week.start_date.isoformat()} 至 {week.end_date.isoformat()}</span>
                  </summary>
                  <div class="roadmap-week-context">
                    <div><span>所属板块</span><strong>{_e(week.workstream)}</strong></div>
                    <div><span>开始条件</span><strong>{_e(week.dependency)}</strong></div>
                    <div><span>周验收结果</span><strong>{_e(week.milestone)}</strong></div>
                  </div>
                  <div class="roadmap-day-list">{''.join(task_rows)}</div>
                </details>
                """
            )
        material_rows = "".join(
            f"""
            <tr>
              <td>{_e(item.workstream)}</td>
              <td><strong>{_e(item.material)}</strong></td>
              <td>{_e(item.business_action)}</td>
              <td>{_e(item.purpose)}</td>
              <td>{item.due_date.isoformat()}</td>
              <td><span class="status-pill {'in-progress' if item.required else 'muted'}">{'必须' if item.required else '建议'}</span></td>
              <td>{_e(item.acceptance)}</td>
              <td>{_e(item.existing_state)}</td>
            </tr>
            """
            for item in MATERIAL_REQUIREMENTS
        )
        body = f"""
        <section class="toolbar roadmap-toolbar">
          <div>
            <span class="badge">零基础执行版</span>
            <h1>数据处理 + AI智能Brief详细开发排期</h1>
            <p>从 2026-07-27 开始，按工作日逐步完成；你负责提供和确认业务资料，我负责代码、数据库、网页和内网部署。</p>
          </div>
          <div class="button-row">
            <a class="button accent" href="/development-roadmap/download">下载Excel排期</a>
            <a class="button secondary" href="/">返回看板</a>
          </div>
        </section>
        <section class="roadmap-kpis">
          <div><span>总周期</span><strong>10周</strong><small>2026-07-27 至 2026-10-02</small></div>
          <div><span>工作日步骤</span><strong>{roadmap_day_count()}</strong><small>每天都有可见结果和验收标准</small></div>
          <div><span>数据处理</span><strong>5周</strong><small>标准数据层、日报、周报</small></div>
          <div><span>AI智能Brief</span><strong>5周</strong><small>选品、文案、视觉、质检、交付</small></div>
        </section>
        <section class="roadmap-beginner-band">
          <div><span>1</span><strong>按模板交资料</strong><p>不要改平台导出文件，只按约定名称上传。</p></div>
          <div><span>2</span><strong>确认业务口径</strong><p>只回答页面列出的选择题和待确认项。</p></div>
          <div><span>3</span><strong>试用并反馈</strong><p>用真实任务跑一遍，填写哪里不对和希望怎么改。</p></div>
          <div><span>4</span><strong>确认上线版本</strong><p>数字、文案和Brief通过后，确认负责人和使用范围。</p></div>
        </section>
        <section class="roadmap-section">
          <div class="section-heading"><h2>当前能力与真实缺口</h2><p>“已有基础”不等于完整上线，后续动作以右侧说明为准。</p></div>
          <div class="table-panel"><table class="data-table roadmap-capability-table">
            <thead><tr><th>板块</th><th>能力</th><th>当前状态</th><th>已有依据</th><th>下一步</th></tr></thead>
            <tbody>{capability_rows}</tbody>
          </table></div>
        </section>
        <section class="roadmap-section">
          <div class="section-heading"><h2>逐周、逐日开发步骤</h2><p>默认只展开第1周；完成一周并通过周验收后，再进入下一周。</p></div>
          <div class="roadmap-timeline">{''.join(week_sections)}</div>
        </section>
        <section class="roadmap-section">
          <div class="section-heading"><h2>你需要准备的素材和资料</h2><p>先处理“必须”项；已有资料不需要重复制作，只需确认是否为最终口径。</p></div>
          <div class="table-panel"><table class="data-table roadmap-material-table">
            <thead><tr><th>板块</th><th>资料</th><th>你需要做</th><th>开发用途</th><th>最晚提供</th><th>级别</th><th>合格标准</th><th>当前情况</th></tr></thead>
            <tbody>{material_rows}</tbody>
          </table></div>
        </section>
        <section class="roadmap-boundary">
          <div><h2>开发边界</h2><p>金额、销量、订单、排名、占比和环比全部由程序计算；AI只负责表达、归纳和基于已确认事实提出建议。</p></div>
          <div><strong>不需要平台API也能先上线</strong><p>业务继续从平台导出Excel/CSV并上传。平台API只影响“自动拉数”，不会阻塞清洗、计算、报表和Brief生产。</p></div>
          <div><strong>缺资料时不猜</strong><p>系统输出“待补资料”或阻断报告，不从历史PPT、参考报表或网络内容复制当期事实。</p></div>
        </section>
        """
        result = self._page("详细开发排期", body)
        assert "数据处理 + AI智能Brief详细开发排期" in result
        assert result.count("roadmap-day-row") == roadmap_day_count()
        return result

    @staticmethod
    def _roadmap_status_class(status: str) -> str:
        if not isinstance(status, str):
            raise TypeError("status must be str")
        mapping = {"已有基础": "complete", "需补强": "developed", "待开发": "in-progress"}
        if status not in mapping:
            raise ValueError("unsupported roadmap status")
        result = mapping[status]
        assert result
        return result

    def _development_roadmap_workbook_path(self) -> Path:
        result = self.config.template_root.parent / "ai_automation_development_plan_latest.xlsx"
        assert result.suffix.lower() == ".xlsx"
        return result

    def _project_stages_page(self, user: UserRecord, success: str, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        feedback_by_project = self.storage.list_project_feedback()
        completed_feedback = self._completed_business_feedback_panel(feedback_by_project, "/project-stages")
        success_html = f"<div class='success'>{_e(success)}</div>" if success else ""
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="toolbar project-stage-toolbar">
          <div>
            <h1>项目开发阶段</h1>
            <p>只维护已开发完整业务反馈；首页表格会从这里提取汇总。</p>
          </div>
          <a class="button secondary" href="/">返回看板</a>
        </section>
        {success_html}{error_html}
        {completed_feedback}
        """
        return self._page("项目开发阶段", body)

    def _ai_settings_page(self, user: UserRecord, success: str, error: str, allow_secret_input: bool) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(allow_secret_input, bool):
            raise TypeError("allow_secret_input must be bool")
        if user.role != "管理员":
            return self._page("无权访问", "<p>只有管理员可以查看AI接口配置。</p>")
        settings = BailianSettings.from_environment()
        status_class = "complete" if settings.is_configured else "in-progress"
        status_text = "已配置" if settings.is_configured else "未配置"
        success_html = f"<div class='success'>{_e(success)}</div>" if success else ""
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        key_form = ""
        if allow_secret_input:
            key_form = """
            <form method="post" action="/admin/ai-settings/save" autocomplete="off">
              <label>百炼通用API Key
                <input type="password" name="api_key" autocomplete="new-password" placeholder="以 sk- 开头" required>
              </label>
              <button class="button secondary" type="submit">保存密钥</button>
            </form>
            """
        else:
            key_form = "<p class='note'>密钥录入框仅在工作台主机的 127.0.0.1 地址显示。</p>"
        body = f"""
        <section class="toolbar">
          <div>
            <h1>AI接口配置</h1>
            <p>仅管理员可见。密钥保存在工作台主机的Windows用户环境变量中。</p>
          </div>
          <a class="button secondary" href="/">返回看板</a>
        </section>
        {success_html}{error_html}
        <section class="split">
          <article>
            <div class="section-heading"><h2>阿里云百炼</h2><span class="status-pill {status_class}">{status_text}</span></div>
            <dl class="settings-list">
              <div><dt>地域</dt><dd>华北2（北京）</dd></div>
              <div><dt>接口协议</dt><dd>OpenAI兼容</dd></div>
              <div><dt>文本模型</dt><dd>{_e(settings.model)}</dd></div>
              <div><dt>密钥状态</dt><dd>{_e(settings.masked_key)}</dd></div>
            </dl>
            <form method="post" action="/admin/ai-settings/test">
              <button class="button" type="submit" {'disabled' if not settings.is_configured else ''}>测试连接</button>
            </form>
          </article>
          <article>
            <h2>本机配置步骤</h2>
            <ol class="setup-steps">
              <li>在百炼控制台创建“通用API Key”。</li>
              <li>在下方密码框粘贴API Key并保存。</li>
              <li>确认左侧显示“已配置”。</li>
              <li>点击“测试连接”。</li>
            </ol>
            {key_form}
            <p class="note">不要把API Key发送到聊天、Excel或业务群。页面不会显示完整密钥。</p>
          </article>
        </section>
        """
        return self._page("AI接口配置", body)

    def _is_loopback_request(self, handler: BaseHTTPRequestHandler) -> bool:
        address = handler.client_address[0]
        if not isinstance(address, str):
            return False
        result = address in {"127.0.0.1", "::1"}
        assert isinstance(result, bool)
        return result

    def _priority_coverage_rows(self) -> str:
        rows: list[str] = []
        for priority, _, _ in PRIORITY_SECTIONS:
            scenarios = [scenario for scenario in self.scenarios.values() if scenario.priority == priority]
            brands = sorted({scenario.brand for scenario in scenarios})
            brand_text = "、".join(brands) if brands else "待接入"
            rows.append(
                f"""
                <tr>
                  <td>{_e(priority)}</td>
                  <td>{len(scenarios)}</td>
                  <td>{_e(brand_text)}</td>
                  <td><a href="/priority/{quote(priority)}">查看</a></td>
                </tr>
                """
            )
        return "".join(rows)

    def _stage_class(self, stage: str) -> str:
        if not isinstance(stage, str) or not stage.strip():
            raise ValueError("stage must not be empty")
        class_by_stage = {
            "开发完成": "status-pill complete",
            "已经开发": "status-pill developed",
            "正在开发": "status-pill in-progress",
        }
        result = class_by_stage.get(stage, "status-pill muted")
        assert result.startswith("status-pill")
        return result

    def _manual_time_by_project(self) -> dict[str, dict[str, str]]:
        path = self.config.template_root / "00_index_dictionary" / "project_manual_hours.csv"
        if not path.exists() or path.stat().st_size == 0:
            return {}
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [dict(row) for row in csv.DictReader(handle)]
        result = {row.get("项目", "").strip(): row for row in rows if row.get("项目", "").strip()}
        assert isinstance(result, dict)
        return result

    def _dashboard_metrics(self) -> tuple[tuple[str, str], ...]:
        job_count = len(self.storage.list_jobs())
        index_count = _csv_record_count(self._archive_index_path())
        dictionary_count = _csv_record_count(self._data_dictionary_path())
        p1_count = len(self._developable_group_items("P1"))
        return (
            ("可提效项目", str(self._developable_project_count())),
            ("P1 项目", str(p1_count)),
            ("资料索引记录", str(index_count)),
            ("数据字典字段", str(dictionary_count)),
            ("最近处理记录", str(job_count)),
            ("当前分级", "P1-P4"),
        )

    def _priority_nav_card(self, priority: str, title: str, description: str) -> str:
        stats = self._priority_development_stats(priority)
        return f"""
        <a class="priority-nav-card" href="/priority/{quote(priority)}">
          <span>{_e(priority)}</span>
          <strong>{_e(title)}</strong>
          <div class="priority-nav-stats">
            <div class="priority-stat-total"><small>项目总数</small><em>{stats.total_count}</em></div>
            <div class="priority-stat-developed"><small>已经开发总数</small><em>{stats.developed_count}</em></div>
            <div class="priority-stat-pending"><small>待开发总数</small><em>{stats.pending_count}</em></div>
          </div>
          <p>{_e(description)}</p>
        </a>
        """

    def _automation_runs_page(self, user: UserRecord, success: str, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        tasks = self.storage.list_automation_tasks()
        runs = self.storage.list_automation_runs(20)
        success_html = f"<div class='success'>{_e(success)}</div>" if success else ""
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        task_rows = "".join(self._automation_task_row(task) for task in tasks)
        if not task_rows:
            task_rows = "<tr><td colspan='10'>暂无任务</td></tr>"
        run_rows = "".join(
            f"""
            <tr>
              <td>{_e(run.created_at)}</td>
              <td>{_e(run.run_date)}</td>
              <td>{_e(run.task_name)}</td>
              <td><span class="status-pill {_e(self._automation_status_class(run.status))}">{_e(self._automation_status_label(run.status))}</span></td>
              <td>{run.downloaded_file_count}</td>
              <td>{run.synced_file_count}</td>
              <td>{_e(run.message)}</td>
              <td>{_e(run.executed_by)}</td>
            </tr>
            """
            for run in runs
        )
        if not run_rows:
            run_rows = "<tr><td colspan='8'>暂无执行记录</td></tr>"
        config = default_meituan_sync_config(Path.cwd())
        body = f"""
        <section class="toolbar automation-toolbar">
          <div>
            <h1>自动化数据执行</h1>
            <p>安排每天需要从平台后台导出的数据任务，并把浏览器插件下载的文件同步到统一入库目录。</p>
          </div>
          <div class="button-row">
            <a class="button secondary" href="/data-foundation">数据入库中心</a>
            <a class="button secondary" href="/">返回首页</a>
          </div>
        </section>
        {success_html}{error_html}
        <section class="dashboard-panel">
          <div class="dashboard-panel-header">
            <div>
              <h2>一键执行</h2>
              <p>选择业务日期后，系统会同步插件下载目录、自动入库到统一基础数据层，并逐项校验四类日报数据是否齐全。</p>
            </div>
          </div>
          <form method="post" action="/automation-runs/execute" class="compact-form">
            <label>业务日期<input type="date" name="run_date" required></label>
            <button class="button" type="submit">执行同步、入库与校验</button>
          </form>
          <p class="note">如果结果显示缺少源数据，请先在美团后台用浏览器插件导出该日期对应报表，再回到这里执行。</p>
        </section>
        <section class="automation-kpis">
          <article><span>启用任务</span><strong>{sum(1 for task in tasks if task.enabled)}</strong><small>按时间顺序执行</small></article>
          <article><span>覆盖平台</span><strong>{len({task.platform for task in tasks})}</strong><small>当前先从美团开始</small></article>
          <article><span>日报字段源</span><strong>4类</strong><small>商品、财务、流量、评价</small></article>
          <article><span>同步目录</span><strong>本机</strong><small>{_e(str(config.target_root))}</small></article>
        </section>
        <section class="automation-layout">
          <article class="automation-main">
            <div class="section-heading">
              <h2>每日任务计划</h2>
              <p>默认预置安踏美团日报四类数据。日报入库后，近7天、周报、月报都从统一基础数据层聚合。</p>
            </div>
            <div class="table-panel">
              <table class="data-table automation-table">
                <thead>
                  <tr><th>时间</th><th>任务</th><th>品牌</th><th>平台</th><th>渠道</th><th>文件类型</th><th>周期</th><th>取数窗口</th><th>状态</th><th>操作</th></tr>
                </thead>
                <tbody>{task_rows}</tbody>
              </table>
            </div>
          </article>
          <aside class="automation-side">
            <h2>新增任务</h2>
            <form method="post" action="/automation-runs/create" class="compact-form">
              <label>任务名称<input name="task_name" value="安踏美团日报-" required></label>
              <label>业务方<input name="business_unit" value="anta_retail_team" required></label>
              <label>品牌ID<input name="brand_id" value="anta_kids" required></label>
              <label>品牌名称<input name="brand_name" value="安踏儿童" required></label>
              <label>平台
                <select name="platform">
                  <option value="meituan">美团</option>
                  <option value="jd">京东</option>
                  <option value="tmall">天猫</option>
                  <option value="mini_program">小程序</option>
                  <option value="official_site">官网</option>
                </select>
              </label>
              <label>渠道
                <select name="channel">
                  <option value="instant_retail">即时零售</option>
                  <option value="ecommerce">电商</option>
                  <option value="private_domain">私域</option>
                  <option value="official_direct">官网直销</option>
                </select>
              </label>
              <label>文件类型
                <select name="file_type">
                  <option value="product_order">商品/订单数据</option>
                  <option value="store_finance">门店财务数据</option>
                  <option value="store_traffic">门店流量数据</option>
                  <option value="service_review">服务评价数据</option>
                </select>
              </label>
              <div class="form-two">
                <label>周期
                  <select name="frequency">
                    <option value="daily">daily</option>
                    <option value="weekly">weekly</option>
                    <option value="monthly">monthly</option>
                  </select>
                </label>
                <label>时间<input name="scheduled_time" value="09:30" required></label>
              </div>
              <label>取数窗口<input name="date_window" value="yesterday" required></label>
              <label>输出目录<input name="output_folder" value="meituan_auto_download/anta_kids/instant_retail" required></label>
              <label>负责人<input name="owner" value="business" required></label>
              <label class="checkbox-line"><input type="checkbox" name="enabled" checked> 启用任务</label>
              <label>备注<textarea name="notes" maxlength="1000" placeholder="说明用途、字段依赖或异常口径"></textarea></label>
              <button class="button" type="submit">新增任务</button>
            </form>
          </aside>
        </section>
        <section class="automation-layout">
          <article class="automation-side">
            <h2>插件与同步</h2>
            <p class="note">Chrome 插件先把文件下载到：{_e(str(config.source_root))}</p>
            <p class="note">本地同步后进入：{_e(str(config.target_root))}</p>
            <form method="post" action="/automation-runs/sync">
              <button class="button secondary" type="submit">仅同步浏览器下载目录</button>
            </form>
            <div class="automation-guide">
              <strong>执行方式</strong>
              <p>1. 业务登录美团后台。</p>
              <p>2. 用插件选择品牌、日期、报表类型。</p>
              <p>3. 手动或自动触发官方导出。</p>
              <p>4. 回到本页点击同步，文件进入统一入库目录。</p>
            </div>
          </article>
          <article class="automation-main">
            <div class="section-heading">
              <h2>最近执行记录</h2>
              <p>记录每天是否下载、同步了多少文件、是否存在异常。后续会继续接入字段校验和入库状态。</p>
            </div>
            <div class="table-panel">
              <table class="data-table automation-table">
                <thead><tr><th>记录时间</th><th>业务日期</th><th>任务</th><th>状态</th><th>下载</th><th>同步</th><th>说明</th><th>执行人</th></tr></thead>
                <tbody>{run_rows}</tbody>
              </table>
            </div>
          </article>
        </section>
        """
        result = self._page("自动化数据执行", body)
        assert "自动化数据执行" in result
        return result

    def _console_dashboard_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not _console_can_open_page(user, "dashboard"):
            return self._page("Forbidden", "<p>Access denied for Developer Console.</p>")
        body = f"""
        {self._console_nav(user, "dashboard")}
        <section class="console-hero">
          <div>
            <span class="console-eyebrow">Developer Console</span>
            <h1>Dashboard</h1>
            <p>Read-only overview for system status, task summary, and recent failures.</p>
          </div>
          <div class="console-status" id="console-dashboard-status">Loading</div>
        </section>
        <section class="console-grid console-grid-four" id="console-task-summary">
          <article><span>Total Tasks</span><strong data-field="total">-</strong></article>
          <article><span>Pending</span><strong data-field="pending">-</strong></article>
          <article><span>Running</span><strong data-field="running">-</strong></article>
          <article><span>Failed</span><strong data-field="failed">-</strong></article>
        </section>
        <section class="console-grid console-grid-two">
          <article class="console-panel">
            <h2>System Status</h2>
            <dl class="console-definition" id="console-system-status">
              <div><dt>Application</dt><dd>Loading</dd></div>
              <div><dt>Database</dt><dd>Loading</dd></div>
              <div><dt>Storage</dt><dd>Loading</dd></div>
              <div><dt>AI</dt><dd>Loading</dd></div>
            </dl>
          </article>
          <article class="console-panel">
            <h2>Recent Failed Tasks</h2>
            <table class="console-table">
              <thead><tr><th>task_id</th><th>task_type</th><th>created_by</th><th>error</th><th>updated_at</th></tr></thead>
              <tbody id="console-failed-tasks"><tr><td colspan="5">Loading</td></tr></tbody>
            </table>
          </article>
        </section>
        {self._console_dashboard_script()}
        """
        page = self._page("Developer Console", body)
        assert "Developer Console" in page
        return page

    def _console_tasks_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not _console_can_open_page(user, "tasks"):
            return self._page("Forbidden", "<p>Access denied for Task Center.</p>")
        body = f"""
        {self._console_nav(user, "tasks")}
        <section class="toolbar console-toolbar">
          <div>
            <span class="console-eyebrow">Task Center</span>
            <h1>Task Center</h1>
            <p>Read-only task list scoped by the current user's permissions.</p>
          </div>
          <a class="button secondary" href="/console">Back to Console</a>
        </section>
        <section class="console-panel">
          <form class="console-filter-bar" id="console-task-filter-form">
            <label>Task Type<input name="task_type" placeholder="REPORT_GENERATE"></label>
            <label>Status
              <select name="status">
                <option value="">All</option>
                <option value="pending">pending</option>
                <option value="running">running</option>
                <option value="success">success</option>
                <option value="failed">failed</option>
                <option value="cancelled">cancelled</option>
              </select>
            </label>
            <label>Created By<input name="created_by" placeholder="username"></label>
            <button class="button" type="submit">Filter</button>
            <button class="button secondary" type="reset">Clear</button>
          </form>
          <div class="console-table-meta" id="console-task-list-meta">Loading</div>
          <table class="console-table console-task-table">
            <thead><tr><th>task_id</th><th>task_type</th><th>status</th><th>created_by</th><th>created_time</th><th>error</th><th>asset</th><th>detail</th></tr></thead>
            <tbody id="console-task-list"><tr><td colspan="8">Loading</td></tr></tbody>
          </table>
        </section>
        {self._console_tasks_script()}
        """
        page = self._page("Task Center", body)
        assert "Task Center" in page
        return page

    def _console_environment_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not _console_can_open_page(user, "environment"):
            return self._page("Forbidden", "<p>Access denied for Environment Center.</p>")
        body = f"""
        {self._console_nav(user, "environment")}
        <section class="toolbar console-toolbar">
          <div>
            <span class="console-eyebrow">Environment Center</span>
            <h1>Environment Center</h1>
            <p>Read-only runtime configuration status. Environment variables are not editable here.</p>
          </div>
          <a class="button secondary" href="/console">Back to Console</a>
        </section>
        <section class="console-panel">
          <dl class="console-definition" id="console-environment-status">
            <div><dt>APP_ENV</dt><dd>Loading</dd></div>
            <div><dt>DATABASE_BACKEND</dt><dd>Loading</dd></div>
            <div><dt>REPORT_TASK_MODE</dt><dd>Loading</dd></div>
            <div><dt>AI_PROVIDER</dt><dd>Loading</dd></div>
            <div><dt>Storage</dt><dd>Loading</dd></div>
          </dl>
        </section>
        {self._console_environment_script()}
        """
        page = self._page("Environment Center", body)
        assert "Environment Center" in page
        return page

    def _console_nav(self, user: UserRecord, active: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(active, str) or not active.strip():
            raise ValueError("active must not be empty")
        links = (
            ("dashboard", "/console", "Dashboard"),
            ("tasks", "/console/tasks", "Task Center"),
            ("environment", "/console/environment", "Environment Center"),
        )
        link_html = "".join(
            f'<a class="console-nav-link {"active" if key == active else ""}" href="{href}">{label}</a>'
            for key, href, label in links
            if key != "environment" or _console_role_key(user.role) in {"admin", "developer"}
        )
        result = f"""
        <section class="console-nav">
          <div>
            <strong>AI Automation Platform</strong>
            <span>{_e(user.display_name)} - {_e(user.role)}</span>
          </div>
          <nav>{link_html}<a class="console-nav-link" href="/">Business Home</a></nav>
        </section>
        """
        assert "console-nav" in result
        return result

    @staticmethod
    def _console_dashboard_script() -> str:
        return """
        <script>
        (() => {
          const text = (value) => value === null || value === undefined || value === "" ? "-" : String(value);
          const statusText = (item) => item && typeof item === "object" ? `${text(item.status)} ${text(item.message)}` : text(item);
          fetch("/api/console/dashboard", {credentials: "same-origin"})
            .then((response) => response.ok ? response.json() : Promise.reject(new Error(String(response.status))))
            .then((data) => {
              document.querySelector("#console-dashboard-status").textContent = "Connected";
              const summary = data.task_summary || {};
              document.querySelectorAll("#console-task-summary [data-field]").forEach((node) => {
                node.textContent = text(summary[node.dataset.field]);
              });
              const system = data.system_status || {};
              document.querySelector("#console-system-status").innerHTML = [
                ["Application", statusText(system.application)],
                ["Database", statusText(system.database)],
                ["Storage", statusText(system.storage)],
                ["AI", statusText(system.ai)]
              ].map(([key, value]) => `<div><dt>${key}</dt><dd>${value}</dd></div>`).join("");
              const failed = Array.isArray(data.recent_failed_tasks) ? data.recent_failed_tasks : [];
              document.querySelector("#console-failed-tasks").innerHTML = failed.length
                ? failed.map((task) => `<tr><td>${text(task.task_id)}</td><td>${text(task.task_type)}</td><td>${text(task.created_by)}</td><td>${text(task.error)}</td><td>${text(task.updated_at)}</td></tr>`).join("")
                : '<tr><td colspan="5">No failed tasks</td></tr>';
            })
            .catch(() => {
              document.querySelector("#console-dashboard-status").textContent = "Unavailable";
              document.querySelector("#console-failed-tasks").innerHTML = '<tr><td colspan="5">No permission or API unavailable</td></tr>';
            });
        })();
        </script>
        """

    @staticmethod
    def _console_tasks_script() -> str:
        return """
        <script>
        (() => {
          const form = document.querySelector("#console-task-filter-form");
          const list = document.querySelector("#console-task-list");
          const meta = document.querySelector("#console-task-list-meta");
          const text = (value) => value === null || value === undefined || value === "" ? "-" : String(value);
          const html = (value) => text(value).replace(/[&<>'"]/g, (char) => ({"&":"&amp;","<":"&lt;",">":"&gt;","'":"&#39;",'"':"&quot;"}[char]));
          const statusClass = (value) => {
            const status = text(value).toLowerCase();
            if (["success", "failed", "running", "pending", "cancelled"].includes(status)) return `console-status-${status}`;
            return "console-status-unknown";
          };
          const errorSummary = (task) => task && task.error ? html(task.error).slice(0, 120) : "-";
          const assetStatus = (task) => {
            const asset = task && task.result_asset && typeof task.result_asset === "object" ? task.result_asset : null;
            if (!asset || !asset.filename) return '<span class="console-asset-missing">No file</span>';
            return `<span class="console-asset-ready">${html(asset.filename)}</span>`;
          };
          const taskUrl = (task) => `/tasks/${encodeURIComponent(text(task.task_id))}`;
          const queryString = () => {
            const params = new URLSearchParams();
            new FormData(form).forEach((value, key) => {
              const normalized = String(value || "").trim();
              if (normalized) params.set(key, normalized);
            });
            const textValue = params.toString();
            return textValue ? `?${textValue}` : "";
          };
          const renderTasks = (tasks) => {
            meta.textContent = `${tasks.length} visible task${tasks.length === 1 ? "" : "s"}`;
            list.innerHTML = tasks.length
              ? tasks.map((task) => `<tr><td>${html(task.task_id)}</td><td>${html(task.task_type)}</td><td><span class="console-task-status ${statusClass(task.status)}">${html(task.status)}</span></td><td>${html(task.created_by)}</td><td>${html(task.created_time)}</td><td>${errorSummary(task)}</td><td>${assetStatus(task)}</td><td><a class="button secondary console-detail-button" href="${taskUrl(task)}">View</a></td></tr>`).join("")
              : '<tr><td colspan="8">No task records</td></tr>';
          };
          const loadTasks = () => {
            meta.textContent = "Loading";
            fetch(`/api/tasks${queryString()}`, {credentials: "same-origin"})
              .then((response) => response.ok ? response.json() : Promise.reject(new Error(String(response.status))))
              .then((data) => renderTasks(Array.isArray(data.tasks) ? data.tasks : []))
              .catch(() => {
                meta.textContent = "Unavailable";
                list.innerHTML = '<tr><td colspan="8">No permission or API unavailable</td></tr>';
              });
          };
          form.addEventListener("submit", (event) => {
            event.preventDefault();
            loadTasks();
          });
          form.addEventListener("reset", () => window.setTimeout(loadTasks, 0));
          loadTasks();
        })();
        </script>
        """

    @staticmethod
    def _console_environment_script() -> str:
        return """
        <script>
        (() => {
          const text = (value) => value === null || value === undefined || value === "" ? "-" : String(value);
          const storageSummary = (storage) => {
            if (!storage || typeof storage !== "object") return "-";
            const provider = text(storage.provider);
            const resultDir = storage.result_dir && typeof storage.result_dir === "object" ? `result writable: ${text(storage.result_dir.writable)}` : "";
            return `${provider} ${resultDir}`.trim();
          };
          fetch("/api/system/config/status", {credentials: "same-origin"})
            .then((response) => response.ok ? response.json() : Promise.reject(new Error(String(response.status))))
            .then((data) => {
              document.querySelector("#console-environment-status").innerHTML = [
                ["APP_ENV", data.app_env],
                ["DATABASE_BACKEND", data.database_backend],
                ["REPORT_TASK_MODE", data.report_task_mode],
                ["AI_PROVIDER", `${text(data.ai_provider)} / ${text(data.ai_model)} / key: ${text(data.ai_api_key_configured)}`],
                ["Storage", storageSummary(data.storage)]
              ].map(([key, value]) => `<div><dt>${key}</dt><dd>${text(value)}</dd></div>`).join("");
            })
            .catch(() => {
              document.querySelector("#console-environment-status").innerHTML = '<div><dt>Status</dt><dd>No permission or API unavailable</dd></div>';
            });
        })();
        </script>
        """

    def _tasks_page(self, user: UserRecord) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        tasks = self._permission_service().filter_visible_tasks(user, self._task_query_service().list_tasks())
        rows = "".join(
            f"""
            <tr>
              <td><a href="/tasks/{task.task_id}">{task.task_id}</a></td>
              <td>{_e(task.task_type)}</td>
              <td><span class="badge">{_e(task.status)}</span></td>
              <td>{_e(task.created_by)}</td>
              <td>{_e(task.created_time)}</td>
              <td>{_e(_task_result_summary_text(task.result))}</td>
            </tr>
            """
            for task in tasks
        )
        if not rows:
            rows = "<tr><td colspan='6'>暂无任务记录</td></tr>"
        body = f"""
        <section class="toolbar">
          <div>
            <h1>任务状态</h1>
            <p>提交人：{_e(user.display_name)} · 最近任务 {len(tasks)} 条</p>
          </div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        <section>
          <article>
            <h2>最近任务</h2>
            <table>
              <thead>
                <tr><th>task_id</th><th>task_type</th><th>status</th><th>created_by</th><th>created_time</th><th>result</th></tr>
              </thead>
              <tbody>{rows}</tbody>
            </table>
          </article>
        </section>
        """
        page = self._page("任务状态", body)
        assert page.strip()
        return page

    def _task_detail_page(self, user: UserRecord, path: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        try:
            task_id = _task_id_from_page_path(path)
            task = self._task_query_service().get_task(task_id)
            if task is None:
                return self._page("?????", "<p>??????</p>")
            if not self._permission_service().can_view_task(user, task):
                return self._page("forbidden", "<p>forbidden</p>")
            error_html = f"<p class='error'>{_e(task.error)}</p>" if task.error else "<p class='note'>?????</p>"
            result_rows = "".join(
                f"<li><strong>{_e(key)}</strong><span>{_e(value)}</span></li>"
                for key, value in _flatten_result(task.result).items()
            )
            if not result_rows:
                result_rows = "<li><strong>result</strong><span>暂无结果。</span></li>"
            body = f"""
            <section class="toolbar">
              <div>
                <h1>任务详情 #{task.task_id}</h1>
                <p>{_e(task.task_type)} ? {_e(task.status)} ? {_e(task.created_by)}</p>
              </div>
              <div class="button-row">
                <a class="button secondary" href="/console/tasks">Console Task Center</a>
                <a class="button secondary" href="/tasks">返回任务列表</a>
              </div>
            </section>
            <section class="split">
              <article>
                <h2>任务信息</h2>
                <ul class="metrics">
                  <li><strong>task_id</strong><span>{task.task_id}</span></li>
                  <li><strong>task_type</strong><span>{_e(task.task_type)}</span></li>
                  <li><strong>created_by</strong><span>{_e(task.created_by)}</span></li>
                  <li><strong>created_time</strong><span>{_e(task.created_time)}</span></li>
                  <li><strong>status</strong><span>{_e(task.status)}</span></li>
                  <li><strong>updated_time</strong><span>{_e(task.updated_at or task.created_time)}</span></li>
                </ul>
                {error_html}
              </article>
              <article>
                <h2>执行结果</h2>
                <ul class="metrics">{result_rows}</ul>
              </article>
            </section>
            <section class="console-grid console-grid-two task-diagnostics-grid">
              {self._task_execution_flow_panel(task)}
              {self._task_result_asset_panel(user, task)}
            </section>
            {self._task_error_diagnostics_panel(task)}
            """
            return self._page("任务详情", body)
        except (ValueError, TypeError) as exc:
            return self._page("??????", f"<p>{_e(exc)}</p>")

    def _task_execution_flow_panel(self, task: object) -> str:
        task_type = _task_text(task, "task_type")
        steps = _task_execution_steps(task_type)
        items = "".join(
            f"<li><span>{index}</span><strong>{_e(step)}</strong></li>"
            for index, step in enumerate(steps, start=1)
        )
        result = f"""
        <article class="console-panel task-flow-panel">
          <h2>Execution Flow</h2>
          <p class="note">Static diagnostic path for task type: {_e(task_type)}</p>
          <ol class="task-flow-list">{items}</ol>
        </article>
        """
        assert "Execution Flow" in result
        return result

    def _task_result_asset_panel(self, user: UserRecord, task: object) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        task_id = _task_id_value(task)
        status = _task_text(task, "status")
        filename = ""
        available = False
        if status == "success" and self._permission_service().can_download_task(user, task):
            try:
                view = self._task_result_service().get_result(task_id)
                filename = view.filename
                available = True
            except (FileNotFoundError, ValueError, TypeError, PermissionError):
                available = False
        download_html = self._task_download_button(user, task)
        status_text = "可下载" if available else _task_download_status_text(status)
        result = f"""
        <article class="console-panel task-asset-panel">
          <h2>结果文件</h2>
          <dl class="console-definition">
            <div><dt>结果文件</dt><dd>{_e(filename or "-")}</dd></div>
            <div><dt>状态</dt><dd>{_e(status_text)}</dd></div>
          </dl>
          {download_html}
        </article>
        """
        assert "结果文件" in result
        return result

    def _task_error_diagnostics_panel(self, task: object) -> str:
        status = _task_text(task, "status")
        error = _task_text(task, "error") or "-"
        created_time = _task_text(task, "created_time")
        updated_time = _task_text(task, "updated_at") or created_time
        result = f"""
        <section class="console-panel task-error-panel">
          <h2>Error Diagnostics</h2>
          <dl class="console-definition">
            <div><dt>status</dt><dd>{_e(status)}</dd></div>
            <div><dt>error message</dt><dd>{_e(error)}</dd></div>
            <div><dt>created_time</dt><dd>{_e(created_time)}</dd></div>
            <div><dt>updated_time</dt><dd>{_e(updated_time)}</dd></div>
          </dl>
        </section>
        """
        assert "Error Diagnostics" in result
        return result

    def _task_download_button(self, user: UserRecord, task: object) -> str:
        task_id = _task_id_value(task)
        status = _task_text(task, "status")
        if status in {"pending", "running"}:
            return "<p class='note'>任务完成后可下载</p>"
        if status == "failed":
            return "<p class='note'>任务失败，无结果文件</p>"
        if status != "success":
            return "<p class='note'>暂无可下载文件</p>"
        if not self._permission_service().can_download_task(user, task):
            return "<p class='note'>结果文件不存在</p>"
        try:
            result = self._task_result_service().get_result(task_id)
        except (FileNotFoundError, ValueError, TypeError, PermissionError):
            return "<p class='note'>结果文件不存在</p>"
        return f"<a class=\"button\" href=\"/api/tasks/{task_id}/download\">下载结果 CSV：{_e(result.filename)}</a>"

    def _automation_task_row(self, task: AutomationTaskRecord) -> str:
        if not isinstance(task, AutomationTaskRecord):
            raise TypeError("task must be AutomationTaskRecord")
        status_class = "complete" if task.enabled else "muted"
        status_text = "启用" if task.enabled else "停用"
        next_enabled = "0" if task.enabled else "1"
        next_label = "停用" if task.enabled else "启用"
        today = date.today().isoformat()
        return f"""
        <tr>
          <td><strong>{_e(task.scheduled_time)}</strong></td>
          <td>{_e(task.task_name)}<small>{_e(task.notes)}</small></td>
          <td>{_e(task.brand_name)}</td>
          <td>{_e(self._platform_label(task.platform))}</td>
          <td>{_e(self._channel_label(task.channel))}</td>
          <td>{_e(self._file_type_label(task.file_type))}</td>
          <td>{_e(task.frequency)}</td>
          <td>{_e(task.date_window)}</td>
          <td><span class="status-pill {status_class}">{status_text}</span></td>
          <td>
            <form method="post" action="/automation-runs/toggle" class="inline-form">
              <input type="hidden" name="task_id" value="{task.id}">
              <input type="hidden" name="enabled" value="{next_enabled}">
              <button class="link-button" type="submit">{next_label}</button>
            </form>
            <form method="post" action="/automation-runs/record" class="inline-form">
              <input type="hidden" name="task_id" value="{task.id}">
              <input type="hidden" name="run_date" value="{today}">
              <input type="hidden" name="status" value="manual_done">
              <input type="hidden" name="downloaded_file_count" value="1">
              <input type="hidden" name="synced_file_count" value="0">
              <input type="hidden" name="message" value="已手动确认完成下载，待同步入库目录。">
              <button class="link-button" type="submit">标记完成</button>
            </form>
          </td>
        </tr>
        """

    @staticmethod
    def _automation_status_class(status: str) -> str:
        if not isinstance(status, str) or not status.strip():
            raise ValueError("status must not be empty")
        mapping = {
            "synced": "complete",
            "foundation_ready": "complete",
            "manual_done": "developed",
            "failed": "danger",
            "missing_source": "danger",
            "unsupported": "muted",
            "pending": "in-progress",
        }
        return mapping.get(status, "muted")

    @staticmethod
    def _automation_status_label(status: str) -> str:
        if not isinstance(status, str) or not status.strip():
            raise ValueError("status must not be empty")
        mapping = {
            "synced": "已同步",
            "foundation_ready": "基础层就绪",
            "manual_done": "已下载",
            "failed": "失败",
            "missing_source": "缺少源数据",
            "unsupported": "未接入",
            "pending": "待执行",
        }
        return mapping.get(status, status)

    @staticmethod
    def _platform_label(platform: str) -> str:
        mapping = {"meituan": "美团", "jd": "京东", "tmall": "天猫", "mini_program": "小程序", "official_site": "官网"}
        return mapping.get(platform, platform)

    @staticmethod
    def _channel_label(channel: str) -> str:
        mapping = {"instant_retail": "即时零售", "ecommerce": "电商", "private_domain": "私域", "official_direct": "官网直销"}
        return mapping.get(channel, channel)

    @staticmethod
    def _file_type_label(file_type: str) -> str:
        mapping = {
            "product_order": "商品/订单数据",
            "store_finance": "门店财务数据",
            "store_traffic": "门店流量数据",
            "service_review": "服务评价数据",
        }
        return mapping.get(file_type, file_type)

    def _data_foundation_page(self, user: UserRecord, success: str, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        success_html = f"<div class='success'>{_e(success)}</div>" if success else ""
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="toolbar">
          <div>
            <h1>数据入库中心</h1>
            <p>先识别和校验原始文件，只有通过规则的数据才允许进入统一基础数据层。</p>
          </div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        {success_html}
        {error_html}
        <section class="split">
          <article>
            <h2>上传并校验</h2>
            <form method="post" action="/data-foundation/check" enctype="multipart/form-data">
              <label>业务方
                <input name="business_unit" value="anta_retail_team" required>
              </label>
              <label>品牌ID
                <input name="brand_id" value="anta_kids" required>
              </label>
              <label>品牌名称
                <input name="brand_name" value="安踏儿童" required>
              </label>
              <label>平台
                <select name="platform">
                  <option value="meituan">美团</option>
                  <option value="jd">京东</option>
                  <option value="tmall">天猫</option>
                  <option value="mini_program">小程序</option>
                  <option value="official_site">官网</option>
                </select>
              </label>
              <label>渠道
                <select name="channel">
                  <option value="instant_retail">即时零售</option>
                  <option value="ecommerce">电商</option>
                  <option value="private_domain">私域</option>
                  <option value="official_direct">官网直销</option>
                </select>
              </label>
              <label>项目
                <input name="project_code" value="p1_p2_anta_meituan" required>
              </label>
              <label>文件类型
                <select name="declared_file_type">
                  <option value="product_order">商品/订单数据</option>
                  <option value="store_finance">门店财务数据</option>
                  <option value="store_traffic">门店流量数据</option>
                  <option value="service_review">服务评价数据</option>
                </select>
              </label>
              <label>开始日期
                <input name="data_start_date" value="20260720" required>
              </label>
              <label>结束日期
                <input name="data_end_date" value="20260726" required>
              </label>
              <label>原始数据文件
                <input type="file" name="data_file" accept=".csv,.xlsx" required>
              </label>
              <button class="button" type="submit">执行入库校验</button>
            </form>
          </article>
          <article>
            <h2>当前规则</h2>
            <ul class="setup-steps">
              <li>原始文件只存档，不直接参与 P1/P2 计算。</li>
              <li>系统按平台和渠道归属数据，同一张基础表通过 platform + channel 区分来源。</li>
              <li>系统按表头识别文件类型，并校验业务方声明的文件类型是否一致。</li>
              <li>字段先映射到统一字段，再做空值、金额、整数和品牌归属校验。</li>
              <li>美团映射已可用；京东、天猫、小程序、官网会进入“映射待配置”，不会误入库。</li>
              <li>品牌分大于等于 90 自动通过；70-89 待人工确认；低于 70 拒绝入库。</li>
              <li>第一版只做预入库校验，不直接写入最终 fact 表。</li>
            </ul>
          </article>
        </section>
        """
        return self._page("数据入库中心", body)

    def _data_foundation_result_page(
        self,
        user: UserRecord,
        import_batch_id: str,
        original_file_name: str,
        plan: IngestionPlan,
    ) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(plan, IngestionPlan):
            raise TypeError("plan must be IngestionPlan")
        for field_name, field_value in (("import_batch_id", import_batch_id), ("original_file_name", original_file_name)):
            if not isinstance(field_value, str) or not field_value.strip():
                raise ValueError(f"{field_name} must not be empty")
        decision_label = {
            "auto_pass": "自动通过",
            "manual_review": "待人工确认",
            "reject": "拒绝入库",
        }[plan.brand_match.decision]
        validation_label = "通过" if plan.validation.passed else "未通过"
        metric_items = (
            ("批次", import_batch_id),
            ("文件", original_file_name),
            ("平台", plan.metadata.platform),
            ("渠道", plan.metadata.channel),
            ("识别类型", plan.recognition.file_type),
            ("识别置信度", str(plan.recognition.confidence)),
            ("校验状态", validation_label),
            ("品牌匹配分", str(plan.brand_match.total_score)),
            ("品牌判定", decision_label),
            ("目标基础表", plan.target_table),
            ("标准化行数", str(len(plan.normalized_rows))),
        )
        metrics_html = "".join(
            f"<li><strong>{_e(label)}</strong><span>{_e(value)}</span></li>"
            for label, value in metric_items
        )
        messages_html = self._foundation_messages(plan)
        preview_rows = []
        for row in plan.normalized_rows[:10]:
            preview_rows.append(
                "<tr>"
                f"<td>{_e(row.get('source_row_number', ''))}</td>"
                f"<td>{_e(row.get('store_name', ''))}</td>"
                f"<td>{_e(row.get('store_id', ''))}</td>"
                f"<td>{_e(row.get('sku_code', row.get('upc_code', '')))}</td>"
                f"<td>{_e(row.get('product_name', row.get('order_products', '')))}</td>"
                f"<td>{_e(row.get('paid_sales_amount', row.get('paid_transaction_amount', '')))}</td>"
                "</tr>"
            )
        body = f"""
        <section class="toolbar">
          <div>
            <h1>入库校验结果</h1>
            <p>本次结果已写入批次表和校验报告表。第一版不会直接写入最终事实表。</p>
          </div>
          <div class="button-row">
            <a class="button" href="/data-foundation">继续校验文件</a>
            <a class="button secondary" href="/">返回首页</a>
          </div>
        </section>
        <section class="split">
          <article>
            <h2>结果概览</h2>
            <ul class="metrics">{metrics_html}</ul>
          </article>
          <article>
            <h2>规则消息</h2>
            {messages_html}
          </article>
        </section>
        <section>
          <div class="section-heading">
            <h2>标准化字段预览</h2>
            <p>仅展示前 10 行，后续确认入库时会按这些标准字段写入统一基础数据层。</p>
          </div>
          <div class="table-panel">
            <table class="data-table">
              <thead><tr><th>行号</th><th>门店</th><th>门店ID</th><th>SKU/UPC</th><th>商品/评价商品</th><th>金额</th></tr></thead>
              <tbody>{''.join(preview_rows)}</tbody>
            </table>
          </div>
        </section>
        """
        page = self._page("入库校验结果", body)
        assert import_batch_id in page
        return page

    def _foundation_messages(self, plan: IngestionPlan) -> str:
        if not isinstance(plan, IngestionPlan):
            raise TypeError("plan must be IngestionPlan")
        messages: list[tuple[str, str]] = []
        messages.extend(("错误", message) for message in plan.validation.errors)
        messages.extend(("提醒", message) for message in plan.validation.warnings)
        messages.extend(("提醒", message) for message in plan.brand_match.warnings)
        if not messages:
            messages.append(("通过", "字段、格式和品牌归属校验均通过。"))
        items = "".join(
            f"<li><strong>{_e(level)}</strong><span>{_e(message)}</span></li>"
            for level, message in messages
        )
        result = f"<ul class='metrics'>{items}</ul>"
        assert result.strip()
        return result

    def _priority_page(self, user: UserRecord, priority: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        section = next((item for item in PRIORITY_SECTIONS if item[0] == priority), None)
        if section is None:
            raise ValueError("priority must be P1-P4")
        _, title, description = section
        cards = "".join(self._scenario_card(key) for key, scenario in self.scenarios.items() if scenario.priority == priority)
        if not cards:
            cards = "<div class='empty-state'>暂无接入项目</div>"
        sibling_links = "".join(
            f"<a class='button {'accent' if item[0] == priority else 'secondary'}' href='/priority/{quote(item[0])}'>{_e(item[0])}</a>"
            for item in PRIORITY_SECTIONS
        )
        featured_entry = self._priority_featured_entry(priority)
        body = f"""
        <section class="toolbar">
          <div>
            <h1>{_e(priority)} · {_e(title)}</h1>
            <p>{_e(description)} 当前登录：{_e(user.display_name)}</p>
          </div>
          <div class="toolbar-actions">
            <a class="button secondary" href="/">返回看板</a>
          </div>
        </section>
        <section class="level-switch">{sibling_links}</section>
        {featured_entry}
        <section class="priority-grid">{cards}</section>
        """
        return self._page(f"{priority}项目", body)

    def _priority_featured_entry(self, priority: str) -> str:
        if not isinstance(priority, str) or not priority.strip():
            raise ValueError("priority must not be empty")
        if priority != "P2":
            return ""
        return """
        <section class="priority-featured-entry priority-border-p2">
          <div>
            <span class="badge">P2</span>
            <h2>P2内容生产中心</h2>
            <p>统一承接 AI选品、人群场景、卖点提炼、文案生成、视觉Brief 和质检交付包。</p>
          </div>
          <a class="button" href="/p2-content-center">进入内容生产中心</a>
        </section>
        """

    def _archive_intake_page(self, user: UserRecord, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        pending_dir = self.config.template_root / "00_intake" / "01_pending"
        index_path = self.config.template_root / "00_index_dictionary" / "archive_index.csv"
        dictionary_path = self.config.template_root / "00_index_dictionary" / "data_dictionary.csv"
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="toolbar">
          <div>
            <h1>资料投递入口</h1>
            <p>上传后自动归档、登记索引、更新数据字典。</p>
          </div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        {error_html}
        <section class="split">
          <article>
            <h2>直接上传</h2>
            <form method="post" action="/archive-intake/upload" enctype="multipart/form-data">
              <label>选择文件<input type="file" name="files" multiple required></label>
              <button class="button" type="submit">上传并自动归档</button>
            </form>
          </article>
          <article>
            <h2>文件夹扫描</h2>
            <p>如果文件很多，也可以先批量放到待处理文件夹，再点击扫描。</p>
            <form method="post" action="/archive-intake/run">
              <button class="button secondary" type="submit">扫描待处理文件夹</button>
            </form>
            <p class="note">待处理：{_e(str(pending_dir))}</p>
          </article>
        </section>
        <section class="split">
          <article>
            <h2>刷新本地资料包</h2>
            <p>把资料包里已经放好的最新业务文件重新登记到资料索引和数据字典，不移动原文件。</p>
            <form method="post" action="/archive-catalog/rebuild">
              <button class="button accent" type="submit">刷新资料索引和数据字典</button>
            </form>
          </article>
          <article>
            <h2>刷新范围</h2>
            <p>会扫描 01 数据处理、02 品牌内容资料、03 配置自动化资料、04 页面巡检复盘等本地目录；00 提交文件夹和 00 资料索引表不会重复纳入。</p>
          </article>
        </section>
        <section class="split">
          <article>
            <h2>命名建议</h2>
            <p>建议文件名包含品牌、平台、项目、资料类型、日期。例：安踏_美团_上下架_商品下载_20260713.xlsx</p>
          </article>
          <article>
            <h2>登记位置</h2>
            <div class="button-row">
              <a class="button accent" href="/archive-index">打开资料索引</a>
              <a class="button secondary" href="/data-dictionary">打开数据字典</a>
            </div>
            <p class="note">索引：{_e(str(index_path))}<br>数据字典：{_e(str(dictionary_path))}</p>
          </article>
        </section>
        """
        return self._page("资料投递入口", body)

    def _archive_csv_page(self, user: UserRecord, title: str, path: Path, description: str, download_path: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        for value_name, value in (("title", title), ("description", description), ("download_path", download_path)):
            if not isinstance(value, str) or not value.strip():
                raise ValueError(f"{value_name} must not be empty")
        if not isinstance(path, Path):
            raise TypeError("path must be pathlib.Path")
        ensure_intake_workspace(ArchiveIntakeConfig(self.config.template_root))
        headers, rows = _read_csv_table(path)
        table_html = self._csv_table(headers, rows)
        body = f"""
        <section class="toolbar ledger-toolbar">
          <div>
            <h1>{_e(title)}</h1>
            <p>{_e(description)} 当前登录：{_e(user.display_name)}</p>
          </div>
          <div class="toolbar-actions">
            <a class="button accent" href="{_e(download_path)}">下载 CSV</a>
            <a class="button secondary" href="/archive-intake">投递资料</a>
            <a class="button secondary" href="/">返回首页</a>
          </div>
        </section>
        <section class="ledger-summary">
          <article>
            <span class="summary-label">记录数</span>
            <strong>{len(rows)}</strong>
          </article>
          <article>
            <span class="summary-label">字段数</span>
            <strong>{len(headers)}</strong>
          </article>
          <article>
            <span class="summary-label">来源文件</span>
            <strong>{_e(path.name)}</strong>
          </article>
        </section>
        <section class="table-panel">
          {table_html}
        </section>
        """
        return self._page(title, body)

    def _p2_content_center_page(self, user: UserRecord, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        settings = BailianSettings.from_environment()
        ai_status = "已配置" if settings.is_configured else "未配置"
        ai_class = "complete" if settings.is_configured else "in-progress"
        body = f"""
        <section class="toolbar p2-toolbar">
          <div>
            <h1>P2内容生产中心</h1>
            <p>从统一基础数据层生成 AI选品、人群场景、卖点、文案、视觉Brief 和质检结果。</p>
          </div>
          <div class="button-row">
            <a class="button secondary" href="/admin/ai-settings">AI接口配置</a>
            <a class="button secondary" href="/">返回看板</a>
          </div>
        </section>
        {error_html}
        <section class="p2-command-panel">
          <article class="p2-run-card">
            <div class="section-heading">
              <h2>生成内容交付包</h2>
              <span class="status-pill {ai_class}">AI接口{ai_status}</span>
            </div>
            <form method="post" action="/p2-content-center/run" class="p2-grid-form">
              <label>品牌
                <select name="brand_id">
                  <option value="anta_kids">安踏儿童</option>
                </select>
              </label>
              <label>平台
                <select name="platform">
                  <option value="meituan">美团</option>
                </select>
              </label>
              <label>渠道
                <select name="channel">
                  <option value="instant_retail">即时零售</option>
                </select>
              </label>
              <label>内容任务
                <select name="task_type">
                  <option value="social_copy">社群文案</option>
                  <option value="xiaohongshu_copy">小红书文案</option>
                  <option value="poster_copy">海报文案</option>
                  <option value="selection_brief">选品Brief</option>
                </select>
              </label>
              <label>开始日期<input type="date" name="start_date" required></label>
              <label>结束日期<input type="date" name="end_date" required></label>
              <label>输出数量<input type="number" name="output_count" min="1" max="20" value="5" required></label>
              <button class="button p2-submit" type="submit">生成P2交付包</button>
            </form>
            <p class="note">正式输出只读取统一基础数据层。若缺少日期数据，请先用美团插件导出并执行入库。</p>
          </article>
          <aside class="p2-rule-card">
            <h2>当前固定流程</h2>
            <ol class="setup-steps">
              <li>业务选择品牌、渠道和日期，不再上传文案模板。</li>
              <li>系统同步插件下载目录，并把通过校验的数据写入统一基础数据层。</li>
              <li>系统按销售额、销量、订单数和评价摘录筛选候选商品。</li>
              <li>AI只基于商品事实、品牌资料和禁用词输出文案与Brief。</li>
              <li>质检结果和历史版本写入处理记录，可下载CSV交付包。</li>
            </ol>
          </aside>
        </section>
        <section class="p2-flow">
          <article><span>1</span><strong>选品Agent</strong><small>销售额 / 销量 / 订单 / 评价</small></article>
          <article><span>2</span><strong>人群场景Agent</strong><small>目标人群与使用场景</small></article>
          <article><span>3</span><strong>卖点Agent</strong><small>仅提炼可证明卖点</small></article>
          <article><span>4</span><strong>文案Agent</strong><small>标题 / 正文 / CTA</small></article>
          <article><span>5</span><strong>视觉BriefAgent</strong><small>尺寸 / 层级 / 素材</small></article>
          <article><span>6</span><strong>质检Agent</strong><small>禁用词 / 事实 / 待确认</small></article>
        </section>
        """
        page = self._page("P2内容生产中心", body)
        assert "P2内容生产中心" in page
        return page

    def _anta_reporting_page(self, user: UserRecord, error: str) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        weekly_dir = self.config.template_root / "01_data_processing" / "01-3_weekly_report" / "anta_weekly_report" / "01_raw_data"
        monthly_dir = self.config.template_root / "01_data_processing" / "01-4_monthly_report" / "anta_monthly_report" / "01_raw_data"
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="toolbar">
          <div>
            <h1>安踏周报/月报</h1>
            <p>从资料包已归档源数据生成报表初稿，先输出核心指标和TOP商品。</p>
          </div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        {error_html}
        <section class="action-row">
          <article class="primary-action action-intake">
            <h2>美团日报交付版</h2>
            <p>由业务选择日报日期。系统先查统一基础数据层；缺数据时会同步插件下载目录并自动入库，仍缺数据则提示用美团插件导出对应日期。</p>
            <form method="post" action="/anta-reporting/meituan-daily/run" class="compact-form">
              <label>日报日期<input type="date" name="report_date" required></label>
              <button class="button" type="submit">生成美团日报</button>
            </form>
          </article>
          <article class="primary-action action-ledger">
            <h2>美团周报交付版</h2>
            <p>从统一基础数据层读取已入库美团数据，按周维度输出核心指标、TOP门店、TOP商品、流量表现、服务评价、下周选品建议和内容文案建议。</p>
            <form method="post" action="/anta-reporting/meituan-weekly/run">
              <button class="button accent" type="submit">生成美团周报</button>
            </form>
          </article>
          <article class="primary-action action-project">
            <h2>数据来源规则</h2>
            <p>插件和下载目录只作为原始文件入口；正式日报、周报必须从统一基础数据层取数，并生成基础数据来源追溯清单。</p>
            <a class="button secondary" href="/automation-runs">查看自动化数据执行</a>
          </article>
        </section>
        <section class="action-row">
          <article class="primary-action action-intake">
            <h2>周报初稿</h2>
            <p>读取安踏周报原始数据目录中的最新美团、京东数据，生成本周销售、销量、订单和TOP商品。</p>
            <form method="post" action="/anta-reporting/weekly/run">
              <button class="button" type="submit">生成周报初稿</button>
            </form>
          </article>
          <article class="primary-action action-ledger">
            <h2>月报初稿</h2>
            <p>读取安踏月报原始数据目录中的商品、门店、财务数据，生成月度经营指标和TOP商品。</p>
            <form method="post" action="/anta-reporting/monthly/run">
              <button class="button accent" type="submit">生成月报初稿</button>
            </form>
          </article>
          <article class="primary-action action-project">
            <h2>内容输出</h2>
            <p>AI选品和文案输出已接入，选择安踏儿童项目或上传品牌内容资料即可生成建议。</p>
            <div class="button-row">
              <a class="button secondary" href="/scenario/ai_selection">AI选品</a>
              <a class="button secondary" href="/scenario/copy_content">文案输出</a>
            </div>
          </article>
        </section>
        <section class="split">
          <article>
            <h2>周报取数目录</h2>
            <p class="note">{_e(str(weekly_dir))}</p>
          </article>
          <article>
            <h2>月报取数目录</h2>
            <p class="note">{_e(str(monthly_dir))}</p>
          </article>
        </section>
        """
        return self._page("安踏周报/月报", body)

    def _csv_table(self, headers: list[str], rows: list[dict[str, str]]) -> str:
        if not isinstance(headers, list):
            raise TypeError("headers must be list[str]")
        if not isinstance(rows, list):
            raise TypeError("rows must be list[dict[str, str]]")
        if not headers:
            return "<div class='empty-state'>暂无字段</div>"
        header_html = "".join(f"<th>{_e(header)}</th>" for header in headers)
        if not rows:
            row_html = f"<tr><td colspan='{len(headers)}'>暂无记录</td></tr>"
        else:
            row_html = "".join(
                "<tr>" + "".join(f"<td>{_e(row.get(header, ''))}</td>" for header in headers) + "</tr>"
                for row in rows[:500]
            )
            if len(rows) > 500:
                row_html += f"<tr><td colspan='{len(headers)}'>仅展示前 500 行，完整内容请下载 CSV。</td></tr>"
        return f"""
        <table class="data-table">
          <thead><tr>{header_html}</tr></thead>
          <tbody>{row_html}</tbody>
        </table>
        """

    def _archive_intake_result_page(self, user: UserRecord, result: ArchiveIntakeResult, title: str = "自动归档完成") -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(result, ArchiveIntakeResult):
            raise TypeError("result must be ArchiveIntakeResult")
        if not isinstance(title, str) or not title.strip():
            raise ValueError("title must not be empty")
        rows = "".join(
            f"""
            <tr>
              <td>{_e(outcome.file_name)}</td>
              <td>{_e(outcome.status)}</td>
              <td>{_e(outcome.priority)}</td>
              <td>{_e(outcome.project)}</td>
              <td>{_e(outcome.material_type)}</td>
              <td>{_e(outcome.archive_path)}</td>
              <td>{_e(outcome.reason)}</td>
            </tr>
            """
            for outcome in result.outcomes
        )
        if not rows:
            rows = "<tr><td colspan='7'>待处理文件夹暂无文件</td></tr>"
        body = f"""
        <section class="toolbar">
          <div>
            <h1>{_e(title)}</h1>
            <p>提交人：{_e(user.display_name)} · 已归档 {result.processed_count} 个，无法判定 {result.unresolved_count} 个，跳过 {result.skipped_count} 个</p>
          </div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        <section class="split">
          <article>
            <h2>登记文件</h2>
            <p class="note">资料索引：{_e(str(result.index_path))}<br>数据字典：{_e(str(result.dictionary_path))}</p>
          </article>
          <article>
            <h2>无法判定处理</h2>
            <p>无法判定的文件会放入 `00_intake/03_unresolved`，改名补充品牌、项目或素材类型后，可重新放回待处理文件夹扫描。</p>
          </article>
        </section>
        <section>
          <h2>处理明细</h2>
          <table>
            <thead><tr><th>文件</th><th>状态</th><th>优先级</th><th>项目</th><th>素材类型</th><th>归档位置</th><th>判定理由</th></tr></thead>
            <tbody>{rows}</tbody>
          </table>
        </section>
        """
        return self._page(title, body)

    def _priority_section(self, priority: str, title: str, description: str) -> str:
        for value_name, value in (("priority", priority), ("title", title), ("description", description)):
            if not isinstance(value, str) or not value.strip():
                raise ValueError(f"{value_name} must not be empty")
        scenario_keys = [key for key, scenario in self.scenarios.items() if scenario.priority == priority]
        cards = "".join(self._scenario_card(key) for key in scenario_keys)
        if not cards:
            cards = "<div class='empty-state'>暂无接入项目</div>"
        return f"""
        <section class="priority-section">
          <div class="priority-header">
            <div class="priority-mark">{_e(priority)}</div>
            <div>
              <h2>{_e(title)}</h2>
              <p>{_e(description)}</p>
            </div>
          </div>
          <div class="priority-grid">{cards}</div>
        </section>
        """

    def _scenario_card(self, scenario_key: str) -> str:
        scenario = self.scenarios[scenario_key]
        href = self._scenario_href(scenario.key)
        button_text = "打开项目" if scenario.key in (ANTA_RETAIL_KEY, anta_reporting.MODULE_KEY) else "进入处理"
        return f"""
        <article class="card">
          <div class="badge">{_e(scenario.priority)}</div>
          <h2>{_e(scenario.name)}</h2>
          <p>{_e(scenario.description)}</p>
          <dl><dt>品牌</dt><dd>{_e(scenario.brand)}</dd><dt>归类</dt><dd>{_e(scenario.business_type)}</dd></dl>
          <a class="button" href="{href}">{_e(button_text)}</a>
        </article>
        """

    def _scenario_href(self, scenario_key: str) -> str:
        if scenario_key == ANTA_RETAIL_KEY:
            return "/anta-retail"
        if scenario_key == anta_reporting.MODULE_KEY:
            return "/anta-reporting"
        return f"/scenario/{quote(scenario_key)}"

    def _anta_retail_page(self, user: UserRecord, error: str = "", anta_url: str = ANTA_RETAIL_DEFAULT_URL) -> str:
        if not isinstance(anta_url, str) or not anta_url.strip():
            raise ValueError("anta_url must not be empty")
        scenario = self.scenarios[ANTA_RETAIL_KEY]
        fields = "".join(f"<li>{_e(field)}</li>" for field in scenario.required_fields)
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="toolbar">
          <div>
            <h1>{_e(scenario.name)}</h1>
            <p>{_e(scenario.priority)} · {_e(scenario.business_type)} · 提交人：{_e(user.display_name)}</p>
          </div>
          <a class="button secondary" href="/">返回</a>
        </section>
        {error_html}
        <section class="split">
          <article>
            <h2>进入安踏网页</h2>
            <p>这个入口连接到已开发好的安踏即时零售网页，当前统一承接上下架筛选、素材筛选、黑名单筛选。</p>
            <a class="button" href="{_e(anta_url)}" target="_blank" rel="noreferrer">打开安踏即时零售网页</a>
            <p class="note">本机试用请运行“start_anta_retail_web.bat”。局域网共享请运行“start_anta_retail_lan.bat”，并保持主工作台和安踏网页都处于启动状态。</p>
          </article>
          <article>
            <h2>资料要求</h2>
            <ul>{fields}</ul>
            <p class="note">资料归档位置：{_e(str(scenario.template_path))}</p>
          </article>
        </section>
        <section class="grid">
          <article class="card">
            <div class="badge">1</div>
            <h2>上下架筛选</h2>
            <p>对比官网标准货盘、美团商品原表、京东商品原表，输出待补上架、平台非候选和异常明细。</p>
          </article>
          <article class="card">
            <div class="badge">2</div>
            <h2>素材筛选</h2>
            <p>按待上架清单核验商品素材，支持素材索引、素材目录和素材站检索，输出待补素材结果。</p>
          </article>
          <article class="card">
            <div class="badge">3</div>
            <h2>黑名单筛选</h2>
            <p>用黑名单明细匹配美团、京东商品和门店信息，生成新增、释放、导入模板和复盘汇总。</p>
          </article>
        </section>
        <section class="split">
          <article>
            <h2>黑名单上传处理</h2>
            <form method="post" action="/anta-retail/blacklist/run" enctype="multipart/form-data">
              <label>任务名称<input name="title" value="安踏即时零售黑名单筛选"></label>
              <label>黑名单明细<input type="file" name="blacklist_file" accept=".xlsx" required></label>
              <label>美团商品表<input type="file" name="meituan_product_file" accept=".xlsx" required></label>
              <label>京东商品表<input type="file" name="jd_product_file" accept=".xlsx,.xls,.html,.htm,.csv" required></label>
              <label>门店信息汇总<input type="file" name="store_summary_file" accept=".xlsx" required></label>
              <label>旧版美团黑名单（可选）<input type="file" name="old_meituan_blacklist_file" accept=".xlsx"></label>
              <button class="button" type="submit">开始黑名单筛选</button>
            </form>
          </article>
          <article>
            <h2>黑名单结果</h2>
            <p>处理完成后会生成一份 Excel 总表，并在首页“最近处理记录”中留痕，可重复下载。</p>
            <ul>
              <li>匹配总览</li>
              <li>美团商品命中明细、美团新增导入模板、美团释放模板</li>
              <li>京东商品命中明细、京东新增导入模板</li>
              <li>异常数据和未匹配款号</li>
            </ul>
          </article>
        </section>
        """
        return self._page(scenario.name, body)

    def _anta_retail_url(self, handler: BaseHTTPRequestHandler) -> str:
        host_header = handler.headers.get("Host", "").strip()
        if host_header == "":
            return ANTA_RETAIL_DEFAULT_URL
        host_name = host_header
        if host_name.startswith("[") and "]" in host_name:
            host_name = host_name.split("]", 1)[0] + "]"
        elif ":" in host_name:
            host_name = host_name.rsplit(":", 1)[0]
        if host_name in ("", "0.0.0.0"):
            host_name = "127.0.0.1"
        result = f"http://{host_name}:8766"
        assert result.startswith("http://")
        return result

    def _archive_index_path(self) -> Path:
        path = self.config.template_root / "00_index_dictionary" / "archive_index.csv"
        assert isinstance(path, Path)
        return path

    def _data_dictionary_path(self) -> Path:
        path = self.config.template_root / "00_index_dictionary" / "data_dictionary.csv"
        assert isinstance(path, Path)
        return path

    def _scenario_page(self, user: UserRecord, scenario_key: str, error: str) -> str:
        scenario = self.scenarios[scenario_key]
        fields = "".join(f"<li>{_e(field)}</li>" for field in scenario.required_fields)
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        template = _e(str(scenario.template_path))
        project_select = self._project_select(scenario_key)
        body = f"""
        <section class="toolbar">
          <div>
            <h1>{_e(scenario.name)}</h1>
            <p>{_e(scenario.brand)} · {_e(scenario.business_type)} · 提交人：{_e(user.display_name)}</p>
          </div>
          <a class="button secondary" href="/">返回</a>
        </section>
        {error_html}
        <section class="split">
          <article>
            <h2>上传处理</h2>
            <form method="post" action="/scenario/{quote(scenario.key)}/run" enctype="multipart/form-data">
              <label>任务名称<input name="title" value="{_e(scenario.name)}"></label>
              {project_select}
              <label>业务数据文件<input type="file" name="data_file" accept=".csv,.xlsx" required></label>
              <button class="button" type="submit">开始处理</button>
            </form>
          </article>
          <article>
            <h2>字段要求</h2>
            <ul>{fields}</ul>
            <p><a class="button secondary" href="/scenario/{quote(scenario.key)}/template">下载标准模板</a></p>
            <p class="note">标准模板位置：{template}</p>
          </article>
        </section>
        """
        return self._page(scenario.name, body)

    def _project_select(self, scenario_key: str) -> str:
        if scenario_key != "ai_selection":
            return ""
        return """
              <label>项目选择
                <select name="project">
                  <option value="通用选品">通用选品</option>
                  <option value="安踏儿童">安踏儿童</option>
                </select>
              </label>
        """

    def _result_page(self, user: UserRecord, job_id: int, result: ProcessingResult) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(job_id, int) or job_id <= 0:
            raise ValueError("job_id must be a positive integer")
        if not isinstance(result, ProcessingResult):
            raise TypeError("result must be ProcessingResult")
        summary_items = "".join(f"<li><strong>{_e(key)}</strong><span>{_e(value)}</span></li>" for key, value in result.summary.items())
        warnings = "".join(f"<li>{_e(item)}</li>" for item in result.warnings) or "<li>无</li>"
        result_preview = self._copy_content_result_preview(result)
        if result.module == P2_CONTENT_MODULE_KEY:
            download_label = "下载P2内容交付包 CSV"
        elif result.module == copy_content.MODULE_KEY:
            download_label = "下载含文案的结果 CSV"
        else:
            download_label = "下载结果 CSV"
        body = f"""
        <section class="toolbar">
          <div><h1>处理完成</h1><p>提交人：{_e(user.display_name)} · 任务编号：{job_id}</p></div>
          <a class="button secondary" href="/">返回首页</a>
        </section>
        <section class="split">
          <article>
            <h2>结果汇总</h2>
            <ul class="metrics">{summary_items}</ul>
            <a class="button" href="/jobs/{job_id}/download">{download_label}</a>
            {"<p class='note'>下载后的 CSV 中，E 列是 AI 标题，F 列是 AI 正文。</p>" if result.module == copy_content.MODULE_KEY else ""}
          </article>
          <article>
            <h2>复核提醒</h2>
            <ul>{warnings}</ul>
          </article>
        </section>
        {result_preview}
        """
        page = self._page("处理完成", body)
        assert page.strip()
        return page

    def _task_result_page(self, user: UserRecord, result: TaskResult) -> str:
        if not isinstance(user, UserRecord):
            raise TypeError("user must be UserRecord")
        if not isinstance(result, TaskResult):
            raise TypeError("result must be TaskResult")
        status_label = "\u6210\u529f" if result.status == WorkerTaskStatus.SUCCESS else "\u5931\u8d25"
        result_items = "".join(
            f"<li><strong>{_e(str(key))}</strong><span>{_e(str(value))}</span></li>"
            for key, value in result.result.items()
        )
        if not result_items:
            result_items = "<li><strong>status</strong><span>empty result</span></li>"
        error_block = f"<p class='error'>{_e(result.error)}</p>" if result.error.strip() else ""
        body = f"""
        <section class="toolbar">
          <div><h1>{status_label}</h1><p>{_e(user.display_name)} - task_id: {result.task_id}</p></div>
          <a class="button secondary" href="/anta-reporting">\u8fd4\u56de\u62a5\u8868\u9875</a>
        </section>
        <section class="split">
          <article>
            <h2>\u4efb\u52a1\u72b6\u6001</h2>
            <ul class="metrics">
              <li><strong>status</strong><span>{_e(result.status.value)}</span></li>
              <li><strong>finished_time</strong><span>{_e(result.finished_time)}</span></li>
            </ul>
            {error_block}
          </article>
          <article>
            <h2>\u4efb\u52a1\u7ed3\u679c</h2>
            <ul class="metrics">{result_items}</ul>
          </article>
        </section>
        """
        page = self._page("\u4efb\u52a1\u7ed3\u679c", body)
        assert page.strip()
        return page

    def _copy_content_result_preview(self, result: ProcessingResult) -> str:
        if not isinstance(result, ProcessingResult):
            raise TypeError("result must be ProcessingResult")
        if result.module == anta_meituan_reporting.MODULE_KEY and result.output_rows:
            return self._meituan_report_result_preview(result)
        if result.module == P2_CONTENT_MODULE_KEY and result.output_rows:
            return self._p2_content_result_preview(result)
        if result.module != copy_content.MODULE_KEY or not result.output_rows:
            return ""
        rows = []
        for row in result.output_rows[:20]:
            if not isinstance(row, dict):
                raise TypeError("each output row must be dict")
            rows.append(
                "<tr>"
                f"<td>{_e(row.get('商品名称', ''))}</td>"
                f"<td>{_e(row.get('AI标题建议', ''))}</td>"
                f"<td>{_e(row.get('AI正文建议', ''))}</td>"
                f"<td>{_e(row.get('合规状态', ''))}</td>"
                "</tr>"
            )
        preview = f"""
        <section>
          <div class="section-heading">
            <h2>文案生成结果</h2>
            <p>页面展示前 20 条，完整结果保存在下载文件中。</p>
          </div>
          <div class="table-panel">
            <table class="data-table copy-result-table">
              <thead><tr><th>商品名称</th><th>AI标题建议</th><th>AI正文建议</th><th>合规状态</th></tr></thead>
              <tbody>{''.join(rows)}</tbody>
            </table>
          </div>
        </section>
        """
        assert "AI正文建议" in preview
        return preview

    def _p2_content_result_preview(self, result: ProcessingResult) -> str:
        if not isinstance(result, ProcessingResult):
            raise TypeError("result must be ProcessingResult")
        rows = []
        for row in result.output_rows[:20]:
            if not isinstance(row, dict):
                raise TypeError("each output row must be dict")
            rows.append(
                "<tr>"
                f"<td>{_e(row.get('款号/SKU', ''))}</td>"
                f"<td>{_e(row.get('商品名称', ''))}</td>"
                f"<td>{_e(row.get('目标人群', ''))}</td>"
                f"<td>{_e(row.get('卖点提炼', ''))}</td>"
                f"<td>{_e(row.get('AI标题', ''))}</td>"
                f"<td>{_e(row.get('AI正文', ''))}</td>"
                f"<td>{_e(row.get('质检风险', ''))}</td>"
                "</tr>"
            )
        preview = f"""
        <section>
          <div class="section-heading">
            <h2>P2内容交付包预览</h2>
            <p>页面展示前 20 条，完整选品、文案、视觉Brief 和质检结果保存在下载文件中。</p>
          </div>
          <div class="table-panel">
            <table class="data-table p2-result-table">
              <thead><tr><th>款号/SKU</th><th>商品名称</th><th>目标人群</th><th>卖点提炼</th><th>AI标题</th><th>AI正文</th><th>质检风险</th></tr></thead>
              <tbody>{''.join(rows)}</tbody>
            </table>
          </div>
        </section>
        """
        assert "P2内容交付包预览" in preview
        return preview

    def _meituan_report_result_preview(self, result: ProcessingResult) -> str:
        if not isinstance(result, ProcessingResult):
            raise TypeError("result must be ProcessingResult")
        rows = []
        for row in result.output_rows[:40]:
            rows.append(
                "<tr>"
                f"<td>{_e(row.get('板块', ''))}</td>"
                f"<td>{_e(row.get('排序', ''))}</td>"
                f"<td>{_e(row.get('名称', ''))}</td>"
                f"<td>{_e(row.get('数值', ''))}</td>"
                f"<td>{_e(row.get('说明', ''))}</td>"
                "</tr>"
            )
        preview = f"""
        <section>
          <div class="section-heading">
            <h2>安踏美团报表预览</h2>
            <p>页面展示前 40 行，完整交付内容在结果 CSV 中。</p>
          </div>
          <div class="table-panel">
            <table class="data-table">
              <thead><tr><th>板块</th><th>排序</th><th>名称</th><th>数值</th><th>说明</th></tr></thead>
              <tbody>{''.join(rows)}</tbody>
            </table>
          </div>
        </section>
        """
        assert "安踏美团报表预览" in preview
        return preview

    def _job_row(self, job: JobRecord) -> str:
        return f"""
        <tr>
          <td>{job.id}</td>
          <td>{_e(job.title)}</td>
          <td>{_e(job.brand)}</td>
          <td>{_e(job.business_type)}</td>
          <td>{_e(job.created_by)}</td>
          <td>{_e(job.created_at)}</td>
          <td><a href="/jobs/{job.id}/download">下载</a></td>
        </tr>
        """

    def _login_page(self, error: str) -> str:
        error_html = f"<div class='error'>{_e(error)}</div>" if error else ""
        body = f"""
        <section class="login">
          <h1>内网自动化工作台</h1>
          <p>用于组内数据处理、配置自动化和结果留痕。</p>
          {error_html}
          <form method="post" action="/login">
            <label>账号<input name="username" autocomplete="username" required></label>
            <label>密码<input type="password" name="password" autocomplete="current-password" required></label>
            <button class="button" type="submit">登录</button>
          </form>
          <p class="note">试点账号：admin / admin123</p>
        </section>
        """
        return self._page("登录", body, is_login=True)

    def _page(self, title: str, body: str, is_login: bool = False) -> str:
        layout_class = "login-shell" if is_login else "app-shell"
        script = "" if is_login else self._page_script()
        return f"""<!doctype html>
        <html lang="zh-CN">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <title>{_e(title)}</title>
          <link rel="stylesheet" href="/static/style.css">
        </head>
        <body><main class="{layout_class}">{body}</main>{script}</body>
        </html>"""

    @staticmethod
    def _page_script() -> str:
        result = """
        <script>
        (() => {
          const parseDurationHours = (value) => {
            if (typeof value !== "string") return null;
            const text = value.trim();
            if (!text) return null;
            const match = text.match(/\\d+(?:\\.\\d+)?/);
            if (!match) return null;
            const amount = Number(match[0]);
            if (!Number.isFinite(amount)) return null;
            return text.includes("分钟") || text.includes("分") ? amount / 60 : amount;
          };
          const normalizeDurationHoursText = (value) => {
            if (typeof value !== "string") return "";
            const text = value.trim();
            if (!text) return "";
            const match = text.match(/\\d+(?:\\.\\d+)?/);
            if (!match) return text;
            const amount = Number(match[0]);
            if (!Number.isFinite(amount)) return text;
            const hours = text.includes("分钟") || text.includes("分") ? amount / 60 : amount;
            return `${Number.isInteger(hours) ? String(hours) : String(Number(hours.toFixed(2)))}小时`;
          };
          const formatTimeSaved = (originalValue, currentValue) => {
            const originalHours = parseDurationHours(originalValue);
            const currentHours = parseDurationHours(currentValue);
            if (originalHours === null || originalHours <= 0 || currentHours === null || currentHours < 0) return "待计算";
            const savedHours = originalHours - currentHours;
            if (savedHours < 0) return "未提效";
            return `${Number.isInteger(savedHours) ? String(savedHours) : String(Number(savedHours.toFixed(2)))}小时`;
          };
          const updateCard = (card) => {
            const originalInput = card.querySelector('input[name="original_manual_time"]');
            const currentInput = card.querySelector('input[name="current_processing_time"]');
            const output = card.querySelector("[data-time-saved-output]");
            if (!originalInput || !currentInput || !output) return;
            output.textContent = formatTimeSaved(originalInput.value, currentInput.value);
          };
          document.querySelectorAll(".completed-feedback-card").forEach((card) => {
            card.querySelectorAll(".efficiency-source-input").forEach((input) => {
              input.addEventListener("input", () => updateCard(card));
              input.addEventListener("change", () => {
                input.value = normalizeDurationHoursText(input.value);
                updateCard(card);
              });
              input.addEventListener("blur", () => {
                input.value = normalizeDurationHoursText(input.value);
                updateCard(card);
              });
              input.value = normalizeDurationHoursText(input.value);
            });
            updateCard(card);
          });
        })();
        </script>
        """
        assert "formatTimeSaved" in result
        return result

    def _send_html(self, handler: BaseHTTPRequestHandler, content: str, status: int = 200) -> None:
        data = content.encode("utf-8")
        handler.send_response(status)
        handler.send_header("Content-Type", "text/html; charset=utf-8")
        handler.send_header("Content-Length", str(len(data)))
        handler.end_headers()
        handler.wfile.write(data)

    def _send_json(self, handler: BaseHTTPRequestHandler, payload: dict[str, object], status: int = 200) -> None:
        if not isinstance(payload, dict):
            raise TypeError("payload must be dict")
        data = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        handler.send_response(status)
        handler.send_header("Content-Type", "application/json; charset=utf-8")
        handler.send_header("Content-Length", str(len(data)))
        handler.end_headers()
        handler.wfile.write(data)

    def _send_file(self, handler: BaseHTTPRequestHandler, path: Path, download_name: str | None = None) -> None:
        if not path.exists() or not path.is_file():
            self._send_html(handler, self._page("文件不存在", "<p>文件不存在。</p>"), status=404)
            return
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        handler.send_response(200)
        handler.send_header("Content-Type", content_type)
        handler.send_header("Content-Length", str(path.stat().st_size))
        if download_name:
            handler.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{quote(download_name)}")
        handler.end_headers()
        with path.open("rb") as handle:
            shutil.copyfileobj(handle, handler.wfile)

    def _redirect(
        self,
        handler: BaseHTTPRequestHandler,
        location: str,
        cookie_value: str | None = None,
        clear_cookie: bool = False,
    ) -> None:
        handler.send_response(302)
        handler.send_header("Location", location)
        if cookie_value:
            handler.send_header("Set-Cookie", f"intranet_session={cookie_value}; HttpOnly; SameSite=Lax; Path=/")
        if clear_cookie:
            handler.send_header("Set-Cookie", "intranet_session=; Max-Age=0; HttpOnly; SameSite=Lax; Path=/")
        handler.end_headers()


class _BytesReader:
    def __init__(self, content: bytes) -> None:
        if not isinstance(content, bytes) or not content:
            raise ValueError("content must be non-empty bytes")
        self._content = content

    def read(self) -> bytes:
        return self._content


def _e(value: object) -> str:
    return html.escape(str(value), quote=True)


def _safe_name(file_name: str) -> str:
    cleaned = "".join(char if char.isalnum() or char in "._-" else "_" for char in Path(file_name).name)
    return cleaned or "upload"


def _safe_serial() -> str:
    from datetime import datetime

    return datetime.now().strftime("%Y%m%d%H%M%S%f")


def _required_json_text(payload: dict[str, object], field_name: str) -> str:
    if not isinstance(payload, dict):
        raise TypeError("payload must be dict")
    value = payload.get(field_name)
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{field_name} must not be empty")
    result = value.strip()
    assert result
    return result


def _required_json_object(payload: dict[str, object], field_name: str) -> dict[str, object]:
    if not isinstance(payload, dict):
        raise TypeError("payload must be dict")
    value = payload.get(field_name)
    if not isinstance(value, dict):
        raise ValueError(f"{field_name} must be object")
    result = dict(value)
    assert isinstance(result, dict)
    return result


def _task_execution_steps(task_type: str) -> tuple[str, ...]:
    if not isinstance(task_type, str) or not task_type.strip():
        raise ValueError("task_type must not be empty")
    normalized = task_type.strip().upper()
    service = {
        "REPORT_GENERATE": "ReportService",
        "DATA_IMPORT": "DataFoundationService",
        "AI_CONTENT_GENERATE": "AIContentService",
    }.get(normalized, "Service")
    executor = {
        "REPORT_GENERATE": "ReportExecutor",
        "DATA_IMPORT": "DataImportExecutor",
        "AI_CONTENT_GENERATE": "AIContentExecutor",
    }.get(normalized, "Executor")
    result = ("TaskSubmitter", "TaskRunner", executor, service, "ResultAsset")
    assert len(result) == 5
    return result


def _task_download_status_text(status: str) -> str:
    if not isinstance(status, str):
        raise TypeError("status must be str")
    normalized = status.strip().lower()
    if normalized in {"pending", "running"}:
        return "任务完成后可下载"
    if normalized == "failed":
        return "任务失败，无结果文件"
    if normalized == "success":
        return "结果文件不存在"
    return "暂无可下载文件"

def _task_id_value(task: object) -> int:
    value = getattr(task, "task_id", 0)
    if not isinstance(value, int) or value <= 0:
        raise ValueError("task_id must be positive int")
    return value


def _task_text(task: object, field_name: str) -> str:
    if not isinstance(field_name, str) or not field_name.strip():
        raise ValueError("field_name must not be empty")
    value = getattr(task, field_name, "")
    if value is None:
        return ""
    return str(value).strip()


def _asset_rows_html(asset: dict[str, object]) -> str:
    if not isinstance(asset, dict):
        raise TypeError("asset must be dict")
    if not asset:
        return "<li><strong>asset</strong><span>-</span></li>"
    rows = "".join(
        f"<li><strong>{_e(key)}</strong><span>{_e(value)}</span></li>"
        for key, value in sorted(asset.items())
    )
    assert rows
    return rows


def _console_can_open_page(user: UserRecord, page: str) -> bool:
    if not isinstance(user, UserRecord):
        raise TypeError("user must be UserRecord")
    if not isinstance(page, str) or not page.strip():
        raise ValueError("page must not be empty")
    role = _console_role_key(user.role)
    if role == "viewer":
        return False
    if page == "environment":
        return role in {"admin", "developer"}
    return role in {"admin", "developer", "business_owner", "user"}


def _console_role_key(role: str) -> str:
    if not isinstance(role, str):
        raise TypeError("role must be str")
    text = role.strip().lower()
    if not text:
        raise ValueError("role must not be empty")
    if any(marker in text for marker in ("admin", "administrator", "???", "?????")):
        return "admin"
    if any(marker in text for marker in ("developer", "dev", "??", "???", "??")):
        return "developer"
    if any(marker in text for marker in ("business_owner", "business owner", "owner", "?????", "???")):
        return "business_owner"
    if any(marker in text for marker in ("viewer", "read_only", "readonly", "???", "??", "??")):
        return "viewer"
    return "user"


def _task_console_filters_from_query(path: str) -> TaskConsoleFilters:
    query = parse_qs(urlparse(path).query)
    result = TaskConsoleFilters(
        task_type=query.get("task_type", [""])[0].strip(),
        status=query.get("status", [""])[0].strip(),
        created_by=query.get("created_by", [""])[0].strip(),
        brand_id=query.get("brand_id", [""])[0].strip(),
        business_unit=query.get("business_unit", [""])[0].strip(),
        platform=query.get("platform", [""])[0].strip(),
        channel=query.get("channel", [""])[0].strip(),
    )
    assert isinstance(result, TaskConsoleFilters)
    return result

def _task_id_from_api_path(path: str, expected_parts: int) -> int:
    if not isinstance(path, str) or not path.strip():
        raise ValueError("path must not be empty")
    parts = [part for part in path.split("/") if part]
    if len(parts) != expected_parts or parts[0] != "api" or parts[1] != "tasks":
        raise ValueError("task api path is invalid")
    if expected_parts == 4 and parts[3] != "download":
        raise ValueError("task download path is invalid")
    try:
        task_id = int(parts[2])
    except ValueError as exc:
        raise ValueError("task_id must be integer") from exc
    if task_id <= 0:
        raise ValueError("task_id must be positive")
    return task_id


def _task_id_from_page_path(path: str) -> int:
    if not isinstance(path, str) or not path.strip():
        raise ValueError("path must not be empty")
    parts = [part for part in path.split("/") if part]
    if len(parts) != 2 or parts[0] != "tasks":
        raise ValueError("task page path is invalid")
    try:
        task_id = int(parts[1])
    except ValueError as exc:
        raise ValueError("task_id must be integer") from exc
    if task_id <= 0:
        raise ValueError("task_id must be positive")
    return task_id


def _task_result_summary_text(result: dict[str, object]) -> str:
    if not isinstance(result, dict):
        raise TypeError("result must be dict")
    if not result:
        return "empty"
    if "summary" in result and isinstance(result["summary"], dict):
        return f"summary: {len(result['summary'])}"
    if "output_row_count" in result:
        return f"rows: {result['output_row_count']}"
    return ", ".join(str(key) for key in list(result.keys())[:3])


def _flatten_result(result: dict[str, object]) -> dict[str, str]:
    if not isinstance(result, dict):
        raise TypeError("result must be dict")
    flattened: dict[str, str] = {}
    for key, value in result.items():
        if isinstance(value, dict):
            for child_key, child_value in value.items():
                flattened[f"{key}.{child_key}"] = str(child_value)
            continue
        flattened[str(key)] = str(value)
    return flattened


def _unique_upload_path(path: Path) -> Path:
    if not isinstance(path, Path):
        raise TypeError("path must be pathlib.Path")
    if not path.exists():
        return path
    for index in range(1, 10000):
        candidate = path.with_name(f"{path.stem}_{index:03d}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise ValueError(f"无法生成不重名文件路径：{path}")


def _read_csv_table(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not isinstance(path, Path):
        raise TypeError("path must be pathlib.Path")
    if not path.exists() or not path.is_file():
        raise ValueError(f"CSV 文件不存在：{path}")
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = [str(header).strip() for header in (reader.fieldnames or []) if str(header).strip()]
        rows = [{header: str(row.get(header, "") or "").strip() for header in headers} for row in reader]
    assert isinstance(headers, list)
    assert isinstance(rows, list)
    return headers, rows


def _csv_record_count(path: Path) -> int:
    if not isinstance(path, Path):
        raise TypeError("path must be pathlib.Path")
    if not path.exists() or not path.is_file() or path.stat().st_size == 0:
        return 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        count = sum(1 for _ in reader)
    if count < 0:
        raise AssertionError("csv record count must be non-negative")
    return count


def _format_decimal(value: Decimal) -> str:
    if not isinstance(value, Decimal):
        raise TypeError("value must be Decimal")
    normalized = value.quantize(Decimal("0.1")) if value != value.to_integral_value() else value.quantize(Decimal("1"))
    result = format(normalized, "f")
    if "." in result:
        result = result.rstrip("0").rstrip(".")
    assert result
    return result


def _parse_duration_hours(value: str) -> Decimal | None:
    if not isinstance(value, str):
        raise TypeError("value must be str")
    text = value.strip()
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    if match is None:
        return None
    try:
        amount = Decimal(match.group(1))
    except InvalidOperation:
        return None
    if "分钟" in text or "分" in text:
        return amount / Decimal("60")
    return amount


def _form_value(fields: dict[str, list[str]], name: str, default: str = "") -> str:
    if not isinstance(fields, dict):
        raise TypeError("fields must be dict")
    if not isinstance(name, str) or not name.strip():
        raise ValueError("name must not be empty")
    if not isinstance(default, str):
        raise TypeError("default must be str")
    values = fields.get(name, [default])
    if not isinstance(values, list) or not values:
        return default
    result = str(values[0]).strip()
    assert isinstance(result, str)
    return result


def _compact_form_date(value: str) -> str:
    if not isinstance(value, str):
        raise TypeError("value must be str")
    text = value.strip()
    if not text:
        raise ValidationError("请选择开始日期和结束日期。")
    compact = text.replace("-", "")
    if len(compact) != 8 or not compact.isdigit():
        raise ValidationError("日期必须是 YYYY-MM-DD 或 YYYYMMDD。")
    try:
        date(int(compact[:4]), int(compact[4:6]), int(compact[6:8]))
    except ValueError as exc:
        raise ValidationError("请选择有效日期。") from exc
    assert compact
    return compact


def _latest_matching_file(root_dir: Path, required_keywords: tuple[str, ...], suffixes: tuple[str, ...]) -> Path:
    if not isinstance(root_dir, Path):
        raise TypeError("root_dir must be pathlib.Path")
    if not isinstance(required_keywords, tuple) or not required_keywords:
        raise ValueError("required_keywords must be a non-empty tuple")
    if not isinstance(suffixes, tuple) or not suffixes:
        raise ValueError("suffixes must be a non-empty tuple")
    if not root_dir.exists():
        raise FileNotFoundError(f"取数目录不存在：{root_dir}")
    lowered_suffixes = tuple(suffix.lower() for suffix in suffixes)
    candidates = [
        path
        for path in root_dir.rglob("*")
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in lowered_suffixes
        and all(keyword in path.name for keyword in required_keywords)
    ]
    if not candidates:
        keyword_text = "、".join(required_keywords)
        raise FileNotFoundError(f"未在{root_dir}找到包含“{keyword_text}”的取数文件")
    result = max(candidates, key=lambda path: (path.stat().st_mtime, path.name))
    assert result.exists()
    return result


def _source_date_range(rows: list[dict[str, str]], date_field: str) -> tuple[str, str]:
    if not isinstance(rows, list) or not rows:
        raise ValidationError("rows must not be empty")
    if not isinstance(date_field, str) or not date_field.strip():
        raise ValueError("date_field must not be empty")
    dates: set[str] = set()
    for row in rows:
        dates.update(_compact_dates_from_source(row.get(date_field, "")))
        if date_field == "开始时间":
            dates.update(_compact_dates_from_source(row.get("结束时间", "")))
        if date_field == "评价提交日期":
            dates.update(_compact_dates_from_source(row.get("评价提交日期", "")))
    ordered_dates = sorted(dates)
    if not ordered_dates:
        raise ValidationError(f"源文件缺少可识别日期字段：{date_field}")
    result = (ordered_dates[0], ordered_dates[-1])
    assert result[0] <= result[1]
    return result


def _compact_date_from_source(value: object) -> str:
    dates = _compact_dates_from_source(value)
    return dates[0] if dates else ""


def _compact_dates_from_source(value: object) -> list[str]:
    text = "" if value is None else str(value).strip()
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 8:
        return []
    dates = [digits[index : index + 8] for index in range(0, len(digits) - 7, 8)]
    return [item for item in dates if len(item) == 8]


def create_intranet_app(config: AppConfig = DEFAULT_CONFIG, container: ApplicationContainer | None = None) -> IntranetApp:
    if not isinstance(config, AppConfig):
        raise TypeError("config must be AppConfig")
    actual_container = container if container is not None else build_application_container(environ=_container_environ_from_app_config(config))
    app = IntranetApp(config, actual_container)
    assert isinstance(app, IntranetApp)
    return app


def _container_environ_from_app_config(config: AppConfig) -> dict[str, str]:
    if not isinstance(config, AppConfig):
        raise TypeError("config must be AppConfig")
    runtime_dir = config.database_path.parent
    result = {
        "APP_ENV": "development",
        "DATABASE_BACKEND": "sqlite",
        "SQLITE_PATH": str(config.database_path),
        "RUNTIME_DIR": str(runtime_dir),
        "UPLOAD_DIR": str(config.upload_dir),
        "RESULT_DIR": str(config.result_dir),
        "LOG_DIR": str(runtime_dir / "logs"),
        "TEMPLATE_ROOT": str(config.template_root),
        "REPORT_TASK_MODE": _report_task_mode(),
    }
    assert result["DATABASE_BACKEND"] == "sqlite"
    return result


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    app = create_intranet_app(DEFAULT_CONFIG)
    app.initialize()
    server = ThreadingHTTPServer((DEFAULT_CONFIG.host, DEFAULT_CONFIG.port), app.make_handler())
    logging.info("server started: http://%s:%s", DEFAULT_CONFIG.host, DEFAULT_CONFIG.port)
    try:
        server.serve_forever()
    finally:
        app.close()


if __name__ == "__main__":
    main()






